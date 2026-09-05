import os
import re
import json
import base64
import traceback
from contextlib import contextmanager

from dotenv import load_dotenv
load_dotenv()
from typing import Optional, List, Union, Any
from datetime import date, datetime
from io import BytesIO

import pg8000.dbapi as pgdb

import hashlib
import secrets
import time

from fastapi import FastAPI, HTTPException, Request, Response, Query, UploadFile, File
from fastapi.responses import HTMLResponse, FileResponse, Response as FastResponse, JSONResponse, StreamingResponse
from pydantic import BaseModel

# ── Anthropic ────────────────────────────────────────────────────────────────
try:
    import anthropic as _anthropic
    ANTHROPIC_OK = True
except ImportError:
    ANTHROPIC_OK = False

# ── Excel ─────────────────────────────────────────────────────────────────────
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ── WeasyPrint (HTML → PDF, requiere GTK/Pango — funciona en Linux/Railway) ───
try:
    import weasyprint as _weasyprint
    WEASYPRINT_OK = True
except Exception:
    WEASYPRINT_OK = False

# ── xhtml2pdf (HTML → PDF, puro Python — funciona en Windows y Linux) ─────────
try:
    from xhtml2pdf import pisa as _pisa
    XHTML2PDF_OK = True
except ImportError:
    XHTML2PDF_OK = False

# ── PDF (ReportLab — fallback) ────────────────────────────────────────────────
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib.colors import HexColor, white, black
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        Image as RLImage, HRFlowable, KeepTogether,
    )
    REPORTLAB_OK = True
except ImportError:
    REPORTLAB_OK = False

# ── DB config ─────────────────────────────────────────────────────────────────
# .strip() defensivo: si al pegar la clave en el entorno quedó un espacio o
# salto de línea invisible, la API la rechazaría; lo quitamos por seguridad.
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "").strip()
# Espacio de trabajo (workspace) de Anthropic. Las API keys "vinculadas a un
# usuario" (identity-linked) EXIGEN enviar este ID en cada solicitud; sin el,
# la API responde 400 y el chat de IA muestra "Connection error".
ANTHROPIC_WORKSPACE_ID = os.environ.get("ANTHROPIC_WORKSPACE_ID", "wrkspc_01GYH8XvAbtR3Y2CrjtBkchQ").strip()


def _anthropic_client():
    """Crea el cliente de Anthropic, agregando el header del workspace si está definido."""
    _kwargs = {"api_key": ANTHROPIC_API_KEY}
    if ANTHROPIC_WORKSPACE_ID:
        _kwargs["default_headers"] = {"anthropic-workspace-id": ANTHROPIC_WORKSPACE_ID}
    return _anthropic.Anthropic(**_kwargs)


DB_HOST     = os.environ.get("DB_HOST",     "localhost")
DB_PORT     = int(os.environ.get("DB_PORT", "5432"))
DB_NAME     = os.environ.get("DB_NAME",     "boom_ofertas")
DB_USER     = os.environ.get("DB_USER",     "postgres")
DB_PASSWORD = os.environ.get("DB_PASSWORD", "natalia2281*")

# ── SMTP config ───────────────────────────────────────────────────────────────
SMTP_HOST     = os.environ.get("SMTP_HOST",     "smtp.office365.com")
SMTP_PORT     = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USER     = os.environ.get("SMTP_USER",     "")
SMTP_PASSWORD = os.environ.get("SMTP_PASSWORD", "")
SMTP_FROM     = os.environ.get("SMTP_FROM",     "nvargas@boomlts.com")

NOTIF_TO = [
    "cnavarro@boomlts.com", "ingenieria@boomlts.com", "jsarmiento@boomlts.com",
    "hseboom@boomlts.com",  "rromerro@boomlts.com",   "analistadocumental@boomlts.com",
    "analistatracking@boomlts.com", "proyectos@boomlts.com",
]
NOTIF_CC = [
    "mjamis@boomlts.com", "operaciones@boomlts.com",
    "bborrego@boomlts.com", "comercial@boomlts.com",
]

print(f"[DB] host={DB_HOST} port={DB_PORT} db={DB_NAME} user={DB_USER}")

# ── BOOM comercial info ───────────────────────────────────────────────────────
COMERCIAL_INFO = {
    # Nombres legales completos (firma de la oferta según el ejecutivo)
    "NATALIA ANDREA VARGAS LEAL":   ("Ejecutiva Comercial",   "nvargas@boomlts.com"),
    "BORIS ANDRES BORREGO JIMENEZ": ("Gerente General",        "bborrego@boomlts.com"),
    "WILLINGTON ORTIZ PLATA":       ("Director Comercial",     "comercial@boomlts.com"),
    # Compatibilidad con nombres cortos anteriores
    "Natalia Vargas":   ("Ejecutiva Comercial",   "nvargas@boomlts.com"),
    "Boris Borrego":    ("Gerente General",        "bborrego@boomlts.com"),
    "Willington Ortiz": ("Director Comercial",    "comercial@boomlts.com"),
}


def _logo_src() -> str:
    try:
        logo_path = os.path.join(os.path.dirname(__file__), "templates", "boom_logo.b64")
        with open(logo_path, "r") as f:
            return f.read().strip()
    except Exception:
        return ""


def _sello_src() -> str:
    """Sello 'Construimos País' (PNG transparente, líneas blancas)."""
    try:
        sello_path = os.path.join(os.path.dirname(__file__), "templates", "sello_pais.b64")
        with open(sello_path, "r") as f:
            return f.read().strip()
    except Exception:
        return ""


def _fmt_cop(v) -> str:
    try:
        return "${:,}".format(int(v)).replace(",", ".")
    except Exception:
        return "$0"


def _fmt_ref(num) -> str:
    n = str(num).zfill(6)
    return f"{n[:2]}-{n[2:]}"


# ── Auth helpers ──────────────────────────────────────────────────────────────
def _hash_pw(password: str) -> str:
    salt = secrets.token_hex(16)
    key = hashlib.pbkdf2_hmac("sha256", password.encode(), salt.encode(), 200_000)
    return f"{salt}:{key.hex()}"


def _verify_pw(password: str, stored: str) -> bool:
    try:
        salt, key_hex = stored.split(":", 1)
        key = hashlib.pbkdf2_hmac("sha256", password.encode(), salt.encode(), 200_000)
        return secrets.compare_digest(key.hex(), key_hex)
    except Exception:
        return False


def _logo_for_pdf(height_cm: float = 1.1, max_width_cm: float = 3.5):
    """Return a reportlab Image for the BOOM logo sized to fit the header cell."""
    if not REPORTLAB_OK:
        return None
    try:
        from reportlab.lib.utils import ImageReader
        logo_path = os.path.join(os.path.dirname(__file__), "templates", "boom_logo.b64")
        with open(logo_path, "r") as f:
            b64_str = f.read().strip()
        if "," in b64_str:
            b64_str = b64_str.split(",", 1)[1]
        img_bytes = base64.b64decode(b64_str)
        iw, ih = ImageReader(BytesIO(img_bytes)).getSize()
        aspect = iw / ih if ih else 1
        target_h = height_cm * cm
        target_w = min(target_h * aspect, max_width_cm * cm)
        target_h = target_w / aspect
        return RLImage(BytesIO(img_bytes), width=target_w, height=target_h)
    except Exception as exc:
        print(f"[PDF] logo error: {exc}")
        return None


def _sello_for_pdf(height_cm: float = 1.5, max_width_cm: float = 2.6):
    """Return a reportlab Image for the 'Construimos País' seal sized for the header."""
    if not REPORTLAB_OK:
        return None
    try:
        from reportlab.lib.utils import ImageReader
        sello_path = os.path.join(os.path.dirname(__file__), "templates", "sello_pais.b64")
        with open(sello_path, "r") as f:
            b64_str = f.read().strip()
        if "," in b64_str:
            b64_str = b64_str.split(",", 1)[1]
        img_bytes = base64.b64decode(b64_str)
        iw, ih = ImageReader(BytesIO(img_bytes)).getSize()
        aspect = iw / ih if ih else 1
        target_h = height_cm * cm
        target_w = min(target_h * aspect, max_width_cm * cm)
        target_h = target_w / aspect
        return RLImage(BytesIO(img_bytes), width=target_w, height=target_h)
    except Exception as exc:
        print(f"[PDF] sello error: {exc}")
        return None


def _parsear_detalle(texto: str) -> list:
    """Parse a pricing block into a list of equipment dicts.

    Expected format (one block per empty line):
        Nombre del equipo
        • Descripción / configuración
        • $XX.XXX.XXX,oo
    """
    items: list = []
    current: dict | None = None

    for raw in texto.strip().split("\n"):
        line = raw.strip()

        if not line:
            if current is not None:
                items.append(current)
                current = None
            continue

        is_bullet = bool(re.match(r"^[•\-\*•–]", line))
        clean = re.sub(r"^[•\-\*•–]\s*", "", line).strip()

        if not is_bullet:
            # New equipment – save previous block first
            if current is not None:
                items.append(current)
            # Try to extract leading quantity:  "5 Contenedor 40HC"  or  "Contenedor x5"
            qty = 1
            name = line
            m_qty = re.match(r"^(\d+)\s+(.+)$", line)
            if m_qty:
                qty = int(m_qty.group(1))
                name = m_qty.group(2)
            else:
                m_qty2 = re.search(r"\s+[xX]\s*(\d+)$", line)
                if m_qty2:
                    qty = int(m_qty2.group(1))
                    name = line[: m_qty2.start()]
            current = {"equipo": name.strip(), "dimensiones": "", "cant": qty,
                       "config": "", "valor_unit": 0}
        else:
            if current is None:
                current = {"equipo": "", "dimensiones": "", "cant": 1,
                           "config": "", "valor_unit": 0}
            price_m = re.search(r"\$\s*([\d.,]+)", clean)
            if price_m:
                p = price_m.group(1)
                p = re.sub(r",?[oO]{2}$", "", p)   # strip trailing ,oo
                p = p.replace(".", "").replace(",", "")
                try:
                    current["valor_unit"] = int(p)
                except Exception:
                    pass
            else:
                current["config"] = (current["config"] + "\n" + clean).strip()

    if current is not None:
        items.append(current)

    return [i for i in items if i.get("equipo") or i.get("config")]


def _auto_notas(equipos: list, texto_cliente: str = "",
                origen: str = "", destino: str = "") -> str:
    """Apply BOOM Logistics business rules to generate standard technical notes."""
    all_cfg = " ".join(
        (eq.get("config", "") or "") + " " + (eq.get("equipo", "") or "")
        for eq in equipos
    ).lower()
    all_text = all_cfg + " " + texto_cliente.lower()

    # ── Detect vehicle type for stand-by rate ─────────────────────────────
    is_cama5 = any(w in all_text for w in ["5 ejes", "5ejes", "cama baja 5", "cb5"])
    is_cama4 = any(w in all_text for w in ["4 ejes", "4ejes", "cama baja 4", "cb4"])
    is_cama3 = any(w in all_text for w in [
        "3 ejes", "3ejes", "cama baja 3", "cb3", "camabaja", "cama baja",
    ])
    is_cama_alta = any(w in all_text for w in [
        "cama alta", "camalta", "cama-alta", "extensible",
    ])

    if is_cama5:
        standby_valor = "$2.600.000"
        standby_tipo  = "Cama Baja 5 Ejes"
    elif is_cama4:
        standby_valor = "$1.800.000"
        standby_tipo  = "Cama Baja 4 Ejes"
    elif is_cama3:
        standby_valor = "$1.500.000"
        standby_tipo  = "Cama Baja 3 Ejes"
    else:
        standby_valor = "$1.200.000"
        standby_tipo  = "Cama Alta Extensible"

    # ── Detect extra/special service ──────────────────────────────────────
    is_izaje = any(w in all_text for w in [
        "izaje", "grúa", "grua", "modular en operación",
        "modular en operacion", "izar", "montaje de", "pluma",
    ])
    is_extra = any(w in all_text for w in [
        "extradi", "extrapesa", "escolta", "tecnólog", "tecnolog",
        "2 esc", "1 esc", "oversize",
    ])
    is_modular = any(w in all_text for w in ["modular", "spmt", "self-propelled"])

    for m_dim in re.finditer(r'(\d+(?:[.,]\d+)?)\s*m(?:\b|\s)', all_text):
        try:
            if float(m_dim.group(1).replace(",", ".")) > 3.0:
                is_extra = True
                break
        except Exception:
            pass

    tiempos_libres = (
        "12 horas para cargue / 12 horas para descargue" if is_modular
        else "6 horas para cargue / 6 horas para descargue"
    )

    notas = []

    # ── Fijos siempre presentes ────────────────────────────────────────────
    notas.append(f"Origen: {origen}" if origen else "Origen: ")
    notas.append(f"Destino: {destino}" if destino else "Destino: ")
    notas.append("Esquema de seguridad: ")

    if is_izaje:
        notas.append(
            f"Stand-by {standby_tipo}: {standby_valor} COP/día por unidad. "
            "Stand-by de grúa/equipo especializado: valor proporcional a la hora según equipo asignado."
        )
        notas.append(
            "Inicio de operación: Se requiere visita técnica previa al inicio "
            "de labores de izaje/montaje. El tiempo de alistamiento está incluido."
        )
    else:
        notas.append(f"Stand-by {standby_tipo}: {standby_valor} COP/día por unidad en espera.")

    if is_extra:
        notas.append(
            "Esquema extradimensionado/extrapesado: se incluyen 2 escoltas + 2 tecnólogos por despacho."
        )

    notas.append(f"Tiempos libres: {tiempos_libres}.")

    notas.append(
        "Devolución: Las tarifas incluyen el retorno (viaje vacío) del equipo "
        "hasta el punto de origen o patio BOOM."
    )
    notas.append(
        "Permisos: Incluye gestión de permisos de tránsito ante autoridades "
        "competentes (según aplique y reglamentación vigente)."
    )


    return "\n".join(notas)


# ── HTML offer generation ─────────────────────────────────────────────────────

_STANDBY_RATES = [
    (["modular 6","modular6","6 cuna 6","6cuna6"],                        "$8.500.000/día", "12 horas", "12 horas"),
    (["modular 5","modular5","5 cuna 5","5cuna5",
      "modular 4","modular4","4 cuna 4","4cuna4"],                        "$8.500.000/día", "12 horas", "12 horas"),
    (["semi modular","semimodular","2v4","modular"],                       "$2.800.000/día", "12 horas", "12 horas"),
    (["cama baja 5","camabaja5","cb5","5 ejes","60 ton","60ton"],          "$2.600.000/día", "6 horas",  "6 horas"),
    (["cama baja 4","camabaja4","cb4","4 ejes","45 ton","45ton"],          "$1.800.000/día", "8 horas",  "8 horas"),
    (["cama baja 3","camabaja3","cb3","3 ejes","30 ton","30ton",
      "cama baja","camabaja","cama plana"],                                "$1.200.000/día", "6 horas",  "6 horas"),
    (["camión turbo","camion turbo","turbo","sencillo","camioneta"],       "$550.000/día",   "6 horas",  "6 horas"),
    (["cama alta","camalta","patineta","extensible"],                      "$1.200.000/día", "6 horas",  "6 horas"),
]

def _standby_for_equipo(eq_name: str, eq_config: str) -> tuple:
    txt = (eq_name + " " + eq_config).lower()
    for keywords, rate, libre_c, libre_d in _STANDBY_RATES:
        if any(k in txt for k in keywords):
            return rate, libre_c, libre_d
    return "$1.200.000/día", "6 horas", "6 horas"


def _personal_from_config(config: str) -> str:
    c = config.lower()
    esc = re.search(r"(\d+)\s*escolt", c)
    tec = re.search(r"(\d+)\s*tecn", c)
    parts = []
    if esc:
        parts.append(f"{esc.group(1)} Esc")
    if tec:
        parts.append(f"{tec.group(1)} Tec")
    return " &bull; ".join(parts) if parts else "&mdash;"


def generar_html_oferta(data: dict) -> str:
    ref_fmt       = _fmt_ref(data.get("ref", "260001"))
    cliente       = data.get("cliente", "")
    contacto      = data.get("contacto", "") or ""
    ref_cliente   = data.get("ref_cliente", "") or ""
    cliente_final = data.get("cliente_final", "") or ""
    origen        = data.get("origen", "")
    destino       = data.get("destino", "")
    mes_anio      = data.get("mes_anio", datetime.now().strftime("%b %Y").upper())
    descripcion   = data.get("descripcion", "")
    comercial     = data.get("comercial", "Natalia Vargas")
    cargo_com, email_com = COMERCIAL_INFO.get(comercial, ("Ejecutiva Comercial", "nvargas@boomlts.com"))

    equipos      = data.get("equipos", []) or []
    cargo_items  = data.get("cargo_items", []) or []
    notas_raw    = data.get("notas", "")
    forma_pago   = data.get("forma_pago", "50% anticipo / 50% a 30 días tras radicación de factura")
    vigencia     = data.get("vigencia", 30)
    poliza_carga = data.get("poliza_carga", "Hasta $4.000.000.000 COP")
    poliza_rc    = data.get("poliza_rc",    "Hasta $4.000.000.000 COP")
    excl_raw     = data.get("exclusiones",
        "Permisos de tránsito, operación y pólizas asociadas (a cargo del cliente)\n"
        "Servicios, recursos o actividades no descritos explícitamente en esta oferta")

    logo_src = _logo_src()
    logo_html = (f'<img src="{logo_src}" alt="BOOM Logistics" style="height:42px;width:auto;">'
                 if logo_src else
                 '<span style="color:white;font-weight:bold;font-size:18px;letter-spacing:1px;">BOOM</span>')

    sello_src = _sello_src()
    sello_html = (f'<img src="{sello_src}" alt="Construimos Pais - Boom Logistics" style="height:90px;width:auto;">'
                  if sello_src else "")

    # ── Ref-bar components ────────────────────────────────────────────────────
    ruta = ""
    if origen and destino:
        ruta = f" &nbsp;|&nbsp; {origen.upper()} &#8594; {destino.upper()}"
    elif origen:
        ruta = f" &nbsp;|&nbsp; {origen.upper()}"

    ref_extra_html = ""
    if ref_cliente:
        ref_extra_html = f" &nbsp;|&nbsp; {ref_cliente}"
        if cliente_final:
            ref_extra_html += f" / {cliente_final}"

    desc_bar = f" &nbsp;&mdash;&nbsp; {descripcion}" if descripcion else ""

    # ── Greeting ──────────────────────────────────────────────────────────────
    saludo_nombre = contacto if contacto else f"equipo {cliente}"
    intro_txt = descripcion if descripcion else "la prestación de servicios de logística especializada"

    # ── Detect service flavors ────────────────────────────────────────────────
    all_eq_txt = " ".join(
        (e.get("equipo","") + " " + e.get("config","")).lower() for e in equipos
    ) + " " + notas_raw.lower()
    has_izaje = any(w in all_eq_txt for w in
                    ["izaje","grúa","grua","modular","patineta","izar","montaje"])

    # ── Cargo technical section ────────────────────────────────────────────────
    cargo_section_html = ""
    valid_cargo = [c for c in cargo_items if c.get("descripcion")]
    if valid_cargo:
        show_dims = any(c.get("dimensiones") for c in valid_cargo)
        show_peso = any(c.get("peso")        for c in valid_cargo)
        show_vol  = any(c.get("volumen")     for c in valid_cargo)

        # Spec-grid from first cargo item that has dimensional data
        spec_grid_html = ""
        fc = next((c for c in valid_cargo if c.get("dimensiones") or c.get("peso")), None)
        if fc:
            spec_items = []
            dims_raw = fc.get("dimensiones", "")
            if dims_raw:
                parts = re.split(r'\s*[×xX*]\s*', dims_raw)
                if len(parts) >= 3:
                    spec_items.append(f'<div class="spec-card"><div class="val">{parts[0].strip()}</div><div class="lbl">Largo c/u</div></div>')
                    spec_items.append(f'<div class="spec-card"><div class="val">{parts[1].strip()}</div><div class="lbl">Ancho c/u</div></div>')
                    spec_items.append(f'<div class="spec-card"><div class="val">{parts[2].strip()}</div><div class="lbl">Alto c/u</div></div>')
            if fc.get("peso"):
                spec_items.append(f'<div class="spec-card"><div class="val">{fc["peso"]}</div><div class="lbl">Peso c/u</div></div>')
            elif fc.get("volumen"):
                spec_items.append(f'<div class="spec-card"><div class="val">{fc["volumen"]}</div><div class="lbl">Volumen</div></div>')
            if spec_items:
                spec_grid_html = '<div class="spec-grid">' + "".join(spec_items) + '</div>'

        # Cargo table headers
        cargo_hdrs = '<th style="text-align:left;">Commodity</th><th>Cant.</th>'
        if show_dims: cargo_hdrs += '<th>Dimensiones</th>'
        if show_peso: cargo_hdrs += '<th>Peso c/u</th>'
        if show_vol:  cargo_hdrs += '<th style="text-align:right;">Volumen</th>'

        # Cargo table rows
        cargo_rows_html = ""
        for ci in valid_cargo:
            desc   = ci.get("descripcion","")
            tipo   = ci.get("tipo","")
            cant_c = ci.get("cant", 1)
            dims   = ci.get("dimensiones","")
            peso   = ci.get("peso","")
            vol    = ci.get("volumen","")
            orig_d = ci.get("origen_detalle","")
            dest_d = ci.get("destino_detalle","")
            sub_sp = []
            if tipo: sub_sp.append(tipo)
            if orig_d or dest_d:
                sub_sp.append((orig_d or "") + (" &#8594; " if orig_d and dest_d else "") + (dest_d or ""))
            sub_html = (f'<br><span style="font-size:10px;color:#666;">{" &bull; ".join(sub_sp)}</span>'
                        if sub_sp else "")
            cargo_rows_html += f"""
      <tr>
        <td style="text-align:left;"><strong>{desc}</strong>{sub_html}</td>
        <td>{cant_c}</td>
        {'<td>' + (dims or '&mdash;') + '</td>' if show_dims else ''}
        {'<td>' + (peso or '&mdash;') + '</td>' if show_peso else ''}
        {'<td style="text-align:right;">' + (vol or '&mdash;') + '</td>' if show_vol else ''}
      </tr>"""

        cargo_section_html = f"""
  <div class="section-title">1. DETALLE T&Eacute;CNICO DE LA CARGA</div>
  {spec_grid_html}
  <div class="table-scroll">
  <table class="det">
    <thead><tr>{cargo_hdrs}</tr></thead>
    <tbody>{cargo_rows_html}
    </tbody>
  </table>
  </div>"""

    # ── Fotos de referencia de la carga (embebidas en Sección 1) ──────────────
    fotos = data.get("fotos", []) or []
    valid_fotos = [f for f in fotos if isinstance(f, str) and f.strip().startswith("data:image")]
    if valid_fotos:
        # Fotos GRANDES y AMPLIAS: a todo el ancho, sin recortar (aspecto natural)
        fotos_imgs = "".join(
            f'<img src="{src}" alt="Foto de referencia de la carga" '
            'style="width:100%;height:auto;display:block;margin:12px auto;'
            'border-radius:6px;border:1px solid #e0e0e0;">'
            for src in valid_fotos
        )
        fotos_grid = ('<div style="margin-top:10px;">' + fotos_imgs + '</div>')
        if cargo_section_html:
            cargo_section_html += (
                '<p style="font-size:12px;font-weight:bold;color:#1B2A4A;'
                'margin:16px 0 6px 0;">Fotos de referencia de la carga</p>' + fotos_grid)
        else:
            cargo_section_html = ('<div class="section-title">1. FOTOS DE REFERENCIA '
                                  'DE LA CARGA</div>' + fotos_grid)

    # ── Dynamic section numbering ─────────────────────────────────────────────
    sec = 2 if cargo_section_html else 1

    # ── Equipment rows: Equipo | Cant | Tarifa c/u | Total ────────────────────
    total = 0
    eq_rows_html      = ""
    summary_rows_html = ""
    standby_rows_html = ""
    n_valued = 0

    for e in equipos:
        cant   = int(e.get("cant", 1) or 1)
        v_unit = int(e.get("valor_unit", 0) or 0)
        sub    = cant * v_unit
        total += sub
        eq_name = e.get("equipo", "")
        config  = (e.get("config") or "")

        # Config sub-line (strip newlines, keep as small bullets)
        config_clean = config.strip().replace("\n", " &bull; ")
        tipo_tag = (f'<br><span style="font-size:10px;color:#666;">{config_clean}</span>'
                    if config_clean else "")

        # ITR inclusion box
        is_itr = "itr" in eq_name.lower()

        unit_str  = _fmt_cop(v_unit) if v_unit else "&mdash;"
        total_str = _fmt_cop(sub)    if sub    else "&mdash;"

        eq_rows_html += f"""
      <tr>
        <td style="text-align:left;"><strong>{eq_name}</strong>{tipo_tag}</td>
        <td>{cant}</td>
        <td>{unit_str}</td>
        <td>{total_str}</td>
      </tr>"""

        if v_unit:
            n_valued += 1
            label = f"{eq_name} &mdash; {cant} &times; {unit_str}" if cant > 1 else f"{eq_name} &mdash; {unit_str}"
            summary_rows_html += f"""
      <tr>
        <td style="text-align:left;font-size:12px;"><strong>{label}</strong></td>
        <td style="text-align:right;">{total_str}</td>
      </tr>"""

        # Stand-by
        sb_rate, sb_libre_c, sb_libre_d = _standby_for_equipo(eq_name, config)
        standby_rows_html += f"""
      <tr>
        <td style="text-align:left;"><strong>{eq_name}</strong></td>
        <td>{sb_libre_c}</td>
        <td>{sb_libre_d}</td>
        <td>{sb_rate}</td>
      </tr>"""

    ruta_fase = (f" &mdash; {origen.upper()} &#8594; {destino.upper()}"
                 if origen and destino else
                 (f" &mdash; {origen.upper()}" if origen else ""))

    # Resumen económico: solo desglosar cuando hay MÁS de una línea con valor.
    # Con una sola línea, el valor ya aparece en la tabla de equipos → mostramos
    # únicamente la barra TOTAL OFERTA (sin repetir la línea).
    summary_html = ""
    if total:
        desglose = summary_rows_html if n_valued > 1 else ""
        summary_html = f"""
  <div class="table-scroll" style="margin-top:12px;">
  <table class="det">
    <tbody>
      {desglose}
      <tr class="total-row">
        <td><strong>TOTAL OFERTA</strong></td>
        <td><strong>{_fmt_cop(total)} COP</strong></td>
      </tr>
    </tbody>
  </table>
  </div>"""

    # ── Notes → clean bullet list ─────────────────────────────────────────────
    skip_pfx = ("origen:", "destino:", "stand-by", "standby", "tiempos libres")
    note_items = [
        ln for ln in (n.strip() for n in notas_raw.split("\n") if n.strip())
        if not any(ln.lower().startswith(p) for p in skip_pfx)
    ]
    notes_section_html = ""
    if note_items:
        notes_li = "".join(f"    <li>{n}</li>\n" for n in note_items)
        notes_section_html = f"""
  <div class="section-title">{sec + 1}. NOTAS T&Eacute;CNICAS DE OPERACI&Oacute;N</div>
  <ul class="notas">
{notes_li}  </ul>"""

    standby_note = "* Las horas adicionales ser&aacute;n cobradas proporcionalmente seg&uacute;n tarifa establecida."
    if has_izaje:
        standby_note += " El cobro del equipo de izaje inicia desde la llegada al sitio designado."

    # ── Exclusiones ───────────────────────────────────────────────────────────
    excl_items = [e.strip() for e in excl_raw.split("\n") if e.strip()]
    excl_html  = "".join(f"    <li>{e}</li>\n" for e in excl_items)

    cond_num = sec + (2 if note_items else 1)
    excl_num = cond_num + 1

    _html_oferta = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Oferta {ref_fmt} | {cliente}</title>
<style>
*{{box-sizing:border-box;}}
body{{font-family:Arial,sans-serif;font-size:13px;color:#1B2A4A;margin:0;padding:16px;background:#f4f4f4;}}
.wrapper{{max-width:860px;margin:0 auto;background:#fff;border-radius:6px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,.1);}}
.header-bar{{background:#1B2A4A;padding:14px 20px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:8px;}}
.header-bar img{{height:42px;width:auto;}}
.header-info{{color:#fff;font-size:11px;line-height:1.6;text-align:right;}}
.header-info strong{{font-size:12px;display:block;}}
.ref-bar{{background:#E8601C;padding:8px 20px;}}
.ref-bar p{{margin:0;color:#fff;font-size:11px;font-weight:bold;}}
.body{{padding:20px;}}
.greeting p{{font-size:13px;line-height:1.6;margin:0 0 10px 0;}}
.section-title{{background:#1B2A4A;color:#fff;font-size:12px;font-weight:bold;padding:7px 12px;margin:20px 0 8px 0;border-radius:3px;}}
.fase-title{{background:#2d4a7a;color:#fff;font-size:11px;font-weight:bold;padding:5px 12px;margin:10px 0 4px 0;border-radius:3px;}}
.spec-grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin:10px 0;}}
.spec-card{{background:#f0f4fa;border-radius:5px;padding:9px 8px;text-align:center;}}
.spec-card .val{{font-size:14px;font-weight:bold;color:#E8601C;}}
.spec-card .lbl{{font-size:10px;color:#666;margin-top:2px;}}
.incl-box{{background:#e8f5e9;border-left:4px solid #43a047;padding:8px 14px;font-size:12px;line-height:1.7;margin:6px 0 10px 0;}}
.note-box{{background:#fff8e1;border-left:4px solid #f9a825;padding:8px 14px;font-size:12px;line-height:1.7;margin:8px 0;}}
.table-scroll{{width:100%;overflow-x:auto;}}
table.det{{width:100%;border-collapse:collapse;font-size:12px;}}
table.det th{{background:#1B2A4A;color:#fff;padding:7px 8px;font-size:11px;text-align:center;white-space:nowrap;}}
table.det th:first-child{{text-align:left;}}
table.det td{{padding:7px 8px;border-bottom:1px solid #e5e5e5;vertical-align:middle;font-size:12px;}}
table.det td:not(:first-child){{text-align:center;}}
table.det td:last-child{{text-align:right;font-weight:bold;}}
table.det tr:nth-child(even) td{{background:#f9f9f9;}}
.subtotal-row td{{background:#e0e8f5!important;font-weight:bold;padding:7px 8px;border:none!important;color:#1B2A4A;}}
.subtotal-row td:last-child{{text-align:right;}}
.total-row td{{background:#E8601C!important;color:#fff;padding:9px 10px;font-weight:bold;font-size:13px;border:none!important;}}
.total-row td:last-child{{text-align:right;}}
ul.notas{{margin:0;padding-left:18px;}}
ul.notas li{{margin-bottom:7px;font-size:13px;line-height:1.6;}}
table.cond{{width:100%;border-collapse:collapse;font-size:13px;}}
table.cond td{{padding:7px 10px;border-bottom:1px solid #e5e5e5;vertical-align:top;line-height:1.5;}}
table.cond td:first-child{{font-weight:bold;width:40%;background:#f5f5f5;}}
.footer{{border-top:1px solid #e0e0e0;margin-top:22px;padding-top:12px;font-size:13px;color:#444;}}
.firma-nombre{{font-weight:bold;color:#1B2A4A;font-size:14px;margin:0 0 2px 0;}}
.firma-cargo{{color:#E8601C;font-size:13px;margin:2px 0;}}
.pie{{color:#aaa;font-size:11px;margin-top:10px;border-top:1px solid #eee;padding-top:8px;text-align:center;}}
</style>
</head>
<body>
<div class="wrapper">
<div class="header-bar">
  {logo_html}
  <div style="display:flex;align-items:center;gap:16px;">
    <div class="header-info">
      <strong>BOOM LOGISTICS COLOMBIA S.A.S.</strong>
      Soluciones de Transporte Especializado
    </div>
    {sello_html}
  </div>
</div>
<div class="ref-bar">
  <p>REF: {ref_fmt} &nbsp;|&nbsp; {cliente.upper()}{ruta} &nbsp;|&nbsp; {mes_anio}{ref_extra_html}{desc_bar}</p>
</div>

<div class="body">
  <div class="greeting">
    <p>Hola {saludo_nombre},</p>
    <p>En atenci&oacute;n a su solicitud, BOOM Logistics Colombia S.A.S. presenta a continuaci&oacute;n su propuesta para <strong>{intro_txt}</strong>.</p>
  </div>

  {cargo_section_html}

  <!-- ===== PROPUESTA ECONÓMICA ===== -->
  <div class="section-title">{sec}. PROPUESTA ECON&Oacute;MICA CON EQUIPO</div>

  <div class="fase-title">TRANSPORTE{ruta_fase}</div>
  <div class="table-scroll">
  <table class="det">
    <thead><tr>
      <th style="text-align:left;">Servicio / Equipo</th>
      <th>Cant.</th>
      <th>Tarifa c/u</th>
      <th style="text-align:right;">Total</th>
    </tr></thead>
    <tbody>
      {eq_rows_html}
    </tbody>
  </table>
  </div>

  {summary_html}

  <!-- Stand-by -->
  <p style="font-size:12px;font-weight:bold;color:#1B2A4A;margin:16px 0 5px 0;">STAND-BY</p>
  <div class="table-scroll">
  <table class="det">
    <thead><tr>
      <th style="text-align:left;">Equipo</th>
      <th>Tiempo libre cargue</th>
      <th>Tiempo libre descargue</th>
      <th style="text-align:right;">Stand-By (por d&iacute;a)</th>
    </tr></thead>
    <tbody>
      {standby_rows_html}
    </tbody>
  </table>
  </div>
  <p style="font-size:11px;color:#666;margin:4px 0 0 2px;">{standby_note}</p>

  {notes_section_html}

  <!-- ===== CONDICIONES COMERCIALES ===== -->
  <div class="section-title">{cond_num}. CONDICIONES COMERCIALES</div>
  <table class="cond">
    <tr><td>Forma de pago</td><td>{forma_pago}</td></tr>
    <tr><td>Moneda</td><td>Pesos colombianos (COP)</td></tr>
    <tr><td>Vigencia</td><td>{vigencia} d&iacute;as calendario a partir de la fecha de emisi&oacute;n</td></tr>
    <tr><td>P&oacute;liza de carga</td><td>{poliza_carga}</td></tr>
    <tr><td>P&oacute;liza RCE</td><td>{poliza_rc}</td></tr>
  </table>

  <!-- ===== EXCLUSIONES ===== -->
  <div class="section-title">{excl_num}. EXCLUSIONES</div>
  <ul class="notas">
{excl_html}  </ul>

  <!-- ===== FIRMA ===== -->
  <div class="footer">
    <p style="margin:0 0 10px 0;">Quedamos atentos a sus comentarios.</p>
    <p style="margin:0 0 2px 0;">Cordialmente,</p>
    <p class="firma-nombre">{comercial}</p>
    <p class="firma-cargo">{cargo_com}</p>
    <p style="margin:2px 0;">BOOM Logistics Colombia S.A.S.</p>
    <p style="margin:2px 0;">{email_com}</p>
    <p class="pie">BOOM LOGISTICS S.A.S. &nbsp;|&nbsp; Oferta v&aacute;lida por {vigencia} d&iacute;as &nbsp;|&nbsp; Ref: {ref_fmt} | {cliente.upper()}</p>
  </div>
</div>
</div>
</body>
</html>"""
    return _inject_anexo(_html_oferta)


# ── PDF helpers ───────────────────────────────────────────────────────────────
def _detect_service_type(equipos: list, notas: str = "") -> dict:
    """Detects service characteristics for dynamic PDF generation."""
    all_text = " ".join(
        (eq.get("equipo", "") + " " + eq.get("config", "")).lower()
        for eq in equipos
    ) + " " + notas.lower()

    has_izaje = any(kw in all_text for kw in [
        "izaje", "grúa", "grua", "modular", "patineta", "izamiento", "montaje con grúa"
    ])
    has_transport = any(kw in all_text for kw in [
        "transporte", "cama baja", "cama alta", "flete", "contenedor"
    ])
    has_security = any(kw in all_text for kw in ["escolta", "tecnólogo", "tecnologo"])

    phase_nums = set()
    for eq in equipos:
        m = re.search(r"fase\s*(\d+)", eq.get("equipo", "").lower())
        if m:
            phase_nums.add(int(m.group(1)))

    return {
        "has_izaje": has_izaje,
        "has_transport": has_transport,
        "has_security": has_security,
        "is_multiphase": len(phase_nums) > 0,
        "phases": sorted(phase_nums),
    }


def _split_notas(notas_raw: str) -> tuple:
    """Splits notas into (izaje_lines, other_lines)."""
    izaje_kws = ["stand-by", "standby", "/hora", "hora adicional", "proporcional",
                 "cobro inicia", "inicio de tiempo", "horas adicionales serán"]
    izaje, other = [], []
    for n in [n.strip() for n in notas_raw.split("\n") if n.strip()]:
        if any(kw in n.lower() for kw in izaje_kws):
            izaje.append(n)
        else:
            other.append(n)
    return izaje, other


# ── PDF offer generation ──────────────────────────────────────────────────────
def generar_pdf_oferta(data: dict) -> bytes:
    if not REPORTLAB_OK:
        raise RuntimeError("reportlab no está instalado")

    # ── Palette ───────────────────────────────────────────────────────────────
    NAVY    = HexColor("#1B2A4A")
    ORANGE  = HexColor("#E8601C")
    LGRAY   = HexColor("#F7F8FA")
    BORDER  = HexColor("#D8DCE4")
    FGRAY   = HexColor("#EEF0F4")
    MED     = HexColor("#5A6373")
    ACCENT  = HexColor("#F0F4FF")   # very light blue tint for alt rows

    # ── Data ──────────────────────────────────────────────────────────────────
    ref_fmt      = _fmt_ref(data.get("ref", "260001"))
    cliente      = data.get("cliente", "")
    contacto     = data.get("contacto", "")
    email_cl     = data.get("email_cliente", "")
    ref_cliente  = data.get("ref_cliente", "") or ""
    cliente_final= data.get("cliente_final", "") or ""
    origen       = data.get("origen", "")
    destino      = data.get("destino", "")
    mes_anio     = data.get("mes_anio", datetime.now().strftime("%b %Y").upper())
    descripcion  = data.get("descripcion", "")
    comercial    = data.get("comercial", "Natalia Vargas")
    cargo_str, email_com = COMERCIAL_INFO.get(comercial, ("Ejecutiva Comercial", "nvargas@boomlts.com"))
    cargo_items  = data.get("cargo_items", []) or []
    equipos      = data.get("equipos", []) or []
    notas_raw    = data.get("notas", "")
    forma_pago   = data.get("forma_pago", "50% anticipo / 50% a 30 días tras radicación de factura")
    vigencia     = data.get("vigencia", 30)
    poliza_carga = data.get("poliza_carga", "Hasta $4.000.000.000 COP por despacho")
    poliza_rc    = data.get("poliza_rc", "Hasta $4.000.000.000 COP")
    resolucion   = data.get("resolucion", "Carga extradimensionada/extrapesada incluida")
    excl_raw     = data.get("exclusiones", "")
    fecha_str    = data.get("fecha", "") or datetime.now().strftime("%d/%m/%Y")

    svc = _detect_service_type(equipos, notas_raw)
    izaje_notas, other_notas = _split_notas(notas_raw)

    # ── Document ──────────────────────────────────────────────────────────────
    buf = BytesIO()
    ML, MR, MT, MB = 1.6*cm, 1.6*cm, 1.4*cm, 1.9*cm
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=ML, rightMargin=MR, topMargin=MT, bottomMargin=MB,
        title=f"Oferta {ref_fmt} — {cliente}",
    )
    W = A4[0] - ML - MR

    # ── Page footer callback ──────────────────────────────────────────────────
    def _page_footer(canvas, doc):
        canvas.saveState()
        pw = A4[0]
        y  = 0.65 * cm
        canvas.setStrokeColor(BORDER)
        canvas.setLineWidth(0.4)
        canvas.line(ML, y + 0.28*cm, pw - MR, y + 0.28*cm)
        canvas.setFont("Helvetica", 6.5)
        canvas.setFillColor(MED)
        canvas.drawString(ML, y,
            f"BOOM Logistics Colombia S.A.S.  ·  NIT 900.548.985-7  ·  Ref. {ref_fmt}")
        canvas.drawRightString(pw - MR, y,
            f"Página {doc.page}  ·  {cliente.upper()}")
        canvas.restoreState()

    # ── Style factory ─────────────────────────────────────────────────────────
    def ps(name, **kw):
        d = dict(fontName="Helvetica", fontSize=8.5, leading=12,
                 textColor=NAVY, spaceAfter=0, spaceBefore=0)
        d.update(kw)
        return ParagraphStyle(name, **d)

    s_body    = ps("body",    fontSize=9,   leading=14)
    s_white   = ps("white",   textColor=white, fontName="Helvetica-Bold", fontSize=8.5)
    s_cell    = ps("cell",    fontSize=8,   leading=12)
    s_cell_c  = ps("cell_c",  fontSize=8,   leading=12, alignment=TA_CENTER)
    s_cell_r  = ps("cell_r",  fontSize=8,   leading=12, alignment=TA_RIGHT)
    s_cell_w  = ps("cell_w",  fontSize=8,   leading=12, textColor=white,
                   fontName="Helvetica-Bold", alignment=TA_CENTER)
    s_cell_wr = ps("cell_wr", fontSize=8,   leading=12, textColor=white,
                   fontName="Helvetica-Bold", alignment=TA_RIGHT)
    s_lbl     = ps("lbl",     fontSize=8,   leading=12, fontName="Helvetica-Bold")
    s_val     = ps("val",     fontSize=8,   leading=12, textColor=MED)
    s_foot    = ps("foot",    fontSize=7.5, leading=11, textColor=MED, alignment=TA_CENTER)

    SP  = lambda n: Spacer(1, n)   # compact spacer shorthand

    def note_para(text):
        m = re.match(r'^([^:]+:)\s*(.*)', text.strip(), re.DOTALL)
        st = ps("nota", fontSize=8, leading=13, leftIndent=10, spaceAfter=3)
        if m:
            return Paragraph(f"<font color='#E8601C'>▸</font> <b>{m.group(1)}</b> {m.group(2)}", st)
        return Paragraph(f"<font color='#E8601C'>▸</font> {text}", st)

    # ── Section title ─────────────────────────────────────────────────────────
    sec = [0]
    def sec_title(title: str):
        sec[0] += 1
        t = Table(
            [[Paragraph(f"<b>{sec[0]}. {title.upper()}</b>",
                        ps("st", fontName="Helvetica-Bold", fontSize=8.5,
                           textColor=white, leading=12))]],
            colWidths=[W],
        )
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), NAVY),
            ("LEFTPADDING",   (0, 0), (-1, -1), 10),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        return t

    # ── Standard table style ─────────────────────────────────────────────────
    def _eq_table_style(data_rows, n_total_span):
        ts = [
            ("BACKGROUND",    (0, 0), (-1, 0),  NAVY),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING",   (0, 0), (-1, -1), 7),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 7),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("GRID",          (0, 0), (-1, -2), 0.4, BORDER),
            ("BACKGROUND",    (0, -1), (-1, -1), ORANGE),
            ("SPAN",          (0, -1), (n_total_span, -1)),
            ("ALIGN",         (0, -1), (0, -1), "RIGHT"),
            ("TOPPADDING",    (0, -1), (-1, -1), 7),
            ("BOTTOMPADDING", (0, -1), (-1, -1), 7),
        ]
        for i in range(1, len(data_rows) - 1):
            ts.append(("BACKGROUND", (0, i), (-1, i), ACCENT if i % 2 == 0 else white))
        return TableStyle(ts)

    story = []

    # ── Header: logo + company ────────────────────────────────────────────────
    logo = _logo_for_pdf(1.1, max_width_cm=3.4)
    company_p = Paragraph(
        '<font name="Helvetica-Bold" size="8.5" color="white">BOOM LOGISTICS COLOMBIA S.A.S.</font><br/>'
        '<font size="6.5" color="#4FC3D8">NIT: 900.548.985-7</font><br/>'
        '<font size="6.5" color="white">Soluciones de Transporte Especializado</font>',
        ps("hdr_r", textColor=white, fontSize=8, leading=10, alignment=TA_RIGHT),
    )
    left_cell = logo if logo else Paragraph("<b>BOOM</b>",
        ps("boom_fb", textColor=white, fontName="Helvetica-Bold", fontSize=15))
    sello = _sello_for_pdf(1.5, max_width_cm=2.6)
    if sello:
        hdr = Table([[left_cell, company_p, sello]], colWidths=[W * 0.20, W * 0.62, W * 0.18])
        hdr_align = ("ALIGN", (2, 0), (2, 0), "RIGHT")
    else:
        hdr = Table([[left_cell, company_p]], colWidths=[W * 0.22, W * 0.78])
        hdr_align = ("ALIGN", (1, 0), (1, 0), "RIGHT")
    hdr.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), NAVY),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN",         (0, 0), (0,  0),  "LEFT"),
        ("ALIGN",         (1, 0), (1,  0),  "RIGHT"),
        hdr_align,
        ("LEFTPADDING",   (0, 0), (-1, -1), 12),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 12),
        ("TOPPADDING",    (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
    ]))
    story.append(hdr)

    # ── Orange identifier bar ─────────────────────────────────────────────────
    ruta = ""
    if origen and destino:
        ruta = f"  ·  {origen.upper()} → {destino.upper()}"
    elif origen:
        ruta = f"  ·  {origen.upper()}"
    ref_extra = f"  ·  {ref_cliente}" if ref_cliente else ""
    if ref_extra and cliente_final:
        ref_extra += f" / {cliente_final}"
    ref_t = Table(
        [[Paragraph(
            f"<b>OFERTA  {ref_fmt}  ·  {cliente.upper()}{ruta}  ·  {mes_anio}{ref_extra}</b>",
            ps("refb", textColor=white, fontSize=8.5, leading=12)
        )]],
        colWidths=[W],
    )
    ref_t.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, -1), ORANGE),
        ("LEFTPADDING",  (0, 0), (-1, -1), 12),
        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ("TOPPADDING",   (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 6),
    ]))
    story.append(ref_t)
    story.append(SP(8))

    # ── Client info block ────────────────────────────────────────────────────
    ruta_display = f"{origen}  →  {destino}" if origen and destino else (origen or destino or "—")
    _fd = fecha_str if isinstance(fecha_str, str) else str(fecha_str)

    if svc["is_multiphase"]:
        svc_label = "Proyecto especializado — multifase"
    elif svc["has_izaje"] and svc["has_transport"]:
        svc_label = "Transporte e izaje especializado"
    elif svc["has_izaje"]:
        svc_label = "Operación de izaje especializado"
    else:
        svc_label = "Transporte especializado"

    info_rows = [
        [Paragraph("<b>Para:</b>",       s_lbl), Paragraph(cliente,        s_val),
         Paragraph("<b>Referencia:</b>", s_lbl), Paragraph(ref_fmt,        s_val)],
        [Paragraph("<b>Atención:</b>",   s_lbl), Paragraph(contacto or "—", s_val),
         Paragraph("<b>Fecha:</b>",      s_lbl), Paragraph(_fd,            s_val)],
        [Paragraph("<b>Ruta:</b>",       s_lbl), Paragraph(ruta_display,   s_val),
         Paragraph("<b>Vigencia:</b>",   s_lbl), Paragraph(f"{vigencia} días", s_val)],
        [Paragraph("<b>Servicio:</b>",   s_lbl), Paragraph(svc_label,      s_val),
         Paragraph("<b>Moneda:</b>",     s_lbl), Paragraph("COP",          s_val)],
    ]
    n_base = len(info_rows)
    if descripcion:
        info_rows.append([
            Paragraph(
                f"<b>Descripción:</b>  "
                f"<font color='#5A6373'>{descripcion[:220]}</font>",
                ps("desc_v", fontSize=8, leading=12, textColor=NAVY)
            ),
            "", "", "",
        ])
    info_t = Table(info_rows, colWidths=[W*0.15, W*0.38, W*0.15, W*0.32])
    info_ts = [
        ("VALIGN",        (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING",   (0, 0), (-1, -1), 7),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 7),
        ("GRID",          (0, 0), (-1, -1), 0.4, BORDER),
        ("BACKGROUND",    (0, 0), (0, n_base - 1), FGRAY),
        ("BACKGROUND",    (2, 0), (2, n_base - 1), FGRAY),
    ]
    if descripcion:
        info_ts += [
            ("SPAN",       (0, n_base), (3, n_base)),
            ("BACKGROUND", (0, n_base), (3, n_base), ACCENT),
        ]
    info_t.setStyle(TableStyle(info_ts))
    story.append(info_t)
    story.append(SP(10))

    # ── Helper: render a section block (title + content) keeping together ──────
    def _section(title, content_flowables):
        return KeepTogether([sec_title(title), SP(3)] + content_flowables + [SP(8)])

    # ── Detect which cargo columns actually have data ─────────────────────────
    def _col_has(items, *keys):
        return any(any(ci.get(k) for k in keys) for ci in items)

    show_dim = _col_has(cargo_items, "dimensiones")
    show_peso = _col_has(cargo_items, "peso")
    show_vol  = _col_has(cargo_items, "volumen")
    show_orig = _col_has(cargo_items, "origen_detalle")
    show_dest = _col_has(cargo_items, "destino_detalle")
    has_detail = show_dim or show_peso or show_vol

    # ── TABLE 1: Physical cargo — full detail or compact ──────────────────────
    if cargo_items:
        if has_detail:
            # ── Full detail table (7 cols, only show populated columns) ───────
            active_cols = ["descripcion"]
            hdr_labels  = ["Descripción"]
            col_ws      = [W * 0.24]
            if show_dim:
                active_cols.append("dimensiones"); hdr_labels.append("Dimensiones"); col_ws.append(W * 0.17)
            if show_peso:
                active_cols.append("peso");        hdr_labels.append("Peso");        col_ws.append(W * 0.10)
            if show_vol:
                active_cols.append("volumen");     hdr_labels.append("Volumen");     col_ws.append(W * 0.11)
            # distribute remaining width to Origen + Destino
            used = sum(col_ws)
            rem  = W - used
            if show_orig:
                active_cols.append("origen_detalle");  hdr_labels.append("Origen");  col_ws.append(rem * 0.45)
            if show_dest:
                active_cols.append("destino_detalle"); hdr_labels.append("Destino"); col_ws.append(rem * 0.55 if show_orig else rem)

            c_hdr = [Paragraph(f"<b>{lbl}</b>", s_cell_w) for lbl in hdr_labels]
            c_data = [c_hdr]
            for ci in cargo_items:
                row = []
                desc = ci.get("descripcion", "") or ""
                tipo = ci.get("tipo", "") or ""
                row.append(Paragraph(
                    f"{desc}<br/><font size='6.5' color='#888'>{tipo}</font>" if tipo else desc,
                    s_cell))
                for col in active_cols[1:]:
                    val = (ci.get(col) or "").replace("\n", "<br/>")
                    row.append(Paragraph(val, s_cell))
                c_data.append(row)

            c_t = Table(c_data, colWidths=col_ws, repeatRows=1)
            c_ts = [
                ("BACKGROUND",    (0, 0), (-1, 0),  NAVY),
                ("VALIGN",        (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING",   (0, 0), (-1, -1), 6),
                ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
                ("TOPPADDING",    (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("GRID",          (0, 0), (-1, -1), 0.4, BORDER),
                ("ALIGN",         (1, 0), (-1, 0),  "CENTER"),
                ("ALIGN",         (1, 1), (-1, -1), "LEFT"),
            ]
            for i in range(1, len(c_data)):
                c_ts.append(("BACKGROUND", (0, i), (-1, i), ACCENT if i % 2 == 0 else white))
            c_t.setStyle(TableStyle(c_ts))
            story.append(_section("DETALLE TÉCNICO DE LA CARGA", [c_t]))

        else:
            # ── Compact combined table: Servicio/Descripción | Equipo | Origen | Destino ──
            # Merge cargo info into the economic table (rendered later as one unified table)
            pass  # handled below via _use_combined flag

    _use_combined = cargo_items and not has_detail

    # ── Security callout ──────────────────────────────────────────────────────
    if svc["has_security"]:
        sec_box = Table(
            [[Paragraph(
                "<b>ESQUEMA DE SEGURIDAD:</b>  2 Escoltas + 2 Tecnólogos incluidos "
                "— carga extradimensionada / extrapesada (ancho &gt; 3.00 m).",
                ps("sec_n", textColor=NAVY, fontSize=8, leading=12)
            )]],
            colWidths=[W],
        )
        sec_box.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), HexColor("#FFF7F3")),
            ("TOPPADDING",    (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ("LEFTPADDING",   (0, 0), (-1, -1), 12),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
            ("BOX",           (0, 0), (-1, -1), 0.4, BORDER),
            ("LINEBELOW",     (0, 0), (-1, -1), 2.5, ORANGE),
        ]))
        story.append(sec_box)
        story.append(SP(8))

    # ── TABLE 2: Economic proposal (or combined if no cargo detail) ──────────
    if _use_combined:
        # ── Combined: Servicio/Descripción | Equipo/Config | Origen | Valor ──
        cb_hdr = [
            Paragraph("<b>Servicio / Descripción</b>", s_cell_w),
            Paragraph("<b>Equipo / Contenedor</b>",    s_cell_w),
            Paragraph("<b>Origen</b>",                 s_cell_w),
            Paragraph("<b>Valor</b>",                  s_cell_w),
        ]
        cw_cb = [W*0.36, W*0.24, W*0.20, W*0.20]
        cb_data = [cb_hdr]
        total_oferta = 0
        for i, eq in enumerate(equipos):
            cant   = int(eq.get("cant", 1) or 1)
            v_unit = int(eq.get("valor_unit", 0) or 0)
            sub    = cant * v_unit
            total_oferta += sub
            cfg = (eq.get("config") or "").strip()
            # match cargo_item if available
            ci = cargo_items[i] if i < len(cargo_items) else {}
            cargo_desc = ci.get("descripcion", "") or ""
            cargo_tipo = ci.get("tipo", "") or ""
            orig_val   = (ci.get("origen_detalle") or origen or "").strip()
            # service cell: equipo name bold + cargo description small
            svc_parts = f"<b>{eq.get('equipo','')}</b>"
            if cargo_desc:
                svc_parts += f"<br/><font size='7' color='#888'>{cargo_desc}"
                if cargo_tipo:
                    svc_parts += f" — {cargo_tipo}"
                svc_parts += "</font>"
            cb_data.append([
                Paragraph(svc_parts, s_cell),
                Paragraph(cfg, s_cell),
                Paragraph(orig_val, s_cell),
                Paragraph(f"<b>{_fmt_cop(sub)}</b>" if sub else "—", s_cell_r),
            ])
        # extra cargo items without matching equipo
        for ci in cargo_items[len(equipos):]:
            dest_val = (ci.get("destino_detalle") or destino or "").strip()
            cb_data.append([
                Paragraph(f"<font color='#888'>{ci.get('descripcion','')}</font>", s_cell),
                Paragraph("—", s_cell_c),
                Paragraph(dest_val, s_cell),
                Paragraph("—", s_cell_c),
            ])
        # total row
        ruta_total = f"{origen.upper()} → {destino.upper()}" if origen and destino else (origen or destino or "")
        cb_data.append([
            Paragraph("<b>TOTAL OFERTA</b>",
                      ps("tot_cb", textColor=white, fontName="Helvetica-Bold",
                         fontSize=10, leading=14, alignment=TA_RIGHT)),
            Paragraph(ruta_total,
                      ps("tot_rt", textColor=white, fontSize=8, leading=12, alignment=TA_CENTER)),
            "",
            Paragraph(f"<b>{_fmt_cop(total_oferta)} COP</b>", s_cell_wr),
        ])
        cb_ts = [
            ("BACKGROUND",    (0, 0), (-1, 0),  NAVY),
            ("VALIGN",        (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING",   (0, 0), (-1, -1), 7),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 7),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("GRID",          (0, 0), (-1, -2), 0.4, BORDER),
            ("BACKGROUND",    (0, -1), (-1, -1), ORANGE),
            ("SPAN",          (0, -1), (2, -1)),
            ("ALIGN",         (3, -1), (3, -1), "RIGHT"),
            ("TOPPADDING",    (0, -1), (-1, -1), 7),
            ("BOTTOMPADDING", (0, -1), (-1, -1), 7),
        ]
        for i in range(1, len(cb_data) - 1):
            cb_ts.append(("BACKGROUND", (0, i), (-1, i), ACCENT if i % 2 == 0 else white))
        eq_t = Table(cb_data, colWidths=cw_cb, repeatRows=1)
        eq_t.setStyle(TableStyle(cb_ts))
        story.append(_section("DETALLE TÉCNICO Y ECONÓMICO", [eq_t]))

    else:
        # ── Standard economic table ───────────────────────────────────────────
        e_hdr = [
            Paragraph("<b>Concepto</b>",                s_cell_w),
            Paragraph("<b>Configuración de Equipo</b>", s_cell_w),
            Paragraph("<b>Cant.</b>",                   s_cell_w),
            Paragraph("<b>Valor</b>",                   s_cell_w),
        ]
        cw_e = [W*0.29, W*0.44, W*0.07, W*0.20]

        if svc["is_multiphase"]:
            LIGHT_BLUE = HexColor("#EBF3FF")
            PHASE_BG   = HexColor("#1E3A8A")
            phase_groups = {}
            no_phase_eqs = []
            for eq in equipos:
                _m = re.search(r"fase\s*(\d+)", eq.get("equipo", "").lower())
                if _m:
                    phase_groups.setdefault(int(_m.group(1)), []).append(eq)
                else:
                    no_phase_eqs.append(eq)

            eq_data   = [e_hdr]
            row_types = ["header"]
            grand_total = 0

            for ph in sorted(phase_groups.keys()):
                phase_total = 0
                eq_data.append([
                    Paragraph(f"<b>FASE {ph}</b>",
                              ps(f"ph{ph}h", fontName="Helvetica-Bold", fontSize=8.5,
                                 textColor=white, leading=12)),
                    "", "", "",
                ])
                row_types.append("phase_hdr")
                for eq in phase_groups[ph]:
                    cant   = int(eq.get("cant", 1) or 1)
                    v_unit = int(eq.get("valor_unit", 0) or 0)
                    sub    = cant * v_unit
                    phase_total += sub
                    grand_total += sub
                    cfg = (eq.get("config") or "").replace("\n", "<br/>")
                    eq_data.append([
                        Paragraph(eq.get("equipo", "") or "", s_cell),
                        Paragraph(cfg, s_cell),
                        Paragraph(str(cant), s_cell_c),
                        Paragraph(f"<b>{_fmt_cop(sub)}</b>", s_cell_r),
                    ])
                    row_types.append("item")
                eq_data.append([
                    Paragraph(f"<b>SUBTOTAL FASE {ph}</b>",
                              ps(f"ph{ph}s", fontName="Helvetica-Bold", fontSize=8,
                                 textColor=NAVY, leading=12, alignment=TA_RIGHT)),
                    "", "",
                    Paragraph(f"<b>{_fmt_cop(phase_total)}</b>",
                              ps(f"ph{ph}v", fontName="Helvetica-Bold", fontSize=8,
                                 textColor=NAVY, leading=12, alignment=TA_RIGHT)),
                ])
                row_types.append("subtotal")

            for eq in no_phase_eqs:
                cant   = int(eq.get("cant", 1) or 1)
                v_unit = int(eq.get("valor_unit", 0) or 0)
                sub    = cant * v_unit
                grand_total += sub
                cfg = (eq.get("config") or "").replace("\n", "<br/>")
                eq_data.append([
                    Paragraph(eq.get("equipo", "") or "", s_cell),
                    Paragraph(cfg, s_cell),
                    Paragraph(str(cant), s_cell_c),
                    Paragraph(f"<b>{_fmt_cop(sub)}</b>", s_cell_r),
                ])
                row_types.append("item")

            total_label = "TOTAL PROYECTO"
            eq_data.append([
                Paragraph(f"<b>{total_label}</b>",
                          ps("tot_lp", textColor=white, fontName="Helvetica-Bold",
                             fontSize=10, leading=14, alignment=TA_RIGHT)),
                "", "",
                Paragraph(f"<b>{_fmt_cop(grand_total)} COP</b>", s_cell_wr),
            ])
            row_types.append("total")

            ts_mp = [
                ("VALIGN",        (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING",   (0, 0), (-1, -1), 7),
                ("RIGHTPADDING",  (0, 0), (-1, -1), 7),
                ("TOPPADDING",    (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("GRID",          (0, 0), (-1, -1), 0.4, BORDER),
            ]
            for i, rt in enumerate(row_types):
                if rt == "header":
                    ts_mp.append(("BACKGROUND", (0, i), (-1, i), NAVY))
                elif rt == "phase_hdr":
                    ts_mp.extend([
                        ("BACKGROUND", (0, i), (-1, i), PHASE_BG),
                        ("SPAN",       (0, i), (-1, i)),
                        ("LEFTPADDING",(0, i), (-1, i), 12),
                        ("TOPPADDING", (0, i), (-1, i), 6),
                        ("BOTTOMPADDING", (0, i), (-1, i), 6),
                    ])
                elif rt == "subtotal":
                    ts_mp.extend([
                        ("BACKGROUND", (0, i), (-1, i), LIGHT_BLUE),
                        ("SPAN",       (0, i), (2,  i)),
                        ("ALIGN",      (3, i), (3,  i), "RIGHT"),
                        ("TOPPADDING", (0, i), (-1, i), 6),
                        ("BOTTOMPADDING", (0, i), (-1, i), 6),
                    ])
                elif rt == "total":
                    ts_mp.extend([
                        ("BACKGROUND",    (0, i), (-1, i), ORANGE),
                        ("SPAN",          (0, i), (2,  i)),
                        ("ALIGN",         (0, i), (0,  i), "RIGHT"),
                        ("TOPPADDING",    (0, i), (-1, i), 7),
                        ("BOTTOMPADDING", (0, i), (-1, i), 7),
                    ])
                else:
                    ts_mp.append(("BACKGROUND", (0, i), (-1, i),
                                  ACCENT if i % 2 == 0 else white))
            eq_t = Table(eq_data, colWidths=cw_e, repeatRows=1)
            eq_t.setStyle(TableStyle(ts_mp))

        else:
            eq_data = [e_hdr]
            total_oferta = 0
            for eq in equipos:
                cant   = int(eq.get("cant", 1) or 1)
                v_unit = int(eq.get("valor_unit", 0) or 0)
                sub    = cant * v_unit
                total_oferta += sub
                cfg = (eq.get("config") or "").replace("\n", "<br/>")
                eq_data.append([
                    Paragraph(eq.get("equipo", "") or "", s_cell),
                    Paragraph(cfg, s_cell),
                    Paragraph(str(cant), s_cell_c),
                    Paragraph(f"<b>{_fmt_cop(sub)}</b>", s_cell_r),
                ])
            eq_data.append([
                Paragraph("<b>TOTAL OFERTA</b>",
                          ps("tot_l", textColor=white, fontName="Helvetica-Bold",
                             fontSize=10, leading=14, alignment=TA_RIGHT)),
                "", "",
                Paragraph(f"<b>{_fmt_cop(total_oferta)} COP</b>", s_cell_wr),
            ])
            eq_t = Table(eq_data, colWidths=cw_e, repeatRows=1)
            eq_t.setStyle(_eq_table_style(eq_data, 2))

        story.append(_section("PROPUESTA ECONÓMICA", [eq_t]))

    # ── Stand-by table ────────────────────────────────────────────────────────
    if equipos:
        sb_hdr = [
            Paragraph("<b>Equipo</b>",                       s_cell_w),
            Paragraph("<b>Tiempo libre cargue</b>",          s_cell_w),
            Paragraph("<b>Tiempo libre descargue</b>",       s_cell_w),
            Paragraph("<b>Stand-By (por día)</b>",      s_cell_w),
        ]
        sb_data = [sb_hdr]
        for eq in equipos:
            sb_rate, sb_lc, sb_ld = _standby_for_equipo(
                eq.get("equipo", ""), eq.get("config", "")
            )
            sb_data.append([
                Paragraph(eq.get("equipo", "") or "", s_cell),
                Paragraph(sb_lc,   s_cell_c),
                Paragraph(sb_ld,   s_cell_c),
                Paragraph(f"<b>{sb_rate}</b>", s_cell_r),
            ])
        sb_ts = [
            ("BACKGROUND",    (0, 0), (-1, 0),  NAVY),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING",   (0, 0), (-1, -1), 7),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 7),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("GRID",          (0, 0), (-1, -1), 0.4, BORDER),
        ]
        for i in range(1, len(sb_data)):
            sb_ts.append(("BACKGROUND", (0, i), (-1, i), ACCENT if i % 2 == 0 else white))
        sb_t = Table(sb_data, colWidths=[W*0.34, W*0.22, W*0.22, W*0.22], repeatRows=1)
        sb_t.setStyle(TableStyle(sb_ts))
        sb_note = Paragraph(
            "* Las horas adicionales serán cobradas proporcionalmente según tarifa establecida."
            + (" El cobro del equipo de izaje inicia desde la llegada al sitio designado."
               if svc["has_izaje"] else ""),
            ps("sb_note", fontSize=7.5, leading=11, textColor=MED)
        )
        story.append(_section("STAND-BY", [sb_t, SP(4), sb_note]))

    # ── Izaje conditions ──────────────────────────────────────────────────────
    if svc["has_izaje"] and izaje_notas:
        story.append(_section("CONDICIONES TÉCNICAS DE IZAJE",
                               [note_para(n) for n in izaje_notas]))

    # ── Other technical notes ─────────────────────────────────────────────────
    if other_notas:
        story.append(_section("NOTAS TÉCNICAS DE OPERACIÓN",
                               [note_para(n) for n in other_notas]))

    # ── Commercial conditions ─────────────────────────────────────────────────
    cond_rows = [
        ("Forma de pago",   forma_pago),
        ("Póliza de carga", poliza_carga),
        ("Póliza RC",       poliza_rc),
        ("Resolución",      resolucion),
    ]
    cond_left  = [
        [Paragraph(f"<b>{lbl}</b>", ps("cl", fontName="Helvetica-Bold", fontSize=8,
                                        textColor=NAVY, leading=12)),
         Paragraph(val, ps("cv", fontSize=8, leading=12, textColor=MED))]
        for lbl, val in cond_rows
    ]
    cond_t = Table(cond_left, colWidths=[W*0.30, W*0.70])
    cond_ts = [
        ("VALIGN",        (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
        ("GRID",          (0, 0), (-1, -1), 0.4, BORDER),
    ]
    for i in range(len(cond_left)):
        cond_ts.append(("BACKGROUND", (0, i), (0, i), FGRAY))
        cond_ts.append(("BACKGROUND", (1, i), (1, i), ACCENT if i % 2 == 0 else white))
    cond_t.setStyle(TableStyle(cond_ts))
    story.append(_section("CONDICIONES COMERCIALES", [cond_t]))

    # ── Exclusions ────────────────────────────────────────────────────────────
    excls = [e.strip() for e in excl_raw.split("\n") if e.strip()]
    if excls:
        story.append(_section("EXCLUSIONES", [note_para(e) for e in excls]))

    # ── Signature block ───────────────────────────────────────────────────────
    story.append(SP(4))
    story.append(HRFlowable(width=W, color=BORDER, thickness=0.5))
    story.append(SP(10))
    # Two-column: each cell is a single Paragraph combining multiple lines via <br/>
    left_txt = (
        f"<b>{comercial}</b><br/>"
        f"<font color='#E8601C'><b>{cargo_str}</b></font><br/>"
        f"<font size='8'>BOOM Logistics Colombia S.A.S.</font><br/>"
        f"<font size='8' color='#5A6373'>{email_com}</font>"
    )
    right_txt = ""
    sig_t = Table(
        [[Paragraph(left_txt,  ps("sig_l", fontSize=9, leading=14, textColor=NAVY)),
          Paragraph(right_txt, ps("sig_r", fontSize=8.5, leading=13, textColor=MED))]],
        colWidths=[W * 0.48, W * 0.52],
    )
    sig_t.setStyle(TableStyle([
        ("VALIGN",       (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING",  (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING",   (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 0),
    ]))
    story.append(sig_t)

    doc.build(story,
              onFirstPage=_page_footer,
              onLaterPages=_page_footer)
    return buf.getvalue()


def _extraer_info(texto: str) -> dict:
    result = {}
    tl = texto.lower()
    m = re.search(r"[\w.\-]+@[\w.\-]+\.\w+", texto)
    if m:
        result["email_cliente"] = m.group(0)
    for pat in [
        r"(?:empresa|cliente|para|de|señores?)[:\s]+([A-ZÁÉÍÓÚÑ][A-Za-záéíóúñ\s&.,]+?(?:S\.A\.S?\.?|LTDA\.?|S\.A\.)?)",
        r"^([A-ZÁÉÍÓÚÑ][A-Za-záéíóúñ\s]{4,35})\s*[\n,]",
    ]:
        m2 = re.search(pat, texto, re.IGNORECASE | re.MULTILINE)
        if m2:
            result["cliente"] = m2.group(1).strip().rstrip(",.")
            break
    for pat in [
        r"(?:desde|origen|salida)[:\s]+([A-Za-záéíóúñ\s]+?)(?:\s+hasta|\s+a\s|\s+hacia|\.|,|\n)",
        r"(?:puerto de|ciudad de)\s+([A-Za-záéíóúñ\s]+?)(?:\s+a\s|\.|,|\n)",
    ]:
        m3 = re.search(pat, texto, re.IGNORECASE)
        if m3:
            result["origen"] = m3.group(1).strip()
            break
    for pat in [r"(?:hasta|destino|hacia|a)\s+([A-Za-záéíóúñ\s,]+?)(?:\.|,|\n|$)"]:
        m4 = re.search(pat, texto, re.IGNORECASE)
        if m4:
            v = m4.group(1).strip()
            if len(v) > 2:
                result["destino"] = v
            break
    if any(x in tl for x in ["izaje", "grúa", "grua", "izamiento", "montaje"]):
        result["tipo"] = "SPOT MIXTO" if "transporte" in tl else "SPOT IZAJE"
    elif any(x in tl for x in ["transporte", "contenedor", "carga", "flete"]):
        result["tipo"] = "SPOT TRANSPORTE"
    frases = [f.strip() for f in re.split(r"[.\n]", texto) if len(f.strip()) > 25]
    if frases:
        result["descripcion"] = frases[0][:220]
    return result


# ── DB init ───────────────────────────────────────────────────────────────────
def _ensure_db():
    try:
        admin = pgdb.connect(host=DB_HOST, port=DB_PORT, database="postgres",
                              user=DB_USER, password=DB_PASSWORD)
        admin.autocommit = True
        cur = admin.cursor()
        cur.execute("SELECT 1 FROM pg_database WHERE datname = %s", (DB_NAME,))
        if not cur.fetchone():
            cur.execute(f'CREATE DATABASE "{DB_NAME}"')
            print(f"[DB] Base de datos '{DB_NAME}' creada.")
        admin.close()
    except Exception as e:
        print(f"[DB] Advertencia al crear BD: {e}")
    try:
        conn = pgdb.connect(host=DB_HOST, port=DB_PORT, database=DB_NAME,
                             user=DB_USER, password=DB_PASSWORD)
        conn.autocommit = True
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS ofertas (
                id bigint generated always as identity primary key,
                num text unique not null,
                mes text, fecha date, cliente text,
                realizada text, formalizada text, unidad text, tipo text,
                valor bigint default 0, estado text default 'ENVIADO',
                respuesta text, facturacion text,
                general text, seguimiento text,
                mes_aceptado text, fecha_facturacion date,
                valor_facturado bigint, no_factura text,
                created_at timestamptz default now()
            )
        """)
        # Migraciones: agrega columnas si la tabla ya existía sin ellas
        cur.execute("ALTER TABLE ofertas ADD COLUMN IF NOT EXISTS facturacion text")
        cur.execute("ALTER TABLE ofertas ADD COLUMN IF NOT EXISTS sector text")
        cur.execute("ALTER TABLE ofertas ADD COLUMN IF NOT EXISTS pdf_data jsonb")
        cur.execute("ALTER TABLE ofertas ADD COLUMN IF NOT EXISTS costo_proyecto bigint DEFAULT NULL")
        # Anulación (ISO 9001 · trazabilidad): la oferta conserva su número,
        # no se borra ni se reutiliza el consecutivo. Queda motivo, quién y cuándo.
        cur.execute("ALTER TABLE ofertas ADD COLUMN IF NOT EXISTS anulada boolean DEFAULT false")
        cur.execute("ALTER TABLE ofertas ADD COLUMN IF NOT EXISTS anulada_motivo text")
        cur.execute("ALTER TABLE ofertas ADD COLUMN IF NOT EXISTS anulada_por text")
        cur.execute("ALTER TABLE ofertas ADD COLUMN IF NOT EXISTS anulada_fecha timestamptz")
        # Ofertas de PRUEBA: no gastan consecutivo real ni afectan la numeración
        cur.execute("ALTER TABLE ofertas ADD COLUMN IF NOT EXISTS es_prueba boolean DEFAULT false")
        # Fecha en que la oferta se marcó ACEPTADA. Sirve para que el módulo de
        # Aprobadas ordene lo último aceptado ARRIBA (sin cambiar el número).
        cur.execute("ALTER TABLE ofertas ADD COLUMN IF NOT EXISTS aceptada_fecha timestamptz")
        # Aprobacion PARCIAL: el cliente autoriza solo una parte de la oferta al notificar.
        # valor = total cotizado ; valor_aprobado = lo realmente autorizado (<= valor).
        cur.execute("ALTER TABLE ofertas ADD COLUMN IF NOT EXISTS valor_aprobado bigint DEFAULT NULL")
        # Moneda de la oferta (COP por defecto). Las ofertas en USD ya no se muestran
        # ni se suman como si fueran pesos en el Control.
        cur.execute("ALTER TABLE ofertas ADD COLUMN IF NOT EXISTS moneda text DEFAULT 'COP'")

        # ── CANDADO DE CONSECUTIVO ────────────────────────────────────────────
        # Evita que dos ofertas ACTIVAS compartan el mismo número (bug de
        # consecutivo duplicado). Tiene dos partes idempotentes:
        #   1) Resolver duplicados que ya existan: entre ofertas activas (no
        #      prueba, no anulada) con el MISMO número, se conserva la MÁS
        #      ANTIGUA (menor id) y se ANULAN las más nuevas (que son las que se
        #      crearon por accidente al chocar el consecutivo). Anular es
        #      reversible y conserva la trazabilidad (ISO 9001): el número no se
        #      borra ni se reutiliza.
        #   2) Crear un índice ÚNICO parcial para que la base RECHACE de raíz un
        #      segundo registro activo con el mismo número. Las ofertas de prueba
        #      y las anuladas quedan por fuera del candado (sí pueden repetir).
        try:
            cur.execute("""
                SELECT CAST(num AS INTEGER) AS n, array_agg(id ORDER BY id) AS ids
                  FROM ofertas
                 WHERE NOT COALESCE(es_prueba, false)
                   AND NOT COALESCE(anulada, false)
                   AND num ~ '^[0-9]+$'
                 GROUP BY CAST(num AS INTEGER)
                HAVING COUNT(*) > 1
            """)
            _dups = cur.fetchall() or []
            for _fila in _dups:
                _n = _fila[0]
                _ids = list(_fila[1] or [])
                _conservar = _ids[0]      # la más antigua queda vigente
                _anular = _ids[1:]        # las más nuevas se anulan
                for _oid in _anular:
                    cur.execute("""
                        UPDATE ofertas
                           SET anulada = true, estado = 'ANULADA',
                               anulada_motivo = COALESCE(anulada_motivo,
                                   'Consecutivo duplicado: el número ya estaba asignado a otra oferta.'),
                               anulada_por = COALESCE(anulada_por, 'Sistema'),
                               anulada_fecha = COALESCE(anulada_fecha, now())
                         WHERE id = %s
                    """, (_oid,))
                    try:
                        cur.execute("""
                            INSERT INTO oferta_historial
                                (oferta_id, oferta_num, campo, valor_ant, valor_nuevo, usuario)
                            VALUES (%s, %s, 'ANULACIÓN', 'VIGENTE',
                                    'ANULADA — consecutivo duplicado (saneamiento de integridad)', 'Sistema')
                        """, (_oid, str(_n)))
                    except Exception:
                        pass
                print(f"[DB][consecutivo] Duplicado {_n}: se conserva id={_conservar}, "
                      f"se anula(n) id={_anular}")
            if not _dups:
                print("[DB][consecutivo] Sin duplicados de consecutivo.")
        except Exception as e:
            print(f"[DB][consecutivo] No se pudieron resolver duplicados: {e}")
        try:
            cur.execute("""
                CREATE UNIQUE INDEX IF NOT EXISTS ofertas_num_activo_uidx
                    ON ofertas (CAST(num AS INTEGER))
                 WHERE NOT COALESCE(es_prueba, false)
                   AND NOT COALESCE(anulada, false)
            """)
            print("[DB][consecutivo] Candado de consecutivo activo listo.")
        except Exception as e:
            print(f"[DB][consecutivo] No se pudo crear el candado: {e}")

        # ── CORRECCIÓN PUNTUAL Y AUTO-SANADORA: el 261161 pertenece a TRANSBORDER ──
        # El consecutivo 261161 es de TRANSBORDER (COTECMAR Mamonal → Puerto
        # Cartagena, $13.600.000 COP). Otras ofertas (p. ej. CONTINENTAL GOLD) lo
        # tomaron por error. Reglas:
        #   1) Si existe la fila de TRANSBORDER → esa queda ACTIVA, las demás anuladas.
        #   2) Si NO existe TRANSBORDER → NUNCA dejar el 261161 sin oferta activa:
        #      se reactiva una sola fila (para que no "desaparezca" el número).
        # El índice único parcial exige que quede EXACTAMENTE una activa. Es idempotente.
        try:
            cur.execute("""
                SELECT id, COALESCE(cliente,'') AS cliente, COALESCE(anulada,false) AS anulada
                  FROM ofertas
                 WHERE num ~ '^[0-9]+$' AND CAST(num AS INTEGER) = 261161
                 ORDER BY id
            """)
            _r261161 = cur.fetchall() or []
            if _r261161:
                _tb = [x for x in _r261161 if str(x[1]).strip().upper().startswith("TRANSBORDER")]
                if _tb:
                    _keep = _tb[0][0]          # preferimos la de TRANSBORDER
                    _motivo = "Consecutivo duplicado: el 261161 pertenece a TRANSBORDER."
                else:
                    # No hay TRANSBORDER: conservamos UNA para no borrar el número.
                    # Preferimos una ya activa; si todas están anuladas, la más reciente.
                    _activas = [x for x in _r261161 if not x[2]]
                    _keep = (_activas[-1][0] if _activas else _r261161[-1][0])
                    _motivo = "Consecutivo duplicado 261161: se conserva una sola oferta activa."
                # 1) Anular todas las demás (libera el candado único).
                for _x in _r261161:
                    if _x[0] != _keep and not _x[2]:
                        cur.execute("""
                            UPDATE ofertas
                               SET anulada = true, estado = 'ANULADA',
                                   anulada_motivo = %s,
                                   anulada_por = 'Sistema', anulada_fecha = now()
                             WHERE id = %s
                        """, (_motivo, _x[0]))
                        print(f"[DB][261161] Anulada id={_x[0]} cliente='{_x[1]}'.")
                # 2) Reactivar SIEMPRE la que conservamos (nunca queda anulada).
                cur.execute("""
                    UPDATE ofertas
                       SET anulada = false,
                           estado = CASE WHEN estado = 'ANULADA' THEN 'ENVIADO' ELSE estado END,
                           anulada_motivo = NULL, anulada_por = NULL, anulada_fecha = NULL
                     WHERE id = %s
                """, (_keep,))
                print(f"[DB][261161] OK: activa id={_keep} (TRANSBORDER preferida).")
            else:
                print("[DB][261161] No existe ninguna fila 261161; nada que corregir.")
        except Exception as e:
            print(f"[DB][261161] No se pudo corregir: {e}")

        print("[DB] Tabla 'ofertas' lista.")

        # Tabla SEPARADA para las ofertas 2025 (histórico facturado).
        # No se mezcla con 'ofertas' (2026) para no afectar la operación viva.
        cur.execute("""
            CREATE TABLE IF NOT EXISTS ofertas_2025 (
                id bigint generated always as identity primary key,
                num text unique not null,
                num_orig text,
                mes text,
                cliente text,
                responsable text,
                descripcion text,
                valor_facturado bigint DEFAULT 0,
                no_factura text,
                fecha_facturacion date,
                seguimiento text DEFAULT 'Facturada',
                created_at timestamptz DEFAULT now()
            )
        """)
        print("[DB] Tabla 'ofertas_2025' lista.")

        # Tabla SEPARADA para los contratos (facturas sin número de oferta),
        # agrupados por cliente. Fuente: hoja "Facturación".
        cur.execute("""
            CREATE TABLE IF NOT EXISTS contratos (
                id bigint generated always as identity primary key,
                cliente text unique not null,
                valor_facturado bigint DEFAULT 0,
                n_facturas int DEFAULT 0,
                no_factura text,
                fecha_facturacion date,
                created_at timestamptz DEFAULT now()
            )
        """)
        print("[DB] Tabla 'contratos' lista.")

        # Facturas SIN número de oferta, categorizadas y agrupadas por cliente:
        # CONTRATO / 2025_ANO_PASADO / OTROS. Reemplaza el uso de 'contratos'.
        cur.execute("""
            CREATE TABLE IF NOT EXISTS facturacion_cat (
                id bigint generated always as identity primary key,
                categoria text NOT NULL,
                cliente text NOT NULL,
                valor_facturado bigint DEFAULT 0,
                n_facturas int DEFAULT 0,
                no_factura text,
                fecha_facturacion date,
                created_at timestamptz DEFAULT now(),
                UNIQUE (categoria, cliente)
            )
        """)
        print("[DB] Tabla 'facturacion_cat' lista.")

        # Catálogo maestro de CLIENTES: fuente única de verdad para el nombre
        # con el que se guardan las ofertas. Evita duplicados/errores de tipeo
        # (auditoría "cero errores"). 'nombre_corto' es el nombre comercial que
        # se muestra y se guarda en ofertas.cliente; razon_social + nit son los
        # datos oficiales (hoja Facturación).
        cur.execute("""
            CREATE TABLE IF NOT EXISTS clientes (
                id bigint generated always as identity primary key,
                nombre_corto text unique not null,
                razon_social text,
                nit text,
                activo boolean not null default true,
                created_at timestamptz DEFAULT now()
            )
        """)
        # Índice único case-insensitive para que "Tiba" y "TIBA" no coexistan.
        cur.execute("""
            CREATE UNIQUE INDEX IF NOT EXISTS clientes_nombre_lower_idx
            ON clientes (lower(nombre_corto))
        """)
        # Semilla idempotente: todo cliente que ya exista en ofertas queda en el
        # catálogo. Así el selector siempre tiene la lista real y unificada.
        cur.execute("""
            INSERT INTO clientes (nombre_corto)
            SELECT DISTINCT TRIM(cliente) FROM ofertas
            WHERE cliente IS NOT NULL AND TRIM(cliente) <> ''
            ON CONFLICT DO NOTHING
        """)
        print("[DB] Tabla 'clientes' lista y sembrada.")

        # Tabla ESPEJO de la facturación real de VULCANO (sistema externo donde
        # Natalia factura). Se llena importando el Excel que ella descarga de
        # Vulcano. 'excluida' marca las facturas que NO cuentan al total real
        # (p. ej. las que contabilidad marca en amarillo). Fuente autoritativa
        # para conciliar el total de la app contra contabilidad ("cero errores").
        cur.execute("""
            CREATE TABLE IF NOT EXISTS vulcano_facturas (
                id bigint generated always as identity primary key,
                factura text unique not null,
                fecha date,
                mes text,
                anio text,
                estado text,
                nit text,
                cliente text,
                subtotal bigint DEFAULT 0,
                total bigint DEFAULT 0,
                valor_pagado bigint DEFAULT 0,
                saldo bigint DEFAULT 0,
                excluida boolean not null default false,
                importado_at timestamptz DEFAULT now()
            )
        """)
        # Migración: N° de oferta y clasificación de origen (2026/2025/CONTRATO)
        # para el panel de Proyección (Bloque 1 "Facturado real").
        cur.execute("ALTER TABLE vulcano_facturas ADD COLUMN IF NOT EXISTS oferta_ref text")
        cur.execute("ALTER TABLE vulcano_facturas ADD COLUMN IF NOT EXISTS clase text")
        print("[DB] Tabla 'vulcano_facturas' lista.")

        # Contratos: pendiente por facturar (Bloque 3 de la Proyección).
        # Líneas manuales que digita Natalia cuando proyectos confirma un
        # servicio de contrato por WhatsApp y Vulcano aún no lo factura.
        cur.execute("""
            CREATE TABLE IF NOT EXISTS contratos_pendientes (
                id bigserial primary key,
                cliente text NOT NULL,
                descripcion text,
                valor bigint NOT NULL DEFAULT 0,
                mes text,
                estado text NOT NULL DEFAULT 'CONFIRMADO',
                oferta_ref text,
                creado_por text,
                created_at timestamptz DEFAULT now(),
                updated_at timestamptz DEFAULT now()
            )
        """)
        print("[DB] Tabla 'contratos_pendientes' lista.")

        # Presupuesto / META mensual de facturación por año. Fuente: Natalia.
        # Sirve para medir el % de cumplimiento (facturado real vs meta).
        cur.execute("""
            CREATE TABLE IF NOT EXISTS presupuesto (
                id bigint generated always as identity primary key,
                anio int NOT NULL,
                mes text NOT NULL,
                monto bigint NOT NULL DEFAULT 0,
                UNIQUE (anio, mes)
            )
        """)
        # Semilla idempotente del presupuesto 2026 (no pisa ediciones posteriores).
        _presu_2026 = [
            ("ENERO", 1250000000), ("FEBRERO", 1450000000), ("MARZO", 1550000000),
            ("ABRIL", 1650000000), ("MAYO", 1850000000), ("JUNIO", 1900000000),
            ("JULIO", 1850000000), ("AGOSTO", 1800000000), ("SEPTIEMBRE", 1600000000),
            ("OCTUBRE", 1650000000), ("NOVIEMBRE", 1450000000), ("DICIEMBRE", 1500000000),
        ]
        for _m, _v in _presu_2026:
            cur.execute("""
                INSERT INTO presupuesto (anio, mes, monto) VALUES (2026, %s, %s)
                ON CONFLICT (anio, mes) DO NOTHING
            """, (_m, _v))
        print("[DB] Tabla 'presupuesto' lista y sembrada (2026).")

        cur.execute("""
            CREATE TABLE IF NOT EXISTS usuarios (
                id        bigserial primary key,
                username  varchar(50) unique not null,
                nombre    varchar(100) not null,
                password_hash text not null,
                rol       varchar(20) not null default 'viewer',
                activo    boolean not null default true,
                creado_en timestamptz default now()
            )
        """)
        _admin_hash = _hash_pw("Boom2025*")
        cur.execute("""
            INSERT INTO usuarios (username, nombre, password_hash, rol)
            VALUES ('admin', 'Administrador', %s, 'admin')
            ON CONFLICT (username) DO NOTHING
        """, (_admin_hash,))
        # Migración: área en usuarios
        cur.execute("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS area varchar(60)")
        cur.execute("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS modulos text[] DEFAULT '{}'::text[]")

        cur.execute("""
            CREATE TABLE IF NOT EXISTS areas (
                id        bigserial primary key,
                nombre    varchar(100) unique not null,
                descripcion text,
                icono     varchar(10) default '🏢',
                activo    boolean not null default true,
                created_at timestamptz default now()
            )
        """)
        cur.execute("""
            INSERT INTO areas (nombre, descripcion, icono)
            VALUES ('Comercial', 'Área comercial — gestión de ofertas', '💼')
            ON CONFLICT (nombre) DO NOTHING
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS area_permisos (
                area_id   bigint references areas(id) on delete cascade,
                modulo    varchar(50) not null,
                activo    boolean not null default true,
                primary key (area_id, modulo)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS notificaciones (
                id          bigserial primary key,
                oferta_id   bigint,
                oferta_num  text,
                cliente     text,
                origen      text,
                destino     text,
                valor       bigint default 0,
                leida       boolean default false,
                created_at  timestamptz default now()
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS osi (
                id           bigserial primary key,
                numero_osi   text unique not null,
                oferta_id    bigint,
                oferta_num   text,
                fecha        date default current_date,
                responsable  text,
                equipo       text,
                cliente      text,
                origen       text,
                destino      text,
                valor        bigint default 0,
                estado       text default 'PROGRAMADO',
                notas        text,
                created_at   timestamptz default now()
            )
        """)
        # Feature 3: OSI nuevas columnas
        cur.execute("ALTER TABLE osi ADD COLUMN IF NOT EXISTS fecha_despacho date")
        cur.execute("ALTER TABLE osi ADD COLUMN IF NOT EXISTS conductor text")
        cur.execute("ALTER TABLE osi ADD COLUMN IF NOT EXISTS placa text")
        cur.execute("ALTER TABLE osi ADD COLUMN IF NOT EXISTS observaciones text")

        # Catálogo de equipos (módulo Operaciones) — propios + subcontratos
        cur.execute("""
            CREATE TABLE IF NOT EXISTS equipos (
                id          bigserial primary key,
                placa       text,
                tipo        text,
                propiedad   text default 'propio',
                tenedor     text,
                conductor   text,
                celular     text,
                estado      text default 'DISPONIBLE',
                created_at  timestamptz default now()
            )
        """)
        # Ampliación flota propia (Fase 1 módulo Operaciones): datos de ficha del equipo.
        for _col, _ddl in [
            ("codigo",      "text"),           # código interno (EXV241, MOD01, GN01…)
            ("categoria",   "text"),           # CABEZOTE / CAMA_ALTA / CAMA_BAJA / SEMI / MODULAR / ACCESORIO
            ("marca",       "text"),           # KENWORTH, SCANIA, GOLDHOFER…
            ("clase",       "text"),           # Extensible, Tecnipesado, Challenger, MODULAR…
            ("config",      "text"),           # CA2, CA3, CB3, CB4, SM5, M6…
            ("ejes",        "text"),           # nº de ejes / líneas
            ("ancho",       "text"),           # dimensiones (texto: admite rangos "14,73 - 23,15")
            ("largo",       "text"),
            ("alto",        "text"),
            ("capacidad",   "text"),
            ("peso",        "text"),
            ("descripcion", "text"),
            ("activo",      "boolean default true"),
        ]:
            cur.execute("ALTER TABLE equipos ADD COLUMN IF NOT EXISTS %s %s" % (_col, _ddl))

        # Feature 4: Historial de cambios
        cur.execute("""
            CREATE TABLE IF NOT EXISTS oferta_historial (
                id          bigserial primary key,
                oferta_id   bigint,
                oferta_num  text,
                campo       text,
                valor_ant   text,
                valor_nuevo text,
                usuario     text,
                created_at  timestamptz default now()
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS sessions (
                token      text primary key,
                user_id    bigint not null,
                username   text not null,
                nombre     text not null,
                rol        text not null,
                expires_at timestamptz not null,
                created_at timestamptz default now()
            )
        """)
        # Versiones del documento de una oferta (re-cotizaciones): cada guardado
        # desde el asistente crea una versión. Mantiene el MISMO número de oferta.
        cur.execute("""
            CREATE TABLE IF NOT EXISTS oferta_versiones (
                id          bigserial primary key,
                oferta_id   bigint not null,
                oferta_num  text,
                version     int not null default 1,
                valor       bigint default 0,
                moneda      text default 'COP',
                descripcion text,
                forma_pago  text,
                html        text,        -- documento de la oferta (HTML)
                pdf_b64     text,        -- PDF original subido (ej. v1 externa)
                resumen     text,        -- qué cambió en esta versión
                creado_por  text,
                vigente     boolean default true,
                created_at  timestamptz default now()
            )
        """)
        cur.execute("CREATE INDEX IF NOT EXISTS idx_oferta_versiones_of ON oferta_versiones(oferta_id)")
        # Metadatos de la app (clave/valor). Se usa para guardar el "sello de versión"
        # desplegada y así forzar re-login cuando se sube una versión nueva.
        cur.execute("""
            CREATE TABLE IF NOT EXISTS app_meta (
                clave text primary key,
                valor text
            )
        """)
        conn.close()
        print("[DB] Tablas 'areas', 'area_permisos', 'notificaciones', 'osi', 'oferta_historial', 'oferta_versiones' y 'sessions' listas.")
    except Exception as e:
        print(f"[DB] Error: {e}")
        raise

_ensure_db()

# ── Session helpers (DB-backed + memory cache) ────────────────────────────────
_sessions: dict = {}   # token -> {id, username, nombre, rol}

def _session_load_from_db():
    """Load all non-expired sessions from DB into memory cache on startup."""
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT s.token, s.user_id, s.username, s.nombre, s.rol, u.modulos "
                "FROM sessions s LEFT JOIN usuarios u ON u.id = s.user_id "
                "WHERE s.expires_at > now()"
            )
            rows = fetchall(cur)
        for r in rows:
            _sessions[r["token"]] = {
                "id": r["user_id"], "username": r["username"],
                "nombre": r["nombre"], "rol": r["rol"],
                "modulos": r.get("modulos") or [],
            }
        print(f"[AUTH] {len(rows)} sesión(es) activa(s) cargada(s) desde DB.")
    except Exception as e:
        print(f"[AUTH] No se pudieron cargar sesiones desde DB: {e}")

def _session_save(token: str, user: dict, max_age_s: int = 86400 * 7):
    _sessions[token] = user
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute(
                """INSERT INTO sessions (token, user_id, username, nombre, rol, expires_at)
                   VALUES (%s, %s, %s, %s, %s, now() + make_interval(secs => %s))
                   ON CONFLICT (token) DO UPDATE
                   SET expires_at = now() + make_interval(secs => %s)""",
                (token, user["id"], user["username"], user["nombre"], user["rol"],
                 max_age_s, max_age_s)
            )
    except Exception as e:
        print(f"[AUTH] Error guardando sesión en DB: {e}")

# ── Cierre por inactividad ────────────────────────────────────────────────────
# La sesión se cierra tras 2 horas SIN actividad real del usuario (mouse/teclado/
# clics). Los refrescos automáticos en segundo plano NO cuentan como actividad;
# solo el "ping" que envía el navegador ante interacción real renueva el reloj.
INACTIVITY_LIMIT_S = 2 * 3600
_session_activity: dict = {}   # token -> epoch de la última actividad real

def _activity_touch(token: str):
    if token:
        _session_activity[token] = time.time()

def _activity_expired(token: str) -> bool:
    """True si la sesión superó el límite de inactividad. Si no hay marca previa
    (p. ej. tras un reinicio del servidor), la inicializa como activa ahora."""
    if not token:
        return False
    last = _session_activity.get(token)
    if last is None:
        _session_activity[token] = time.time()   # gracia: primera vez que la vemos
        return False
    return (time.time() - last) > INACTIVITY_LIMIT_S

def _session_delete(token: str):
    _sessions.pop(token, None)
    _session_activity.pop(token, None)
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("DELETE FROM sessions WHERE token = %s", (token,))
    except Exception as e:
        print(f"[AUTH] Error eliminando sesión de DB: {e}")

def _session_get(token: str) -> dict | None:
    """Check memory first, then DB (and repopulate memory on DB hit)."""
    if not token:
        return None
    if token in _sessions:
        return _sessions[token]
    # Fallback: look up DB (handles restarts where memory was cleared)
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT s.user_id, s.username, s.nombre, s.rol, u.modulos "
                "FROM sessions s LEFT JOIN usuarios u ON u.id = s.user_id "
                "WHERE s.token = %s AND s.expires_at > now()",
                (token,)
            )
            row = fetchone(cur)
        if row:
            user = {"id": row["user_id"], "username": row["username"],
                    "nombre": row["nombre"], "rol": row["rol"],
                    "modulos": row.get("modulos") or []}
            _sessions[token] = user   # repopulate cache
            return user
    except Exception as e:
        print(f"[AUTH] Error consultando sesión en DB: {e}")
    return None

# ── Cierre de sesiones al desplegar una versión nueva ─────────────────────────
# Cada deploy en Railway trae un identificador distinto (commit o deployment id).
# Si cambia respecto al último guardado, se borran TODAS las sesiones activas para
# que todos los usuarios vuelvan a ingresar con usuario y contraseña.
APP_BUILD = (
    os.environ.get("RAILWAY_GIT_COMMIT_SHA")
    or os.environ.get("RAILWAY_DEPLOYMENT_ID")
    or "dev"
)

def _enforce_build_logout():
    if APP_BUILD == "dev":
        return  # entorno local: no cerramos sesiones
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT valor FROM app_meta WHERE clave = 'build'")
            row = fetchone(cur)
            stored = row["valor"] if row else None
            if stored != APP_BUILD:
                cur.execute("DELETE FROM sessions")
                cur.execute(
                    "INSERT INTO app_meta (clave, valor) VALUES ('build', %s) "
                    "ON CONFLICT (clave) DO UPDATE SET valor = EXCLUDED.valor",
                    (APP_BUILD,)
                )
                _sessions.clear()
                _session_activity.clear()
                print(f"[AUTH] Nueva versión desplegada ({APP_BUILD[:12]}). "
                      f"Sesiones cerradas: todos deben volver a ingresar.")
    except Exception as e:
        print(f"[AUTH] No se pudo aplicar el cierre por nueva versión: {e}")

_enforce_build_logout()
_session_load_from_db()


# ── Email notification ────────────────────────────────────────────────────────
def _enviar_notificacion_osi(oferta_row: dict, pdf_data: dict):
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText

    if not SMTP_USER or not SMTP_PASSWORD:
        print("[EMAIL] SMTP_USER/SMTP_PASSWORD no configurados — email omitido")
        return

    num     = oferta_row.get("num", "")
    cliente = oferta_row.get("cliente", "")
    valor   = oferta_row.get("valor", 0) or 0
    origen  = pdf_data.get("origen", "") if pdf_data else ""
    destino = pdf_data.get("destino", "") if pdf_data else ""
    equipos = pdf_data.get("equipos", []) if pdf_data else []
    cargo_items = pdf_data.get("cargo_items", []) if pdf_data else []
    notas   = pdf_data.get("notas", "") if pdf_data else ""
    forma_pago = pdf_data.get("forma_pago", "50% anticipo / 50% a 30 días tras radicación de factura") if pdf_data else ""
    vigencia   = pdf_data.get("vigencia", 30) if pdf_data else 30

    # ── Tabla carga
    if cargo_items:
        rows_c = "".join(f"""<tr><td style='padding:6px 10px;border-bottom:1px solid #e5e5e5;'>{c.get('descripcion','')}</td>
          <td style='padding:6px 10px;border-bottom:1px solid #e5e5e5;text-align:center;'>{c.get('cant',1)}</td>
          <td style='padding:6px 10px;border-bottom:1px solid #e5e5e5;text-align:center;'>{c.get('dimensiones','')}</td>
          <td style='padding:6px 10px;border-bottom:1px solid #e5e5e5;text-align:center;'>{c.get('peso','')}</td></tr>"""
            for c in cargo_items if c.get('descripcion'))
        tabla_carga = f"""<table style='width:100%;border-collapse:collapse;font-size:13px;margin-top:8px;'>
          <thead><tr style='background:#1B2A4A;color:#fff;'>
            <th style='padding:7px 10px;text-align:left;'>Descripción</th>
            <th style='padding:7px 10px;'>Cant.</th>
            <th style='padding:7px 10px;'>Dimensiones</th>
            <th style='padding:7px 10px;'>Peso</th>
          </tr></thead><tbody>{rows_c}</tbody></table>"""
    else:
        tabla_carga = "<p style='color:#666;font-size:13px;'>Ver detalle en oferta adjunta.</p>"

    # ── Tabla equipos
    total = sum(int(e.get("cant",1) or 1) * int(e.get("valor_unit",0) or 0) for e in equipos)
    rows_e = "".join(f"""<tr><td style='padding:6px 10px;border-bottom:1px solid #e5e5e5;'>{e.get('equipo','')}</td>
      <td style='padding:6px 10px;border-bottom:1px solid #e5e5e5;text-align:center;'>{e.get('cant',1)}</td>
      <td style='padding:6px 10px;border-bottom:1px solid #e5e5e5;'>{e.get('config','')}</td>
      <td style='padding:6px 10px;border-bottom:1px solid #e5e5e5;text-align:right;font-weight:bold;'>{_fmt_cop(int(e.get('cant',1) or 1)*int(e.get('valor_unit',0) or 0))}</td></tr>"""
        for e in equipos if e.get('equipo'))
    total_row = f"""<tr style='background:#E8601C;color:#fff;'><td colspan='3' style='padding:8px 10px;font-weight:bold;'>TOTAL OFERTA</td>
      <td style='padding:8px 10px;text-align:right;font-weight:bold;'>{_fmt_cop(total)} COP</td></tr>"""
    tabla_equipos = f"""<table style='width:100%;border-collapse:collapse;font-size:13px;margin-top:8px;'>
      <thead><tr style='background:#1B2A4A;color:#fff;'>
        <th style='padding:7px 10px;text-align:left;'>Equipo</th>
        <th style='padding:7px 10px;'>Cant.</th>
        <th style='padding:7px 10px;text-align:left;'>Esquema de Seguridad</th>
        <th style='padding:7px 10px;text-align:right;'>Tarifa Neto (COP)</th>
      </tr></thead><tbody>{rows_e}{total_row}</tbody></table>"""

    # ── Stand-by desde notas
    sb_lines = [l.strip() for l in notas.split("\n") if "stand-by" in l.lower()]
    sb_html = "".join(f"<li style='margin-bottom:4px;'>{l}</li>" for l in sb_lines)
    if sb_html:
        sb_html = f"<ul style='margin:6px 0 0 0;padding-left:18px;font-size:13px;'>{sb_html}</ul>"

    # ── Condiciones operativas desde notas
    op_lines = [l.strip() for l in notas.split("\n")
                if l.strip() and "stand-by" not in l.lower()
                and not l.strip().lower().startswith(("origen:","destino:"))]
    op_html = "".join(f"<li style='margin-bottom:4px;'>{l}</li>" for l in op_lines)
    if op_html:
        op_html = f"<ul style='margin:6px 0 0 0;padding-left:18px;font-size:13px;'>{op_html}</ul>"

    html_body = f"""<!DOCTYPE html><html><head><meta charset='UTF-8'></head>
<body style='font-family:Arial,sans-serif;font-size:14px;color:#1B2A4A;margin:0;padding:20px;background:#f4f4f4;'>
<div style='max-width:700px;margin:0 auto;background:#fff;border-radius:6px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,.1);'>
  <div style='background:#1B2A4A;padding:14px 20px;display:flex;align-items:center;justify-content:space-between;'>
    <span style='color:#fff;font-weight:700;font-size:18px;letter-spacing:1px;'>BOOM</span>
    <div style='color:#fff;font-size:11px;text-align:right;'><strong style='font-size:12px;display:block;'>BOOM LOGISTICS COLOMBIA S.A.S.</strong>Soluciones de Transporte Especializado</div>
  </div>
  <div style='background:#E8601C;padding:8px 20px;'>
    <p style='margin:0;color:#fff;font-size:11px;font-weight:bold;'>🔔 NOTIFICACIÓN DE INICIO DE SERVICIO &nbsp;|&nbsp; REF: {num} &nbsp;|&nbsp; {cliente.upper()} &nbsp;|&nbsp; {(origen+' → '+destino).upper() if origen and destino else ''}</p>
  </div>
  <div style='padding:24px;'>
    <p style='margin:0 0 16px 0;'>Cordial Saludo,</p>
    <p style='margin:0 0 20px 0;'>Para conocimiento de todos, a continuación notificamos lo siguiente:</p>

    <div style='background:#f0f4ff;border-left:4px solid #1B2A4A;padding:12px 16px;border-radius:4px;margin-bottom:20px;'>
      <p style='margin:0 0 4px 0;font-weight:700;font-size:15px;'>📄 NOTIFICACIÓN DE INICIO DE SERVICIO</p>
      <p style='margin:2px 0;font-size:13px;'><strong>PARA:</strong> Operaciones, Seguimiento y Control Documental</p>
      <p style='margin:2px 0;font-size:13px;'><strong>REF:</strong> OFERTA MERCANTIL No. {num}</p>
      <p style='margin:2px 0;font-size:13px;'><strong>CLIENTE:</strong> {cliente.upper()}</p>
      <p style='margin:2px 0;font-size:13px;'><strong>ESTADO:</strong> 🟢 PROGRAMADO</p>
    </div>

    <p style='background:#1B2A4A;color:#fff;font-weight:700;font-size:12px;padding:7px 12px;border-radius:3px;margin:16px 0 8px 0;'>1. INFORMACIÓN DEL PROYECTO</p>
    <ul style='margin:0;padding-left:18px;font-size:13px;line-height:1.8;'>
      <li><strong>Origen:</strong> {origen or "A confirmar con cliente"}</li>
      <li><strong>Destino:</strong> {destino or "A confirmar con cliente"}</li>
      <li><strong>Fecha de cargue:</strong> A confirmar con cliente</li>
    </ul>

    <p style='background:#1B2A4A;color:#fff;font-weight:700;font-size:12px;padding:7px 12px;border-radius:3px;margin:16px 0 8px 0;'>2. DETALLE TÉCNICO DE LA CARGA</p>
    {tabla_carga}

    <p style='background:#1B2A4A;color:#fff;font-weight:700;font-size:12px;padding:7px 12px;border-radius:3px;margin:16px 0 8px 0;'>3. CONFIGURACIÓN DE TRANSPORTE Y TARIFA</p>
    {tabla_equipos}

    <p style='background:#1B2A4A;color:#fff;font-weight:700;font-size:12px;padding:7px 12px;border-radius:3px;margin:16px 0 8px 0;'>4. CONDICIONES OPERATIVAS Y SEGUROS</p>
    {sb_html}{op_html if op_html else "<p style='font-size:13px;color:#666;margin:6px 0 0 0;'>Póliza de carga y RCE hasta $4.000.000.000 COP. GPS incluido.</p>"}

    <p style='background:#1B2A4A;color:#fff;font-weight:700;font-size:12px;padding:7px 12px;border-radius:3px;margin:16px 0 8px 0;'>5. CONDICIONES COMERCIALES</p>
    <ul style='margin:0;padding-left:18px;font-size:13px;line-height:1.8;'>
      <li><strong>Forma de pago:</strong> {forma_pago}</li>
      <li><strong>Vigencia:</strong> {vigencia} días calendario</li>
    </ul>

    <p style='margin:20px 0 0 0;border-top:1px solid #e0e0e0;padding-top:14px;font-size:13px;color:#666;'>Lo anterior para lo correspondiente.</p>
    <p style='margin:8px 0 2px 0;font-weight:bold;color:#1B2A4A;'>Natalia Vargas</p>
    <p style='margin:2px 0;color:#E8601C;font-size:13px;'>Ejecutiva Comercial</p>
    <p style='margin:2px 0;font-size:12px;color:#666;'>BOOM Logistics Colombia S.A.S.</p>
  </div>
</div>
</body></html>"""

    msg = MIMEMultipart("alternative")
    msg["Subject"] = f"🔔 NOTIFICACIÓN DE INICIO DE SERVICIO — REF: {num} — {cliente.upper()}"
    msg["From"]    = SMTP_FROM
    msg["To"]      = ", ".join(NOTIF_TO)
    msg["Cc"]      = ", ".join(NOTIF_CC)
    msg.attach(MIMEText(html_body, "html", "utf-8"))

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=15) as server:
            server.ehlo()
            server.starttls()
            server.login(SMTP_USER, SMTP_PASSWORD)
            all_recipients = NOTIF_TO + NOTIF_CC
            server.sendmail(SMTP_FROM, all_recipients, msg.as_string())
        print(f"[EMAIL] Notificación OSI enviada para oferta {num}")
    except Exception as exc:
        print(f"[EMAIL] Error al enviar notificación: {exc}")


def new_conn():
    conn = pgdb.connect(host=DB_HOST, port=DB_PORT, database=DB_NAME,
                         user=DB_USER, password=DB_PASSWORD)
    conn.autocommit = True
    return conn


@contextmanager
def get_conn():
    conn = new_conn()
    try:
        yield conn
    finally:
        conn.close()


def _cols(cursor):
    return [col[0] for col in cursor.description]


def fetchall(cursor):
    if not cursor.description:
        return []
    cols = _cols(cursor)
    return [_serialize(dict(zip(cols, row))) for row in cursor.fetchall()]


def fetchone(cursor):
    if not cursor.description:
        return None
    cols = _cols(cursor)
    row = cursor.fetchone()
    return _serialize(dict(zip(cols, row))) if row is not None else None


def _serialize(d: dict) -> dict:
    return {k: v.isoformat() if isinstance(v, (date, datetime)) else v
            for k, v in d.items()}


# ── App ───────────────────────────────────────────────────────────────────────
app = FastAPI(title="BOOM Logistics - Control de Ofertas")

_AUTH_PUBLIC = {"", "/", "/manual", "/anexo-legal", "/auth/login", "/auth/logout", "/auth/me", "/api/logo", "/api/_diagtiba2"}
_WRITE_METHODS = {"POST", "PUT", "DELETE", "PATCH"}
# Rutas de escritura del módulo OPERACIONES (OSI, equipos, alertas). Un usuario
# 'viewer' que tenga el módulo 'operaciones' puede ESCRIBIR sólo aquí (crear/editar
# OSI, agregar equipos, atender alertas); en todo lo demás (ofertas, etc.) sigue
# siendo de solo lectura.
_OPERACIONES_WRITE_PREFIXES = ("/api/osi", "/api/equipos", "/api/notificaciones")


@app.middleware("http")
async def auth_middleware(request: Request, call_next):
    path = request.url.path
    if path in _AUTH_PUBLIC:
        return await call_next(request)

    token = request.cookies.get("boom_session")
    user = _session_get(token)

    if not user:
        return JSONResponse({"detail": "No autenticado"}, status_code=401)

    # Cierre por inactividad (2h). Los refrescos automáticos pasan por aquí pero
    # NO renuevan el reloj; solo lo comprueban. El reloj se renueva vía /auth/ping.
    if _activity_expired(token):
        _session_delete(token)
        return JSONResponse({"detail": "Sesión cerrada por inactividad"}, status_code=401)

    request.state.user = user

    if user["rol"] == "viewer" and request.method in _WRITE_METHODS:
        # Excepción: un viewer con el módulo 'operaciones' puede editar SÓLO en
        # las rutas de Operaciones. Fuera de ahí sigue siendo de solo lectura.
        puede_operaciones = (
            "operaciones" in (user.get("modulos") or [])
            and path.startswith(_OPERACIONES_WRITE_PREFIXES)
        )
        if not puede_operaciones:
            return JSONResponse({"detail": "Acceso de solo lectura"}, status_code=403)

    if path.startswith("/api/usuarios") and request.method in _WRITE_METHODS and user["rol"] != "admin":
        return JSONResponse({"detail": "Solo administradores pueden gestionar usuarios"}, status_code=403)

    return await call_next(request)


# ── Pydantic models ───────────────────────────────────────────────────────────
class OfertaCreate(BaseModel):
    num: Optional[str] = None
    mes: Optional[str] = None
    fecha: Optional[str] = None
    cliente: Optional[str] = None
    realizada: Optional[str] = None
    formalizada: Optional[str] = None
    unidad: Optional[str] = None
    tipo: Optional[str] = None
    sector: Optional[str] = None
    valor: Optional[int] = 0
    estado: Optional[str] = "ENVIADO"
    respuesta: Optional[str] = None
    facturacion: Optional[str] = None
    general: Optional[str] = None
    seguimiento: Optional[str] = None
    mes_aceptado: Optional[str] = None
    fecha_facturacion: Optional[str] = None
    valor_facturado: Optional[int] = None
    no_factura: Optional[str] = None
    valor_aprobado: Optional[int] = None
    pdf_data: Optional[dict] = None


class OfertaUpdate(BaseModel):
    mes: Optional[str] = None
    fecha: Optional[str] = None
    cliente: Optional[str] = None
    realizada: Optional[str] = None
    formalizada: Optional[str] = None
    unidad: Optional[str] = None
    tipo: Optional[str] = None
    sector: Optional[str] = None
    valor: Optional[int] = None
    estado: Optional[str] = None
    respuesta: Optional[str] = None
    facturacion: Optional[str] = None
    general: Optional[str] = None
    seguimiento: Optional[str] = None
    mes_aceptado: Optional[str] = None
    fecha_facturacion: Optional[str] = None
    valor_facturado: Optional[int] = None
    no_factura: Optional[str] = None
    valor_aprobado: Optional[int] = None
    pdf_data: Optional[dict] = None
    costo_proyecto: Optional[int] = None


class EquipoItem(BaseModel):
    equipo: Optional[str] = ""
    dimensiones: Optional[str] = ""
    cant: Optional[int] = 0
    config: Optional[str] = ""
    valor_unit: Optional[int] = 0


class CargoItem(BaseModel):
    descripcion: Optional[str] = ""
    tipo: Optional[str] = ""
    cant: Optional[int] = 1
    dimensiones: Optional[str] = ""
    peso: Optional[str] = ""
    volumen: Optional[str] = ""
    origen_detalle: Optional[str] = ""
    destino_detalle: Optional[str] = ""


class OfertaHtml(BaseModel):
    ref: Optional[str] = None
    cliente: Optional[str] = None
    contacto: Optional[str] = None
    email_cliente: Optional[str] = None
    ref_cliente: Optional[str] = None
    cliente_final: Optional[str] = None
    origen: Optional[str] = None
    destino: Optional[str] = None
    mes_anio: Optional[str] = None
    descripcion: Optional[str] = None
    comercial: Optional[str] = "Natalia Vargas"
    cargo_items: Optional[List[CargoItem]] = []
    equipos: Optional[List[EquipoItem]] = []
    notas: Optional[str] = ""
    forma_pago: Optional[str] = "50% anticipo / 50% a 30 días tras radicación de factura"
    vigencia: Optional[int] = 30
    poliza_carga: Optional[str] = "Hasta $4.000.000.000 COP por despacho"
    poliza_rc: Optional[str] = "Hasta $4.000.000.000 COP"
    resolucion: Optional[str] = "Carga extradimensionada/extrapesada incluida"
    exclusiones: Optional[str] = (
        "Permisos de tránsito, operación y pólizas asociadas (a cargo del cliente)\n"
        "Servicios, recursos o actividades no descritos explícitamente en esta oferta"
    )
    fotos: Optional[List[str]] = []  # data URIs (base64) de fotos de la carga


class LoginBody(BaseModel):
    username: str
    password: str


class UsuarioCreate(BaseModel):
    username: str
    nombre: str
    password: str
    rol: str = "viewer"
    modulos: Optional[List[str]] = None


class UsuarioUpdate(BaseModel):
    nombre: Optional[str] = None
    password: Optional[str] = None
    rol: Optional[str] = None
    activo: Optional[bool] = None
    modulos: Optional[List[str]] = None


class ClienteCreate(BaseModel):
    nombre_corto: str
    razon_social: Optional[str] = None
    nit: Optional[str] = None


class ClienteUpdate(BaseModel):
    nombre_corto: Optional[str] = None
    razon_social: Optional[str] = None
    nit: Optional[str] = None


class VulcanoExcluir(BaseModel):
    excluida: bool


class PresupuestoItem(BaseModel):
    anio: int
    mes: str
    monto: int


class TextoCliente(BaseModel):
    texto: str


class ChatMsg(BaseModel):
    role: str
    content: Union[str, List[Any]]


class ChatOfertaBody(BaseModel):
    messages: List[ChatMsg]
    estado_actual: Optional[str] = None


class ParseDetalle(BaseModel):
    texto: str


class AutoNotasRequest(BaseModel):
    equipos: Optional[List[EquipoItem]] = []
    texto_cliente: Optional[str] = ""
    origen: Optional[str] = ""
    destino: Optional[str] = ""


# ── Claude API extraction ─────────────────────────────────────────────────────
BOOM_SYSTEM_PROMPT = """ERES EL ASISTENTE DE GENERACIÓN DE OFERTAS COMERCIALES DE BOOM LOGISTICS COLOMBIA S.A.S.
Trabajas directamente con Natalia Vargas (Naty), Ejecutiva Comercial. Respondes SIEMPRE en español.

PRINCIPIO: interpreta directamente como lo hace Naty y aplica los estándares BOOM. No pidas
confirmaciones innecesarias. PERO sí pregunta en el chat cuando falte información CRÍTICA
(ver "CUÁNDO GENERAR / CUÁNDO PREGUNTAR").

═══ IDENTIFICACIÓN DEL CLIENTE ═══
Antes de titular la oferta, confirma el cliente por la FIRMA del correo (nombre, empresa, cargo).
No asumas el cliente por el destinatario del hilo si la firma indica una empresa distinta.

═══ COMUNICACIONES INTERNAS (CONFIDENCIAL) ═══
NUNCA incluyas en la oferta mensajes, cotizaciones, nombres ni anotaciones internas de
Willington Ortiz u otro personal interno. Las anotaciones internas de precio ("Nata cotiza así")
se usan SOLO para extraer las tarifas, JAMÁS se citan textualmente en el documento.
Cualquier dato faltante o inconsistencia repórtalo en el chat a Naty de forma genérica,
sin atribuirlo a ninguna persona interna.

═══ CUÁNDO GENERAR / CUÁNDO PREGUNTAR ═══
GENERA la oferta directo cuando tengas cliente + equipo principal + tarifa.
PREGUNTA a Naty en el chat (sin generar aún) cuando:
- Falte PESO o DIMENSIONES de la carga → NO uses "A confirmar" ni placeholders; pide los datos.
- Falte el cliente o el equipo principal.
- Algo sea técnicamente inconsistente (p.ej. peso 100 ton en Cama Alta).

═══ ESTRUCTURA OFERTA (NO VARIAR) ═══
- Header: azul #1B2A4A, logo base64, barra #E8601C
- Secciones: Detalle técnico → Económico → Notas → Condiciones → Exclusiones → Firma
- Firma: la define el sistema según el ejecutivo que genera la oferta (NO la escribas tú en el cuerpo)
- Stand-by: OBLIGATORIO
- Póliza de carga y RCE: "Hasta $4.000.000.000 COP". Si el valor declarado supera ese límite,
  indicar que se expide póliza específica con nota de USD + IVA + deducible a cargo del cliente.
- Pago: "50% anticipo / 50% a 30 días" (salvo que se indique otra cosa)
- valor_unit: entero sin separadores (ej: 19500000). Sin precio → 0.
- TABLA DE DETALLE TÉCNICO: si NO hay peso/dimensiones, se OMITE por completo (cargo_items vacío);
  nunca dejar "A confirmar".

═══ TARIFAS STAND-BY 2026 (fijas — nunca "a confirmar") ═══
- Cama Alta 3 ejes: $1.200.000/día
- Cama Baja 3 ejes: $1.500.000/día
- Cama Baja 4 ejes: $1.800.000/día
- Cama Baja 5 ejes: $2.600.000/día
- Cama Alta/Baja Extensible o Semi Modular/Extensible: $2.500.000/día
- Modular 2 Cuna / Modular Cuna 2 Líneas: $4.800.000/día
- Modular 6 Cuna / 6–8 Líneas / Modular 12 líneas: $8.500.000/día
- Modular 18 líneas: $15.000.000/día
- Jacking Skidding: $15.000.000/día
- Camión Turbo: $550.000/día
- Camión Grúa: tarifa pactada (NO es fija)
Si el equipo no coincide exacto con una categoría, usa la más cercana por tipo de cama/estructura.
Nunca dejes la tarifa en blanco.

═══ NOTAS OBLIGATORIAS EN "notas" ═══
- Siempre: "Origen: [origen]\nDestino: [destino]\nEsquema de seguridad: \nStand-by [equipo]: $X.XXX.XXX/día. Las horas adicionales serán cobradas proporcionalmente según tarifa establecida.\nTiempos libres: 6 horas para cargue / 6 horas para descargue."
- Con equipo de IZAJE (grúa, montacargas, modular, skidding, camión grúa): la nota de horas
  adicionales debe decir EXACTAMENTE: "Las horas adicionales para los equipos de izaje serán
  cobradas proporcionalmente según tarifa establecida." Y agregar: "El cobro del equipo de izaje
  inicia desde la llegada del equipo al sitio designado."
- Con skidding/jacking: agregar "Para la operación de skidding se hace necesario expedir póliza
  específica de montaje."
- Modular/multi-punto: tiempos libres 12h en lugar de 6h.

═══ EXTRADIMENSIÓN / MANEJO ESPECIAL ═══
Ancho > 3,00 m o alto > 4,20 m → activa manejo especial: agregar "2 Escoltas + 2 Tecnólogos"
en la config del equipo y nota de gestión de permisos de carga extradimensionada.

═══ OTRAS REGLAS ═══
- Convierte unidades cuando aplique (lb→kg, mm→cm).
- Pricing multi-puerto (ej. Cartagena vs. Barranquilla): preséntalo como OPCIONES dentro de la
  misma oferta.
- Versiona (V2, V3) cuando cambien tarifas o condiciones.
- Referencias: número consecutivo 26-0XXX (5 dígitos).

═══ VOZ NATY ═══
- Formal, directo, sin bullets en textos.
- ALL CAPS: códigos (26-0720), referencias (SOL #673).
- Cierre: "Quedamos atentos a cualquier inquietud".
- Párrafos cortos, concisión extrema.
"""

def _extraer_info_claude(texto: str) -> dict:
    if not ANTHROPIC_OK:
        raise RuntimeError("Instala el paquete 'anthropic': pip install anthropic")
    if not ANTHROPIC_API_KEY:
        raise RuntimeError("Configura ANTHROPIC_API_KEY en el archivo .env")

    client = _anthropic_client()

    user_prompt = f"""Extrae del siguiente texto la información para la oferta BOOM y devuelve SOLO un JSON válido (sin markdown, sin texto extra):

TEXTO:
{texto}

JSON a devolver:
{{"cliente":"","contacto":"","email_cliente":"","ref_cliente":"","cliente_final":"","origen":"","destino":"","descripcion":"","cargo_items":[{{"descripcion":"","tipo":"","cant":1,"dimensiones":"","peso":"","volumen":"","origen_detalle":"","destino_detalle":""}}],"equipos":[{{"equipo":"","config":"","cant":1,"valor_unit":0}}],"notas":"Origen: ...\nDestino: ...\nEsquema de seguridad: \nStand-by [equipo]: $X.XXX.XXX/día. Las horas adicionales serán cobradas proporcionalmente según tarifa establecida.\nTiempos libres: 6 horas para cargue / 6 horas para descargue.","forma_pago":"50% anticipo / 50% a 30 días tras radicación de factura","vigencia":30}}"""

    message = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=2000,
        system=BOOM_SYSTEM_PROMPT,
        messages=[{"role": "user", "content": user_prompt}],
    )
    raw = message.content[0].text.strip()
    if raw.startswith("```"):
        lines = raw.split("\n")
        inner = lines[1:-1] if lines[-1].strip() == "```" else lines[1:]
        raw = "\n".join(inner)
    return json.loads(raw.strip())


# ── Chat IA multi-turno ───────────────────────────────────────────────────────
CHAT_SYSTEM_PROMPT = BOOM_SYSTEM_PROMPT + """

════════════════════════════════════════
MODO CHAT — INSTRUCCIONES
════════════════════════════════════════
Eres el asistente personal de Natalia Vargas en BOOM Logistics. Conversas directamente con ella.
Responde SIEMPRE en español, de forma breve y profesional.

FORMATO DE RESPUESTA:
1. Responde conversacionalmente (1-4 oraciones según la complejidad).
2. Si tienes suficiente información para llenar el formulario de oferta, incluye al FINAL este bloque:

<<<DATOS>>>
{"cliente":"...","contacto":"...","email_cliente":"...","ref_cliente":"...","cliente_final":"...","origen":"...","destino":"...","descripcion":"...","cargo_items":[{"descripcion":"","tipo":"","cant":1,"dimensiones":"","peso":"","volumen":"","origen_detalle":"","destino_detalle":""}],"equipos":[{"equipo":"","config":"","cant":1,"valor_unit":0}],"notas":"...","forma_pago":"...","vigencia":30}
<<<FIN>>>

3. Si el usuario pide un ajuste puntual (ej: "cambia la vigencia", "agrega un escolta"), incluye el bloque JSON completo con el ajuste aplicado.
4. Si falta información clave, pídela conversacionalmente sin incluir el bloque JSON.
5. valor_unit: entero sin separadores (ej: 19500000). Sin precio → 0.
6. Aplica SIEMPRE las reglas de negocio BOOM: stand-by, escoltas, tiempos libres.
"""


def _chat_oferta(messages: list) -> dict:
    if not ANTHROPIC_OK:
        raise RuntimeError("Instala el paquete 'anthropic': pip install anthropic")
    if not ANTHROPIC_API_KEY:
        raise RuntimeError("Configura ANTHROPIC_API_KEY en el archivo .env")

    client = _anthropic_client()
    api_messages = [{"role": m.role, "content": m.content} for m in messages]

    message = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=2000,
        system=CHAT_SYSTEM_PROMPT,
        messages=api_messages,
    )
    raw = message.content[0].text.strip()

    fields = None
    reply = raw
    start = raw.find("<<<DATOS>>>")
    end = raw.find("<<<FIN>>>")
    if start != -1 and end != -1:
        json_str = raw[start + len("<<<DATOS>>>"):end].strip()
        reply = raw[:start].strip()
        try:
            fields = json.loads(json_str)
        except Exception:
            fields = None

    return {"reply": reply, "fields": fields}


# ══════════════════════════════════════════════════════════════════════════════
# MODO AVANZADO — la IA genera la OFERTA HTML COMPLETA directamente (como el chat
# de Claude que usa Boris). Sirve para ofertas COMPLEJAS: multi-moneda (USD),
# alertas, recomendaciones, opciones, mínimos, cuotas de reserva, fotos, ES/EN.
# ══════════════════════════════════════════════════════════════════════════════
_BOOM_OFERTA_CSS = """*{box-sizing:border-box;}
body{font-family:Arial,sans-serif;font-size:13px;color:#1B2A4A;margin:0;padding:16px;background:#f4f4f4;}
.wrapper{max-width:1000px;margin:0 auto;background:#fff;border-radius:6px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,.1);}
.header-bar{background:#1B2A4A;padding:14px 20px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:8px;}
.header-bar img{height:42px;width:auto;}
.header-info{color:#fff;font-size:11px;line-height:1.6;text-align:right;}
.header-info strong{font-size:12px;display:block;}
.ref-bar{background:#E8601C;padding:8px 20px;}
.ref-bar p{margin:0;color:#fff;font-size:11px;font-weight:bold;}
.body{padding:20px;}
.greeting p{font-size:13px;line-height:1.6;margin:0 0 10px 0;}
.section-title{background:#1B2A4A;color:#fff;font-size:12px;font-weight:bold;padding:7px 12px;margin:20px 0 8px 0;border-radius:3px;}
.spec-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin:10px 0;}
.spec-card{background:#f0f4fa;border-radius:5px;padding:9px 8px;text-align:center;}
.spec-card .val{font-size:13px;font-weight:bold;color:#E8601C;}
.spec-card .lbl{font-size:10px;color:#666;margin-top:2px;}
.alerta{background:#fff3cd;border-left:4px solid #E8601C;padding:8px 12px;border-radius:3px;font-size:12px;color:#7a4f00;margin:10px 0;}
.alerta-critica{background:#fdecea;border-left:4px solid #c0392b;padding:8px 12px;border-radius:3px;font-size:12px;color:#7b1a1a;margin:10px 0;font-weight:bold;}
.destacado{background:#eaf2fb;border-left:4px solid #1B6FA8;padding:8px 12px;border-radius:3px;font-size:12px;color:#0d3d61;margin:10px 0;}
.table-scroll{width:100%;overflow-x:auto;}
table.det{width:100%;border-collapse:collapse;font-size:12px;}
table.det th{background:#1B2A4A;color:#fff;padding:7px 8px;font-size:11px;text-align:center;white-space:nowrap;}
table.det th:first-child{text-align:left;}
table.det td{padding:7px 8px;border-bottom:1px solid #e5e5e5;vertical-align:middle;font-size:12px;}
table.det td:not(:first-child){text-align:center;}
table.det td:last-child{text-align:right;font-weight:bold;}
table.det tr:nth-child(even) td{background:#f9f9f9;}
.subtotal-row td{background:#e0e8f5!important;font-weight:bold;color:#1B2A4A;}
.total-row td{background:#E8601C!important;color:#fff;padding:9px 10px;font-weight:bold;font-size:13px;border:none!important;}
.si{color:#1a7d3c;font-weight:bold;}.incluido{color:#1B6FA8;font-weight:bold;}.pendiente{color:#888;font-style:italic;}
ul.notas{margin:0;padding-left:18px;}
ul.notas li{margin-bottom:7px;font-size:13px;line-height:1.6;}
table.cond{width:100%;border-collapse:collapse;font-size:13px;}
table.cond td{padding:7px 10px;border-bottom:1px solid #e5e5e5;vertical-align:top;line-height:1.5;}
table.cond td:first-child{font-weight:bold;width:40%;background:#f5f5f5;}
.footer{border-top:1px solid #e0e0e0;margin-top:22px;padding-top:12px;font-size:13px;color:#444;}
.firma-nombre{font-weight:bold;color:#1B2A4A;font-size:14px;margin:0 0 2px 0;}
.firma-cargo{color:#E8601C;font-size:13px;margin:2px 0;}
.pie{color:#aaa;font-size:11px;margin-top:10px;border-top:1px solid #eee;padding-top:8px;text-align:center;}
.foto-grid{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin:14px 0;}
.foto-grid.una{grid-template-columns:1fr;max-width:820px;margin-left:auto;margin-right:auto;}
.foto-card img{width:100%;height:460px;object-fit:contain;background:#f7f7f7;display:block;border-radius:5px;border:1px solid #ddd;}
.foto-grid.una .foto-card img{height:560px;}
.foto-cap{font-size:11px;color:#555;text-align:center;margin-top:5px;}
@media (max-width:600px){.foto-grid{grid-template-columns:1fr;}}"""

ADVANCED_OFERTA_PROMPT = """Eres el GENERADOR EXPERTO de OFERTAS COMERCIALES HTML de BOOM LOGISTICS COLOMBIA S.A.S.,
empresa colombiana de transporte pesado y carga sobredimensionada. Trabajas para el equipo comercial.

PRINCIPIO: RESULTADO PRIMERO. Deduce el contexto y aplica los estándares BOOM en vez de preguntar.
Si algo es ambiguo pero NO crítico, toma la decisión más sensata y déjala como supuesto/pendiente
DENTRO de la oferta (nota o alerta), nunca como pregunta al usuario. NUNCA inventes cifras, dimensiones
ni datos que no puedas conocer: si falta un dato, muéstralo como pendiente en la oferta (clase .pendiente
o una nota), no lo inventes. Solo pide algo en el chat si es imprescindible para poder cotizar.

ADAPTABILIDAD: sirves para CUALQUIER oferta de BOOM, no solo grandes operaciones. Ajusta el tamaño de la
oferta al negocio: una cotización simple (un flete, un alquiler de camabaja, un stand-by) debe quedar
CORTA y limpia (pocas secciones, sin alertas de colores si no hacen falta); una operación compleja
(modular, izaje, ruta difícil, sobredimensión) usa toda la estructura. No fuerces secciones ni recuadros
que no aporten. No inventes datos de ejemplos que recuerdes de tu entrenamiento: usa SOLO lo que el
usuario te dé en este chat (texto, correo, fotos o archivos adjuntos). OJO: una oferta anterior que el
usuario ADJUNTE o PEGUE en el chat SÍ es fuente válida y DEBES reutilizar sus datos cuando te lo pida
(ver la sección "REPLICAR / VERSIONAR UNA OFERTA ANTERIOR").

IDIOMA: genera la oferta en el idioma del cliente. Si el texto viene en inglés o el cliente lo requiere,
genera TODA la oferta en inglés con la misma estructura y clases.
MONEDA: por DEFECTO la oferta va SIEMPRE en PESOS COLOMBIANOS (COP). El símbolo "$" SIEMPRE significa
pesos colombianos, NUNCA dólares. Genera en dólares (USD) ÚNICAMENTE si el usuario lo escribe de forma
EXPLÍCITA usando la palabra "dólares" / "dollars" / "USD" / "US$" (un simple "$" NO basta y NO autoriza
dólares), o si los valores que te entrega ya vienen rotulados claramente como USD. Si solo ves "$" o
cifras sin rótulo de moneda, o ante CUALQUIER duda, usa COP. NUNCA conviertas de una moneda a otra ni
asumas USD por tu cuenta (tampoco por tratarse de carga importada, comercio exterior o cliente extranjero).
FORMATO DE CIFRAS: separador de miles con PUNTO estilo Colombia SIEMPRE, incluso en inglés
(ej. "USD 6.500", "USD 39.000", "$1.800.000"). Nunca uses coma para miles.

═══ ESTRUCTURA (secciones numeradas con class="section-title", en este orden cuando apliquen) ═══
1. DETALLE TÉCNICO DE LA CARGA / ACLARACIÓN TÉCNICA
2. ALCANCE DEL SERVICIO  (usa .spec-grid con 3 tarjetas para datos clave: peso, equipo, soportes…)
3. FOTOS DE REFERENCIA  (SOLO si hay fotos adjuntas; ver instrucción de fotos)
4. PROPUESTA ECONÓMICA / BUDGETARY PROPOSAL  (table.det: Concepto | Cant | Valor unit | Valor total; fila final class="total-row")
5. NOTAS TÉCNICAS
6. CONDICIONES COMERCIALES  (table.cond)
7. EXCLUSIONES  (ul.notas)
Renumera según las secciones que realmente incluyas. Cierra con footer + firma.

═══ RECUADROS (usa el color correcto) ═══
- class="destacado" (AZUL): recomendaciones técnicas, ventajas de un método, experiencia en vía difícil.
  Inícialo con "✅ <strong>RECOMMENDATION</strong> —" (o RECOMENDACIÓN en español).
- class="alerta" (AMARILLO): "por confirmar", advertencias estándar, IVA no incluido, sobredimensión normal.
  Inícialo con "⚠️ <strong>TO BE CONFIRMED</strong> —" (o POR CONFIRMAR).
- class="alerta-critica" (ROJO): SOLO discrepancias graves, sobredimensión extrema o pendientes críticos.

═══ REGLAS DE NEGOCIO ═══
- Valor Total = Cantidad de EQUIPOS × Valor Unitario (NUNCA piezas × unitario, salvo que el cliente diga que la tarifa es por unidad).
- Ítems con cantidad NO definida: NO los sumes al total; déjalos fuera con nota explícita.
- Cargos mínimos (ej. mínimo 3 turnos): refléjalos en la columna Cantidad (no "1") y explícalos en Notas.
- Cuota de reserva / disponibilidad: línea propia, independiente de los turnos operativos.
- Opciones alternativas A/B: preséntalas etiquetadas, sin asumir cuál elige el cliente.
- STAND-BY 2026 (tarifas FIJAS preestablecidas — úsalas SIEMPRE, salvo que quien realiza la oferta
  indique en el chat otro valor). Toda oferta debe llevar la nota de stand-by con la tarifa del equipo:
    · Cama Alta 3 ejes: $1.200.000/día
    · Cama Baja 3 ejes: $1.500.000/día
    · Cama Baja 4 ejes: $1.800.000/día
    · Cama Baja 5 ejes: $2.600.000/día
    · Cama Alta/Baja Extensible o Semi Modular/Extensible: $2.500.000/día
    · Modular 2 Cuna / Modular Cuna 2 Líneas: $4.800.000/día
    · Modular 6 Cuna / 6–8 Líneas / Modular 12 líneas: $8.500.000/día
    · Modular 18 líneas: $15.000.000/día
    · Jacking Skidding: $15.000.000/día
    · Camión Turbo: $550.000/día
    · Camión Grúa: tarifa pactada (NO es fija)
  Si el equipo no coincide exacto con una categoría, usa la más cercana por tipo de cama/estructura.
  Nunca dejes el stand-by en blanco ni en "a confirmar". Solo cámbialo si el usuario lo pide expresamente.
  ⚠️ OBLIGATORIO Y NO NEGOCIABLE: el stand-by DEBE aparecer como una nota/línea PROPIA y visible dentro
  de las Condiciones Comerciales (o su propia sección), con el VALOR EN PESOS del equipo de la oferta,
  por ejemplo: "Stand-by Cama Alta 3 ejes: $1.200.000/día. Las horas adicionales se cobran
  proporcionalmente según tarifa establecida." NO es suficiente mencionarlo solo en las exclusiones.
  QUEDA PROHIBIDO escribir el stand-by como "según tarifa vigente", "según tarifa", "a confirmar" o
  cualquier frase sin número: SIEMPRE va el valor fijo en pesos. Si la oferta tiene varios equipos,
  incluye el stand-by de CADA uno. Antes de entregar la oferta, verifica que esta nota con su valor esté
  presente; si falta, agrégala.
- Seguros: póliza RCE y de carga hasta $4.000.000.000 COP c/u. Maquinaria USADA: la carga solo ampara
  pérdida total. Si el valor supera el límite, alerta y recomienda top-up. Si no se sabe nueva/usada,
  deja la pregunta abierta en el cierre.
- Sobredimensión: ancho > 3,0 m o alto > 4,40 m ⇒ permiso especial (.alerta). Exceso muy amplio ⇒
  .alerta-critica con el cálculo explícito.
- Pago por defecto: "50% anticipo / 50% contra factura" salvo otra indicación.
- TIEMPOS LIBRES (predeterminado, inclúyelo SIEMPRE en las notas salvo que el usuario lo cambie):
  "Tiempos libres: 6 horas para cargue / 6 horas para descargue."
- PERMISOS (predeterminado, inclúyelo SIEMPRE salvo que no aplique o el usuario lo cambie):
  "Permisos: Incluye gestión de permisos de tránsito ante autoridades competentes (según aplique y reglamentación vigente)."

═══ NOTAS Y AJUSTES DEL USUARIO (RESPETAR TAL CUAL — MUY IMPORTANTE) ═══
Compórtate como un asistente que edita el texto exactamente como se lo piden (igual que Claude en una
conversación normal), SIEMPRE y para CUALQUIER usuario que use el chat (no solo para un jefe o rol
concreto). Cuando cualquier usuario escriba en el chat una NOTA, condición, aclaración, redacción o
corrección con un texto concreto ("agrega esta nota: …", "que diga exactamente …", "cambia X por Y",
"pon en condiciones …"), incorpóralo TAL CUAL, palabra por palabra, en el lugar que corresponda de la
oferta: NO lo parafrasees, NO lo resumas, NO lo suavices, NO cambies el orden ni la redacción (respeta
mayúsculas, cifras, signos y saltos). Lo que el usuario indica MANDA sobre cualquier valor por defecto:
si pide algo distinto a un estándar (stand-by, tiempos libres, permisos, forma de pago, exclusiones,
etc.), aplica exactamente lo que él diga. Solo completa o reordena lo que él NO haya especificado.

═══ REPLICAR / VERSIONAR UNA OFERTA ANTERIOR (MUY IMPORTANTE) ═══
Cuando el usuario adjunte o pegue una oferta anterior (HTML, PDF o texto) y te pida algo como "haz la
misma oferta" / "la misma exacta oferta" para OTRO cliente y/o con un AJUSTE de precio, DEBES:
1. REPRODUCIR exactamente la misma estructura, secciones, conceptos, cantidades, notas y condiciones de
   esa oferta. No la simplifiques ni omitas ítems: es la MISMA oferta.
2. CAMBIAR el nombre del cliente al nuevo que te indiquen (en la ref-bar y donde aparezca). El cliente
   anterior (ej. Hansa) NO debe quedar en ningún lugar de la oferta nueva.
3. APLICAR el ajuste de precio pedido a CADA valor unitario. Ejemplo "+15%": nuevo unitario = unitario
   anterior × 1,15. Luego recalcula TODOS los valores totales de cada línea y el TOTAL general con los
   nuevos unitarios. Si el ajuste es otro (−10%, un valor fijo nuevo, etc.), aplícalo igual a todas las
   líneas que corresponda. Mantén el formato de cifras con punto de miles.
4. Usar el NUEVO número de referencia {{REF}} que ya asignó el sistema (NO el de la oferta vieja) y el
   mes/año actuales, salvo que el usuario indique otra fecha.
5. En la frase de resumen del chat, di explícitamente qué cambió: cliente nuevo, % aplicado y el nuevo TOTAL.
NUNCA devuelvas la oferta anterior sin aplicar los cambios pedidos. Si un precio no puede ajustarse porque
está "por confirmar", déjalo indicado como pendiente; no lo inventes.

═══ FIRMANTE ═══
Firma SIEMPRE con el ejecutivo indicado en "FIRMANTE POR DEFECTO" (más abajo): es el usuario que está
creando la oferta. Usa su nombre, cargo y correo tal cual. NO incluyas NINGÚN teléfono al pie de la firma.
SOLO usa otro firmante si el usuario lo pide explícitamente en el chat (ej. "fírmala Boris"). Nunca pongas a Boris por defecto.
En meta.realizada escribe el nombre corto del firmante que realmente usaste.

═══ PRESENTACIÓN (OBLIGATORIO) ═══
- NUNCA agregues una columna, tarjeta, celda ni fila de "FORMALIZADA POR" / "FORMALIZED BY". Quien REALIZA
  la oferta la formaliza: muestra a lo sumo UN responsable (el firmante). No pongas dos roles separados
  (realizada/formalizada); si vas a indicar quién la hizo, una sola casilla basta.
- NOMBRE DEL CLIENTE: escríbelo EXACTAMENTE como te lo da el usuario (ej. "Geodis" se queda "Geodis", NO
  "GEODIS" ni una versión "unificada"/traducida). No cambies mayúsculas/minúsculas, no abrevies, no traduzcas
  ni fusiones el nombre del cliente. Respeta la marca tal cual está escrita, incluida la ref-bar.
- VALOR TOTAL: preséntalo en tamaño DISCRETO. Úsalo dentro de la fila class="total-row" de la tabla
  económica. NO lo pongas en tipografía gigante ni en un recuadro enorme aparte; si lo destacas, que su
  fuente no supere ~15px y luzca equilibrado con el resto de la oferta.
- PROHIBIDO el "encabezado resumen": NUNCA agregues una tabla, banda ni grid de columnas que repita
  datos como REALIZADA POR / FORMALIZADA POR / CLIENTE / TIPO NEGOCIO / VALOR COP. Esos datos ya van en
  la ref-bar (cliente y ref) y el valor SOLO va en la fila total-row de la tabla económica. No dupliques
  esa información en una banda superior, ni en <table>, ni en <div> con columnas. El único encabezado
  permitido es el header-bar (logo) + la ref-bar naranja del esqueleto.

═══ ESQUELETO OBLIGATORIO DEL HTML ═══
Devuelve un documento HTML completo (<!DOCTYPE html> … </html>) con <head> que incluya EXACTAMENTE el
bloque <style> que se te entrega (no cambies las clases). El <body> debe ser:
<div class="wrapper">
<div class="header-bar">{{LOGO}}<div style="display:flex;align-items:center;gap:16px;"><div class="header-info"><strong>BOOM LOGISTICS COLOMBIA S.A.S.</strong>Specialized Transportation Solutions</div>{{SELLO}}</div></div>
<div class="ref-bar"><p>REF: {{REF}} &nbsp;|&nbsp; CLIENTE &nbsp;|&nbsp; RUTA/DESCRIPCIÓN &nbsp;|&nbsp; MES AÑO[ &nbsp;|&nbsp; ESTADO]</p></div>
<div class="body"> … secciones … <div class="footer"> … firma con .firma-nombre y .firma-cargo … </div></div>
</div>
NO cambies los marcadores {{LOGO}}, {{SELLO}}, {{REF}}: déjalos literales, el sistema los reemplaza.
Usa el número de referencia {{REF}} tal cual (el sistema ya asignó el consecutivo).

═══ FOTOS ═══
Si hay N fotos adjuntas (te diré cuántas), incluye la sección "FOTOS DE REFERENCIA / REFERENCE PHOTOS"
con una <div class="foto-grid"> que contenga EXACTAMENTE N <div class="foto-card"> con
<img src="{{FOTO_1}}"> … <img src="{{FOTO_N}}"> y un <p class="foto-cap"> descriptivo en cada una
(aclara "operación similar, no la carga real" cuando sea referencia). Deja los marcadores {{FOTO_n}}
literales; el sistema pega la imagen real. Si NO hay fotos, omite la sección por completo.
IMPORTANTE (encuadre de la foto): la foto SIEMPRE debe verse ajustada y ordenada. Cuando haya UNA sola
foto usa <div class="foto-grid una"> (así queda centrada y con buen ancho); con 2 o más usa solo
<div class="foto-grid">. El tamaño y encuadre lo controla el CSS (no pongas width/height en línea en
las <img>); la clase .foto-card ya deja la imagen contenida sin deformarla ni desbordarse.

═══ SALIDA (OBLIGATORIO, respeta el formato) ═══
1) Una frase breve (1-3 líneas) para el chat con los montos clave / cambios / pendientes.
2) El HTML completo entre los marcadores <<<HTML>>> y <<<FINHTML>>> (sin ``` , sin markdown).
3) Los metadatos entre <<<META>>> y <<<FINMETA>>>: un JSON con
   {"cliente":"","valor":0,"moneda":"COP|USD","mes":"ENE..DIC","tipo":"","descripcion":"","realizada":"Boris Borrego|Natalia Vargas|Willington Ortiz","forma_pago":""}
   donde "valor" es el TOTAL numérico entero sin separadores (ej. 39000) y "mes" el mes de la oferta en
   mayúsculas de 3 letras. "moneda" debe coincidir con la moneda usada en el HTML: "COP" por DEFECTO;
   pon "USD" SOLO si la oferta realmente se generó en dólares porque el usuario lo pidió. "cliente": pon el nombre del cliente EXACTAMENTE como lo escribió el usuario en
   el chat / correo (mismas mayúsculas y minúsculas, misma marca). PROHIBIDO cambiarlo: no lo pongas en
   MAYÚSCULAS, no lo traduzcas, no lo abrevies, no lo "unifiques" ni lo fusiones con otro. Debe coincidir
   letra por letra con el que aparece en la ref-bar del HTML. Si el usuario no dio un nombre claro, deja
   "cliente" vacío (no inventes uno). En "forma_pago" pon EXACTAMENTE la forma de pago que quedó escrita en la
   oferta: si el usuario indicó en el chat un anticipo distinto (ej. "70% de anticipo", "60% anticipo /
   40% a 30 días"), reporta ESA, no la predeterminada. Si algún dato no aplica, déjalo vacío o en 0."""


class OfertaIABody(BaseModel):
    messages: List[ChatMsg]
    fotos: Optional[List[str]] = []
    ref: Optional[str] = None
    firmante: Optional[dict] = None  # {nombre, cargo, email, telefono} del usuario logueado
    forma_pago: Optional[str] = None  # forma de pago elegida en la tirilla; la oferta la usa tal cual


def _limpiar_oferta_html(html: str) -> str:
    """Red de seguridad determinista sobre el HTML de la oferta IA.
    Elimina cualquier tabla/banda de resumen que contenga la columna 'FORMALIZADA POR'
    (que a veces el modelo agrega pese al prompt). Esa banda es la que además muestra
    el cliente en mayúsculas y el valor en tipografía gigante, así que al quitarla se
    corrigen los tres problemas de una sola vez. El dato real de cliente/valor queda en
    la ref-bar y en la tabla económica (fila total-row, discreta)."""
    if not html or "formaliza" not in html.lower():
        return html
    # 1) Quita tablas completas que incluyan "FORMALIZADA/FORMALIZED"
    html = re.sub(r"<table\b[^>]*>.*?</table>",
                  lambda m: "" if re.search(r"formaliz", m.group(0), re.I) else m.group(0),
                  html, flags=re.I | re.S)
    # 2) Fallback: si quedó una fila <tr> suelta con "formaliza", quítala
    html = re.sub(r"<tr\b[^>]*>(?:(?!</tr>).)*?formaliz(?:(?!</tr>).)*?</tr>", "",
                  html, flags=re.I | re.S)
    return html


def _inyectar_recursos_oferta(html: str, ref_fmt: str, fotos: list) -> str:
    """Reemplaza los marcadores {{LOGO}}, {{SELLO}}, {{REF}}, {{FOTO_n}} por los recursos reales."""
    logo_src = _logo_src()
    sello_src = _sello_src()
    logo_html = (f'<img src="{logo_src}" alt="BOOM Logistics"/>' if logo_src
                 else '<span style="color:#fff;font-weight:bold;font-size:18px;">BOOM</span>')
    sello_html = (f'<img src="{sello_src}" alt="Construimos Pais - Boom Logistics" style="height:90px;width:auto;"/>'
                  if sello_src else "")
    html = html.replace("{{LOGO}}", logo_html).replace("{{SELLO}}", sello_html)
    html = html.replace("{{REF}}", ref_fmt)
    for i, src in enumerate(fotos or [], start=1):
        if isinstance(src, str) and src.strip().startswith("data:image"):
            html = html.replace("{{FOTO_%d}}" % i, src)
    # Limpia cualquier marcador de foto que el modelo dejara de más
    html = re.sub(r"\{\{FOTO_\d+\}\}", "", html)
    return html


def _oferta_ia(messages: list, fotos: list, ref: str, firmante: dict = None, forma_pago: str = None) -> dict:
    if not ANTHROPIC_OK:
        raise RuntimeError("Instala el paquete 'anthropic': pip install anthropic")
    if not ANTHROPIC_API_KEY:
        raise RuntimeError("Configura ANTHROPIC_API_KEY en el archivo .env")

    ref_fmt = _fmt_ref(ref or "260001")
    n_fotos = len([f for f in (fotos or []) if isinstance(f, str) and f.strip().startswith("data:image")])

    _MESES = ["ENE","FEB","MAR","ABR","MAY","JUN","JUL","AGO","SEP","OCT","NOV","DIC"]
    hoy = datetime.now()
    mes_actual, anio_actual = _MESES[hoy.month - 1], hoy.year

    sys = (ADVANCED_OFERTA_PROMPT
           + "\n\nBLOQUE <style> A USAR TAL CUAL:\n<style>" + _BOOM_OFERTA_CSS + "</style>"
           + f"\n\nNÚMERO DE REFERENCIA ASIGNADO: {ref_fmt} (úsalo en {{{{REF}}}})."
           + f"\nFECHA ACTUAL: {mes_actual} {anio_actual} (usa este mes/año en la ref-bar y en meta.mes, "
             f"salvo que el cliente indique otra fecha)."
           + f"\nFOTOS ADJUNTAS: {n_fotos}.")

    if firmante and isinstance(firmante, dict) and firmante.get("nombre"):
        _f_partes = [str(firmante.get("nombre", "")).strip()]
        if firmante.get("cargo"):
            _f_partes.append(str(firmante["cargo"]).strip())
        if firmante.get("email"):
            _f_partes.append(str(firmante["email"]).strip())
        sys += ("\n\nFIRMANTE POR DEFECTO: " + " — ".join(_f_partes)
                + " (firma la oferta con este ejecutivo salvo que el usuario pida otro en el chat; NO pongas teléfono al pie de la firma).")

    if forma_pago and str(forma_pago).strip():
        sys += ("\n\nFORMA DE PAGO A USAR EN LA OFERTA: \"" + str(forma_pago).strip()
                + "\" — escríbela TAL CUAL en Condiciones Comerciales, salvo que el usuario indique en el "
                  "chat otra forma de pago (en ese caso manda lo que diga el chat).")

    client = _anthropic_client()
    api_messages = _trim_api_messages([{"role": m.role, "content": m.content} for m in messages])
    message = client.messages.create(
        model="claude-sonnet-4-5",
        max_tokens=8000,
        system=sys,
        messages=api_messages,
    )
    raw = message.content[0].text

    def _between(t, a, b):
        i, j = t.find(a), t.find(b)
        if i != -1 and j != -1 and j > i:
            return t[i + len(a):j].strip()
        return ""

    html = _between(raw, "<<<HTML>>>", "<<<FINHTML>>>")
    meta_str = _between(raw, "<<<META>>>", "<<<FINMETA>>>")
    reply = raw.split("<<<HTML>>>")[0].strip() if "<<<HTML>>>" in raw else raw.strip()

    meta = None
    if meta_str:
        try:
            meta = json.loads(meta_str)
        except Exception:
            meta = None

    if html:
        html = _inyectar_recursos_oferta(html, ref_fmt, fotos)
        html = _limpiar_oferta_html(html)
        html = _inject_anexo(html)

    return {"reply": reply, "html": html, "meta": meta, "ref": ref_fmt}


# ── Routes ────────────────────────────────────────────────────────────────────
@app.get("/", response_class=HTMLResponse)
def index():
    # Sin caché: así el equipo siempre carga la última versión al recargar,
    # sin tener que hacer Ctrl+Shift+R.
    return FileResponse("templates/index.html", headers={
        "Cache-Control": "no-cache, no-store, must-revalidate",
        "Pragma": "no-cache",
        "Expires": "0",
    })


@app.get("/manual", response_class=HTMLResponse)
def manual():
    """Manual interactivo para el equipo (Willy, Boris). Público, sin login."""
    return FileResponse("templates/manual.html")


@app.get("/anexo-legal")
def anexo_legal():
    """Anexo 1 – Condiciones Legales de la Cotización. Público (el cliente lo
    abre desde el enlace de la oferta). Fuente única: se actualiza aquí y todas
    las ofertas nuevas apuntan a la versión vigente."""
    path = os.path.join(os.path.dirname(__file__), "templates", "anexo1_cotizacion.pdf")
    return FileResponse(
        path,
        media_type="application/pdf",
        headers={"Content-Disposition": "inline; filename=BOOM_Anexo1_Cotizacion.pdf"},
    )


@app.get("/api/logo")
def get_logo():
    return {"src": _logo_src()}


@app.get("/api/consecutivo")
def get_consecutivo():
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT COALESCE(MAX(CAST(num AS INTEGER)), 260000) + 1 AS next FROM ofertas WHERE NOT COALESCE(es_prueba, false)")
            return {"consecutivo": fetchone(cur)["next"]}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


@app.get("/api/rentabilidad")
def get_rentabilidad(anio: Optional[str] = None, mes: Optional[str] = None):
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            filters, params = ["UPPER(respuesta) = 'ACEPTADA'"], []
            if anio:
                filters.append("EXTRACT(YEAR FROM COALESCE(fecha, created_at))::text = %s")
                params.append(anio)
            if mes:
                filters.append("UPPER(mes_aceptado) = %s")
                params.append(mes.upper())
            where = " AND ".join(filters)
            cur.execute(f"""
                SELECT id, num, cliente, mes_aceptado AS mes, valor AS ingreso, costo_proyecto AS costo
                FROM ofertas
                WHERE {where}
                ORDER BY CAST(num AS INTEGER) DESC
            """, params)
            rows = fetchall(cur)
        return {"rows": rows}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


@app.get("/api/ofertas")
def list_ofertas():
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT * FROM ofertas ORDER BY CAST(num AS INTEGER) DESC")
            return fetchall(cur)
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


# Facturación de CONTRATOS (sin número de oferta) según hoja "Facturación"
# del Excel de agosto. Es una línea resumen; el detalle se cargará después.
CONTRATOS_FACTURADO_2026 = 2_623_851_133


@app.get("/api/facturas")
def list_facturas(oferta: Optional[str] = Query(None)):
    """Facturas (1 oferta -> N facturas). Cada fila de la hoja 'Ofertas Aprobadas'
    es UNA línea facturable. Un mismo N° de factura puede cubrir varias ofertas.
    Si se pasa ?oferta=260123 devuelve solo las líneas de esa oferta."""
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            if oferta:
                cur.execute(
                    "SELECT oferta_num,ref_original,mes,cliente,descripcion,origen,destino,"
                    "valor,estado_proyecto,responsable,no_factura "
                    "FROM facturas WHERE oferta_num=%s ORDER BY id", (oferta,))
            else:
                cur.execute(
                    "SELECT oferta_num,ref_original,mes,cliente,descripcion,origen,destino,"
                    "valor,estado_proyecto,responsable,no_factura FROM facturas ORDER BY id")
            return fetchall(cur)
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


@app.get("/api/facturas/por_oferta")
def facturas_por_oferta():
    """Agrupa las facturas por número de oferta para el módulo Ofertas Aprobadas.
    Devuelve un dict { oferta_num: {n, facturado_real, total_lineas, lineas:[...]} }.
    'facturado_real' suma solo estados FACTURADO% (facturación real);
    'total_lineas' suma todo (incluye proyección: POR EJECUTAR / EN EJECUCIÓN / EJECUTADO)."""
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT oferta_num,mes,estado_proyecto,valor,no_factura,descripcion "
                "FROM facturas WHERE oferta_num IS NOT NULL ORDER BY oferta_num, id")
            out = {}
            for num, mes, estado, valor, nofact, desc in cur.fetchall():
                g = out.setdefault(num, {"n": 0, "facturado_real": 0,
                                         "total_lineas": 0, "lineas": []})
                v = int(valor or 0)
                g["n"] += 1
                g["total_lineas"] += v
                if (estado or "").upper().startswith("FACTURADO"):
                    g["facturado_real"] += v
                g["lineas"].append({
                    "mes": mes, "estado": estado, "valor": v,
                    "no_factura": nofact, "descripcion": desc,
                })
            return out
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


@app.get("/api/ofertas_2025")
def list_ofertas_2025():
    """Ofertas 2025 (histórico facturado), tabla separada."""
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT * FROM ofertas_2025 ORDER BY valor_facturado DESC")
            return fetchall(cur)
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


@app.get("/api/contratos")
def list_contratos():
    """Contratos reales (mensuales), por cliente."""
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT * FROM facturacion_cat WHERE categoria='CONTRATO' "
                        "ORDER BY valor_facturado DESC")
            return fetchall(cur)
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


@app.get("/api/facturacion_cat")
def list_facturacion_cat(categoria: Optional[str] = Query(None)):
    """Facturas sin número de oferta, categorizadas por cliente.
    categoria opcional: CONTRATO | 2025_ANO_PASADO | OTROS."""
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            if categoria:
                cur.execute("SELECT * FROM facturacion_cat WHERE categoria=%s "
                            "ORDER BY valor_facturado DESC", (categoria,))
            else:
                cur.execute("SELECT * FROM facturacion_cat "
                            "ORDER BY categoria, valor_facturado DESC")
            return fetchall(cur)
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


@app.get("/api/facturacion/resumen")
def facturacion_resumen():
    """Resumen general que concilia con el Excel, en 4 categorías:
    Ofertas 2026 + Año Anterior 2025 (numeradas + año pasado) + Contratos + Otros."""
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT COALESCE(SUM(valor_facturado),0) FROM ofertas "
                        "WHERE UPPER(respuesta)='ACEPTADA'")
            fact_2026 = cur.fetchone()[0] or 0
            cur.execute("SELECT COALESCE(SUM(valor_facturado),0), COUNT(*) FROM ofertas_2025")
            row = cur.fetchone()
            f2025_num, n_2025 = (row[0] or 0), (row[1] or 0)
            cur.execute("SELECT categoria, COALESCE(SUM(valor_facturado),0) "
                        "FROM facturacion_cat GROUP BY categoria")
            cat = {r[0]: (r[1] or 0) for r in cur.fetchall()}
        contratos = int(cat.get("CONTRATO", 0))
        ano_pasado = int(cat.get("2025_ANO_PASADO", 0))
        otros = int(cat.get("OTROS", 0))
        fact_2025 = int(f2025_num) + ano_pasado
        # Respaldo: si aún no se cargó facturacion_cat, usar el valor fijo.
        if contratos == 0 and ano_pasado == 0 and otros == 0:
            contratos = CONTRATOS_FACTURADO_2026
        return {
            "ofertas_2026": int(fact_2026),
            "ofertas_2025": fact_2025,
            "ofertas_2025_numeradas": int(f2025_num),
            "ofertas_2025_ano_pasado": ano_pasado,
            "contratos": contratos,
            "otros": otros,
            "total": int(fact_2026) + fact_2025 + contratos + otros,
            "n_ofertas_2025": int(n_2025),
        }
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


@app.get("/api/facturacion/estado_proyecto")
def facturacion_estado_proyecto():
    """Reproduce la tabla dinámica de control (Natalia): agrupa la tabla
    `facturas` por `estado_proyecto` sumando el valor. Devuelve el facturado
    por mes, los estados de ejecución y la proyección (lo que aún NO se ha
    facturado = POR EJECUTAR + EN EJECUCIÓN + EJECUTADO)."""
    _ORDEN_MES = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
                  "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT UPPER(TRIM(COALESCE(estado_proyecto,''))) e, "
                "COUNT(*) n, COALESCE(SUM(valor),0) v "
                "FROM facturas GROUP BY 1")
            filas = fetchall(cur)
            # Detalle línea a línea de las ofertas que aún NO se han facturado
            # (por ejecutar / en ejecución / ejecutado) para desplegar en el panel.
            cur.execute(
                "SELECT oferta_num, cliente, descripcion, mes, responsable, "
                "COALESCE(valor,0) valor, UPPER(TRIM(COALESCE(estado_proyecto,''))) e "
                "FROM facturas "
                "WHERE UPPER(TRIM(COALESCE(estado_proyecto,''))) IN "
                "('POR EJECUTAR','EN EJECUCION','EN EJECUCIÓN','EJECUTADO') "
                "ORDER BY valor DESC")
            filas_det = fetchall(cur)

        facturado_mes = {}     # mes -> {valor, n}
        proyeccion = {"POR EJECUTAR": {"valor": 0, "n": 0},
                      "EN EJECUCION": {"valor": 0, "n": 0},
                      "EJECUTADO": {"valor": 0, "n": 0}}
        detalle = {"POR EJECUTAR": [], "EN EJECUCION": [], "EJECUTADO": []}
        for r in filas_det:
            e_norm = (r.get("e") or "").strip().replace("Ó", "O").replace("Í", "I")
            if e_norm not in detalle:
                continue
            detalle[e_norm].append({
                "oferta_num": r.get("oferta_num") or "",
                "cliente": r.get("cliente") or "",
                "descripcion": r.get("descripcion") or "",
                "mes": r.get("mes") or "",
                "responsable": r.get("responsable") or "",
                "valor": int(r.get("valor") or 0),
            })
        cancelado = {"valor": 0, "n": 0}
        otros = {"valor": 0, "n": 0}

        for r in filas:
            e = (r.get("e") or "").strip()
            v = int(r.get("v") or 0)
            n = int(r.get("n") or 0)
            e_norm = e.replace("Ó", "O").replace("Í", "I")
            if e_norm.startswith("FACTURADO"):
                mes = e_norm.replace("FACTURADO", "").strip() or "(SIN MES)"
                d = facturado_mes.setdefault(mes, {"valor": 0, "n": 0})
                d["valor"] += v
                d["n"] += n
            elif e_norm == "POR EJECUTAR":
                proyeccion["POR EJECUTAR"]["valor"] += v
                proyeccion["POR EJECUTAR"]["n"] += n
            elif e_norm in ("EN EJECUCION", "EN EJECUCION "):
                proyeccion["EN EJECUCION"]["valor"] += v
                proyeccion["EN EJECUCION"]["n"] += n
            elif e_norm == "EJECUTADO":
                proyeccion["EJECUTADO"]["valor"] += v
                proyeccion["EJECUTADO"]["n"] += n
            elif e_norm == "CANCELADO":
                cancelado["valor"] += v
                cancelado["n"] += n
            elif e_norm and e_norm != "(SIN MES)":
                otros["valor"] += v
                otros["n"] += n

        meses = []
        for m in _ORDEN_MES:
            if m in facturado_mes:
                d = facturado_mes[m]
                meses.append({"mes": m, "valor": d["valor"], "n": d["n"]})
        # Meses que no estén en el orden estándar (por si acaso)
        for m, d in facturado_mes.items():
            if m not in _ORDEN_MES:
                meses.append({"mes": m, "valor": d["valor"], "n": d["n"]})

        facturado_total = sum(x["valor"] for x in meses)
        proy_total = sum(p["valor"] for p in proyeccion.values())
        total_general = facturado_total + proy_total + cancelado["valor"] + otros["valor"]

        return {
            "facturado_por_mes": meses,
            "facturado_total": facturado_total,
            "proyeccion": {
                "por_ejecutar": proyeccion["POR EJECUTAR"],
                "en_ejecucion": proyeccion["EN EJECUCION"],
                "ejecutado": proyeccion["EJECUTADO"],
                "total": proy_total,
            },
            "detalle": {
                "por_ejecutar": detalle["POR EJECUTAR"],
                "en_ejecucion": detalle["EN EJECUCION"],
                "ejecutado": detalle["EJECUTADO"],
            },
            "cancelado": cancelado,
            "otros": otros,
            "total_general": total_general,
        }
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


@app.get("/api/facturacion/real")
def facturacion_real():
    """Bloque 1 de la Proyección: facturado REAL de Vulcano (hoja Facturación),
    por mes y clasificado por origen (Ofertas 2026 / 2025 / Contratos y otros).
    Fuente: tabla `vulcano_facturas` (excluye las marcadas como excluida)."""
    _ORDEN_MES = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
                  "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT UPPER(TRIM(COALESCE(mes,''))) m, "
                "UPPER(TRIM(COALESCE(clase,'CONTRATO'))) c, "
                "COUNT(*) n, COALESCE(SUM(subtotal),0) v "
                "FROM vulcano_facturas WHERE excluida = false GROUP BY 1,2")
            filas = fetchall(cur)

        # mes -> {of2026, of2025, contrato, total, n}
        por_mes = {}
        tot = {"of2026": 0, "of2025": 0, "contrato": 0, "total": 0, "n": 0}
        for r in filas:
            m = (r.get("m") or "").strip() or "(SIN MES)"
            c = (r.get("c") or "CONTRATO").strip()
            v = int(r.get("v") or 0)
            n = int(r.get("n") or 0)
            key = "of2026" if c == "2026" else ("of2025" if c == "2025" else "contrato")
            d = por_mes.setdefault(m, {"of2026": 0, "of2025": 0, "contrato": 0, "total": 0, "n": 0})
            d[key] += v
            d["total"] += v
            d["n"] += n
            tot[key] += v
            tot["total"] += v
            tot["n"] += n

        meses = []
        for m in _ORDEN_MES:
            if m in por_mes:
                d = por_mes[m]
                meses.append({"mes": m, **d})
        for m, d in por_mes.items():
            if m not in _ORDEN_MES:
                meses.append({"mes": m, **d})

        return {"por_mes": meses, "totales": tot}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


class ContratoPendCreate(BaseModel):
    cliente: str
    descripcion: Optional[str] = ""
    valor: Optional[int] = 0
    mes: Optional[str] = ""
    estado: Optional[str] = "CONFIRMADO"
    oferta_ref: Optional[str] = ""


class ContratoPendUpdate(BaseModel):
    cliente: Optional[str] = None
    descripcion: Optional[str] = None
    valor: Optional[int] = None
    mes: Optional[str] = None
    estado: Optional[str] = None
    oferta_ref: Optional[str] = None


def _cp_valor(v):
    """Convierte '35.500.000', '$35,500,000' o 35500000 a entero de pesos."""
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    s = re.sub(r"[^\d]", "", str(v))
    return int(s) if s else 0


@app.get("/api/contratos_pendientes")
def contratos_pendientes_list():
    """Bloque 3: líneas manuales de contratos pendientes por facturar."""
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT id, cliente, descripcion, valor, mes, estado, oferta_ref, "
                "creado_por, created_at FROM contratos_pendientes ORDER BY id DESC")
            filas = fetchall(cur)
        pendiente = sum(int(r.get("valor") or 0) for r in filas
                        if (r.get("estado") or "").upper() != "FACTURADO")
        total = sum(int(r.get("valor") or 0) for r in filas)
        return {"items": filas, "pendiente_total": pendiente, "total": total, "n": len(filas)}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


@app.post("/api/contratos_pendientes")
def contratos_pendientes_create(body: ContratoPendCreate, request: Request):
    try:
        cliente = re.sub(r"\s+", " ", (body.cliente or "").strip()).upper()
        if not cliente:
            raise HTTPException(400, "El cliente es obligatorio")
        estado = (body.estado or "CONFIRMADO").strip().upper()
        mes = (body.mes or "").strip().upper()
        creado_por = request.state.user.get("nombre", "") if hasattr(request.state, "user") else ""
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute(
                "INSERT INTO contratos_pendientes "
                "(cliente, descripcion, valor, mes, estado, oferta_ref, creado_por) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id",
                (cliente, (body.descripcion or "").strip(), _cp_valor(body.valor),
                 mes, estado, (body.oferta_ref or "").strip(), creado_por))
            new_id = fetchone(cur)["id"]
        return {"ok": True, "id": new_id}
    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


@app.patch("/api/contratos_pendientes/{cid}")
def contratos_pendientes_update(cid: int, body: ContratoPendUpdate, request: Request):
    try:
        fields = {}
        if body.cliente is not None:
            fields["cliente"] = re.sub(r"\s+", " ", body.cliente.strip()).upper()
        if body.descripcion is not None:
            fields["descripcion"] = body.descripcion.strip()
        if body.valor is not None:
            fields["valor"] = _cp_valor(body.valor)
        if body.mes is not None:
            fields["mes"] = body.mes.strip().upper()
        if body.estado is not None:
            fields["estado"] = body.estado.strip().upper()
        if body.oferta_ref is not None:
            fields["oferta_ref"] = body.oferta_ref.strip()
        if not fields:
            return {"ok": True, "sin_cambios": True}
        sets = ", ".join(f"{k} = %s" for k in fields)
        vals = list(fields.values()) + [cid]
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute(
                f"UPDATE contratos_pendientes SET {sets}, updated_at = now() WHERE id = %s",
                vals)
            if cur.rowcount == 0:
                raise HTTPException(404, "Línea no encontrada")
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


@app.delete("/api/contratos_pendientes/{cid}")
def contratos_pendientes_delete(cid: int, request: Request):
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("DELETE FROM contratos_pendientes WHERE id = %s", (cid,))
            if cur.rowcount == 0:
                raise HTTPException(404, "Línea no encontrada")
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


@app.get("/api/ofertas/{oferta_id}")
def get_oferta(oferta_id: int):
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT * FROM ofertas WHERE id = %s", (oferta_id,))
            row = fetchone(cur)
            if row is None:
                raise HTTPException(404, "Oferta no encontrada")
            return row
    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


@app.post("/api/ofertas", status_code=201)
def create_oferta(oferta: OfertaCreate):
    """Crea una oferta asignando el consecutivo de forma segura ante varios
    usuarios simultáneos: si el número choca con uno ya existente, el servidor
    calcula el siguiente número disponible y reintenta, para que nunca se pierda
    la oferta ni se repita el consecutivo."""
    pdf_json = json.dumps(oferta.pdf_data) if oferta.pdf_data else None
    # Normaliza el nombre del cliente (quita espacios dobles/extremos y lo pasa a
    # MAYÚSCULA) para que coincida con el catálogo oficial y no se creen duplicados
    # por tipeo ni por mayúsculas/minúsculas distintas.
    if oferta.cliente:
        oferta.cliente = re.sub(r"\s+", " ", oferta.cliente.strip()).upper()
    # El nombre del comercial (quien realiza/formaliza) siempre en MAYÚSCULA.
    if oferta.realizada:
        oferta.realizada = re.sub(r"\s+", " ", oferta.realizada.strip()).upper()
    if oferta.formalizada:
        oferta.formalizada = re.sub(r"\s+", " ", oferta.formalizada.strip()).upper()
    MAX_INTENTOS = 10
    for intento in range(MAX_INTENTOS):
        try:
            with get_conn() as conn:
                cur = conn.cursor()
                # Primer intento: respeta el número que envió el cliente (si vino).
                # Si no vino, o si un intento anterior chocó, recalcula el siguiente
                # consecutivo disponible en este preciso momento.
                if not oferta.num or intento > 0:
                    cur.execute("SELECT COALESCE(MAX(CAST(num AS INTEGER)), 260000) + 1 AS next FROM ofertas WHERE NOT COALESCE(es_prueba, false)")
                    oferta.num = str(fetchone(cur)["next"])
                cur.execute(
                    """INSERT INTO ofertas
                       (num,mes,fecha,cliente,realizada,formalizada,unidad,tipo,sector,
                        valor,estado,respuesta,facturacion,general,seguimiento,mes_aceptado,
                        fecha_facturacion,valor_facturado,no_factura,valor_aprobado,pdf_data)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                       RETURNING *""",
                    (oferta.num, oferta.mes, oferta.fecha or None, oferta.cliente,
                     oferta.realizada, oferta.formalizada, oferta.unidad, oferta.tipo,
                     oferta.sector,
                     oferta.valor or 0, oferta.estado or "ENVIADO", oferta.respuesta,
                     oferta.facturacion,
                     oferta.general, oferta.seguimiento, oferta.mes_aceptado,
                     oferta.fecha_facturacion or None, oferta.valor_facturado, oferta.no_factura,
                     oferta.valor_aprobado,
                     pdf_json),
                )
                nueva = fetchone(cur)
                # Mantiene el catálogo sincronizado: si el cliente de esta oferta
                # aún no está registrado, lo agrega (así nunca queda por fuera).
                if oferta.cliente:
                    cur.execute(
                        """INSERT INTO clientes (nombre_corto)
                           SELECT %s
                           WHERE NOT EXISTS (
                               SELECT 1 FROM clientes WHERE lower(nombre_corto) = lower(%s)
                           )""",
                        (oferta.cliente, oferta.cliente),
                    )
                return nueva
        except pgdb.DatabaseError as e:
            msg = str(e).lower()
            if "unique" in msg or "duplicate" in msg:
                # Choque de consecutivo (otro usuario tomó ese número casi al
                # mismo tiempo): reintenta con el siguiente número disponible.
                oferta.num = None
                continue
            traceback.print_exc()
            raise HTTPException(500, str(e))
        except Exception as e:
            traceback.print_exc()
            raise HTTPException(500, str(e))
    raise HTTPException(409, "No se pudo asignar un consecutivo único; intenta guardar de nuevo")


@app.patch("/api/ofertas/{oferta_id}")
def update_oferta(oferta_id: int, oferta: OfertaUpdate, request: Request):
    try:
        fields = {k: v for k, v in oferta.dict().items() if v is not None}
        if not fields:
            raise HTTPException(400, "No hay campos para actualizar")

        # Nombres SIEMPRE en MAYÚSCULA también al editar (cliente + comercial),
        # para que una edición no vuelva a dejarlos en minúscula/mixta.
        for _campo in ("cliente", "realizada", "formalizada"):
            if fields.get(_campo):
                fields[_campo] = re.sub(r"\s+", " ", str(fields[_campo]).strip()).upper()

        # Fields tracked for history
        TRACKED = {"respuesta", "estado", "valor", "facturacion", "mes_aceptado",
                   "seguimiento", "no_factura", "valor_facturado", "valor_aprobado"}

        with get_conn() as conn:
            cur = conn.cursor()
            # Fetch current state before update — all tracked fields
            cur.execute("SELECT estado, respuesta, valor, facturacion, mes_aceptado, "
                        "seguimiento, no_factura, valor_facturado, valor_aprobado, num FROM ofertas WHERE id = %s",
                        (oferta_id,))
            prev = fetchone(cur)
            if prev is None:
                raise HTTPException(404, "Oferta no encontrada")

        if "pdf_data" in fields and isinstance(fields["pdf_data"], dict):
            fields["pdf_data"] = json.dumps(fields["pdf_data"])
        set_clause = ", ".join(f"{k} = %s" for k in fields)
        values = list(fields.values()) + [oferta_id]

        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute(f"UPDATE ofertas SET {set_clause} WHERE id = %s RETURNING *", values)
            row = fetchone(cur)
            if row is None:
                raise HTTPException(404, "Oferta no encontrada")

        # Feature 4: Insert historial for tracked fields that changed
        usuario = ""
        try:
            usuario = request.state.user.get("nombre", "") if hasattr(request.state, "user") else ""
        except Exception:
            pass
        oferta_num = prev.get("num", "")
        hist_entries = []
        for campo in TRACKED:
            if campo in fields:
                ant = str(prev.get(campo) or "")
                nuevo = str(fields[campo] or "")
                if ant != nuevo:
                    hist_entries.append((oferta_id, oferta_num, campo, ant, nuevo, usuario))
        if hist_entries:
            try:
                with get_conn() as conn_h:
                    cur_h = conn_h.cursor()
                    for entry in hist_entries:
                        cur_h.execute(
                            """INSERT INTO oferta_historial
                               (oferta_id, oferta_num, campo, valor_ant, valor_nuevo, usuario)
                               VALUES (%s, %s, %s, %s, %s, %s)""",
                            entry
                        )
            except Exception as he:
                print(f"[HISTORIAL] Error guardando historial: {he}")

        # ── Trigger notification when respuesta changes to ACEPTADA ──────────
        nueva_respuesta = (fields.get("respuesta") or "").upper()
        prev_respuesta  = (prev.get("respuesta") or "").upper()
        if nueva_respuesta == "ACEPTADA" and prev_respuesta != "ACEPTADA":
            import threading
            # Sella la fecha de aceptación → el módulo de Aprobadas la pone ARRIBA.
            try:
                with get_conn() as connA:
                    curA = connA.cursor()
                    curA.execute("UPDATE ofertas SET aceptada_fecha = now() WHERE id = %s", (oferta_id,))
            except Exception as _e:
                print(f"[ACEPTADA] No se pudo sellar aceptada_fecha: {_e}")
            pdf_payload = row.get("pdf_data")
            if isinstance(pdf_payload, str):
                try:
                    pdf_payload = json.loads(pdf_payload)
                except Exception:
                    pdf_payload = {}
            pdf_payload = pdf_payload or {}

            # Persist notification in DB.
            # La OSI ya NO se crea automáticamente: la genera el líder desde la
            # tarjeta con el botón "Crear OSI" (formulario pre-llenado).
            try:
                with get_conn() as conn2:
                    cur2 = conn2.cursor()
                    cur2.execute("""
                        INSERT INTO notificaciones (oferta_id, oferta_num, cliente, origen, destino, valor)
                        VALUES (%s,%s,%s,%s,%s,%s)
                    """, (oferta_id, row.get("num",""), row.get("cliente",""),
                          pdf_payload.get("origen",""), pdf_payload.get("destino",""),
                          row.get("valor",0) or 0))
            except Exception as db_exc:
                print(f"[OSI] Error guardando notificación/OSI: {db_exc}")

            # Send email in background
            threading.Thread(
                target=_enviar_notificacion_osi,
                args=(row, pdf_payload),
                daemon=True
            ).start()

        # ── Cleanup when respuesta stops being ACEPTADA ──────────────────────
        # Si la oferta deja de estar aceptada, retira su notificación pendiente
        # y su OSI automática (solo si el líder aún no la ha trabajado).
        elif prev_respuesta == "ACEPTADA" and nueva_respuesta != "ACEPTADA":
            try:
                with get_conn() as conn3:
                    cur3 = conn3.cursor()
                    cur3.execute("DELETE FROM notificaciones WHERE oferta_id=%s", (oferta_id,))
                    n_notif = cur3.rowcount
                    # Solo borra la OSI si sigue siendo la placeholder automática
                    # (sin conductor, sin placa, sin fecha de despacho)
                    cur3.execute("""
                        DELETE FROM osi
                        WHERE oferta_id=%s
                          AND COALESCE(conductor,'')=''
                          AND COALESCE(placa,'')=''
                          AND fecha_despacho IS NULL
                    """, (oferta_id,))
                    n_osi = cur3.rowcount
                    print(f"[OSI] Limpieza por des-aceptar oferta {oferta_id}: "
                          f"{n_notif} notif, {n_osi} OSI borradas")
            except Exception as clean_exc:
                print(f"[OSI] Error limpiando notificación/OSI: {clean_exc}")

        return row
    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


class AnularBody(BaseModel):
    motivo: str


@app.post("/api/ofertas/{oferta_id}/anular")
def anular_oferta(oferta_id: int, body: AnularBody, request: Request):
    """Anula una oferta SIN borrarla (ISO 9001 · trazabilidad). La oferta
    conserva su número: el consecutivo no se libera ni se reutiliza. Queda
    registro del motivo, quién y cuándo, y una entrada en el historial."""
    motivo = (body.motivo or "").strip()
    if not motivo:
        raise HTTPException(400, "Debes indicar el motivo de la anulación")
    usuario = ""
    if hasattr(request.state, "user") and request.state.user:
        usuario = request.state.user.get("nombre", "") or ""
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT num, anulada FROM ofertas WHERE id = %s", (oferta_id,))
            prev = fetchone(cur)
            if prev is None:
                raise HTTPException(404, "Oferta no encontrada")
            if prev.get("anulada"):
                raise HTTPException(409, "La oferta ya está anulada")
            cur.execute("""
                UPDATE ofertas
                   SET anulada = true, estado = 'ANULADA',
                       anulada_motivo = %s, anulada_por = %s, anulada_fecha = now()
                 WHERE id = %s
                 RETURNING *
            """, (motivo, usuario, oferta_id))
            row = fetchone(cur)
            # Traza en el historial de la oferta
            try:
                cur.execute("""
                    INSERT INTO oferta_historial
                        (oferta_id, oferta_num, campo, valor_ant, valor_nuevo, usuario)
                    VALUES (%s, %s, 'ANULACIÓN', 'VIGENTE', %s, %s)
                """, (oferta_id, prev.get("num", ""), f"ANULADA — {motivo}", usuario))
            except Exception as he:
                print(f"[ANULAR] No se pudo guardar historial: {he}")
        return row
    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


@app.delete("/api/ofertas/{oferta_id}")
def delete_oferta(oferta_id: int, request: Request):
    # Borrado FÍSICO: solo admin, y solo para casos excepcionales. Para el uso
    # normal se debe ANULAR (conserva el número y la trazabilidad ISO).
    rol = ""
    if hasattr(request.state, "user") and request.state.user:
        rol = request.state.user.get("rol", "") or ""
    if rol != "admin":
        raise HTTPException(403, "Solo un administrador puede eliminar. Usa 'Anular' para conservar el número.")
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("DELETE FROM ofertas WHERE id = %s RETURNING id", (oferta_id,))
            row = cur.fetchone()
            if row is None:
                raise HTTPException(404, "Oferta no encontrada")
            return {"deleted": oferta_id}
    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


@app.get("/api/ofertas/{oferta_id}/pdf")
def download_oferta_pdf(oferta_id: int):
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT pdf_data, num, cliente FROM ofertas WHERE id = %s", (oferta_id,))
            row = fetchone(cur)
        if row is None:
            raise HTTPException(404, "Oferta no encontrada")
        stored = row.get("pdf_data")
        if not stored:
            raise HTTPException(404, "Esta oferta no tiene datos de PDF guardados. Regenera el PDF desde Generar Oferta.")
        if isinstance(stored, str):
            payload = json.loads(stored)
        else:
            payload = stored
        # Ofertas del modo IA avanzada guardan el HTML completo ya renderizado.
        if payload.get("ia_html"):
            pdf_bytes = _html_to_pdf_bytes(_limpiar_oferta_html(payload["ia_html"]))
        else:
            payload["equipos"] = [e for e in (payload.get("equipos") or []) if e.get("equipo") or e.get("cant")]
            payload["cargo_items"] = [c for c in (payload.get("cargo_items") or []) if c.get("descripcion") or c.get("dimensiones")]
            pdf_bytes = generar_pdf_oferta(payload)
        ref_fmt = _fmt_ref(row.get("num") or oferta_id)
        cliente_slug = re.sub(r"[^a-zA-Z0-9]", "_", (row.get("cliente") or "BOOM"))[:20]
        filename = f"Oferta_{ref_fmt}_{cliente_slug}.pdf"
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


@app.post("/api/extraer-info")
def extraer_info(body: TextoCliente):
    try:
        return _extraer_info(body.texto)
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/api/extraer-info-claude")
def extraer_info_claude(body: TextoCliente):
    try:
        return _extraer_info_claude(body.texto)
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


@app.post("/api/chat-oferta")
def chat_oferta_endpoint(body: ChatOfertaBody):
    try:
        return _chat_oferta(body.messages)
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


_MAX_HISTORY   = 8       # turnos máximos de historial enviados a la API
_MAX_TEXT_CHARS = 45000  # caracteres máx por bloque de texto adjunto (permite adjuntar una oferta completa para replicarla)

def _trim_api_messages(messages: list) -> list:
    """Recorta el historial a los últimos _MAX_HISTORY mensajes y trunca bloques de texto grandes."""
    msgs = messages[-_MAX_HISTORY:] if len(messages) > _MAX_HISTORY else messages

    def _trim_content(c):
        if isinstance(c, str):
            return c[:_MAX_TEXT_CHARS] + "\n[…truncado]" if len(c) > _MAX_TEXT_CHARS else c
        if isinstance(c, list):
            out = []
            for blk in c:
                if isinstance(blk, dict) and blk.get("type") == "text":
                    t = blk.get("text", "")
                    if len(t) > _MAX_TEXT_CHARS:
                        blk = {**blk, "text": t[:_MAX_TEXT_CHARS] + "\n[…truncado]"}
                out.append(blk)
            return out
        return c

    return [{"role": m["role"], "content": _trim_content(m["content"])} for m in msgs]


@app.post("/api/chat-oferta-stream")
def chat_oferta_stream(body: ChatOfertaBody):
    """Streaming version — devuelve tokens SSE en tiempo real."""
    def generate():
        if not ANTHROPIC_OK or not ANTHROPIC_API_KEY:
            yield f"data: {json.dumps({'err': 'API no configurada'})}\n\n"
            return
        try:
            client = _anthropic_client()
            api_messages = _trim_api_messages(
                [{"role": m.role, "content": m.content} for m in body.messages]
            )

            full_text  = ""
            datos_seen = False

            with client.messages.stream(
                model="claude-haiku-4-5-20251001",
                max_tokens=2000,
                system=CHAT_SYSTEM_PROMPT,
                messages=api_messages,
            ) as stream:
                for chunk in stream.text_stream:
                    full_text += chunk
                    if not datos_seen:
                        if "<<<DATOS>>>" in full_text:
                            datos_seen = True
                            # emitir solo el texto previo al marcador
                            pre = full_text[:full_text.index("<<<DATOS>>>")]
                            # ya emitimos parte del texto en chunks anteriores;
                            # el frontend acumuló todo — no emitir más texto
                        else:
                            yield f"data: {json.dumps({'t': chunk})}\n\n"

            # Parsear campos al terminar
            fields = None
            reply  = full_text
            s = full_text.find("<<<DATOS>>>")
            e2 = full_text.find("<<<FIN>>>")
            if s != -1 and e2 != -1:
                reply = full_text[:s].strip()
                try:
                    fields = json.loads(full_text[s + len("<<<DATOS>>>"):e2].strip())
                except Exception:
                    fields = None

            yield f"data: {json.dumps({'done': True, 'reply': reply, 'fields': fields})}\n\n"

        except Exception as exc:
            traceback.print_exc()
            yield f"data: {json.dumps({'err': str(exc)})}\n\n"

    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.post("/api/parsear-detalle")
def parsear_detalle(body: ParseDetalle):
    try:
        return {"equipos": _parsear_detalle(body.texto)}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/api/auto-notas")
def api_auto_notas(data: AutoNotasRequest):
    try:
        equipos = [e.dict() for e in (data.equipos or [])]
        return {"notas": _auto_notas(
            equipos,
            data.texto_cliente or "",
            data.origen or "",
            data.destino or "",
        )}
    except Exception as e:
        raise HTTPException(500, str(e))


_PDF_PAGE_CSS = """
<style>
@page { size: A4; margin: 1.5cm 1.5cm 2cm 1.5cm; }
@media print {
  body { padding: 0 !important; background: white !important; }
  .wrapper { box-shadow: none !important; border-radius: 0 !important; max-width: 100% !important; }
}
</style>"""

# Ajustes para xhtml2pdf (no soporta CSS Grid ni Google Fonts externos)
_XHTML2PDF_CSS_FIXES = """
<style>
/* Reemplaza grid por tabla para spec-card */
.spec-grid { display: table; width: 100%; border-spacing: 6px; }
.spec-card  { display: table-cell; background: #f0f4fa; border-radius: 5px;
              padding: 9px 8px; text-align: center; width: 25%; }
.spec-card .val { font-size: 13px; font-weight: bold; color: #E8601C; }
.spec-card .lbl { font-size: 10px; color: #666; }
/* Fuentes del sistema como fallback */
body, table, p, ul, li { font-family: Helvetica, Arial, sans-serif !important; }
</style>"""


def _inject_pdf_css(html_str: str, extra_css: str = "") -> str:
    """Elimina fuentes externas e inyecta estilos de impresión."""
    # Quitar link de Google Fonts (no disponible sin red en xhtml2pdf)
    html_str = re.sub(r'<link[^>]+fonts\.googleapis\.com[^>]*>', '', html_str)
    return html_str.replace("</head>", _PDF_PAGE_CSS + extra_css + "\n</head>", 1)


# ── Anexo 1 legal (enlace en cada oferta) ────────────────────────────────────
ANEXO_URL = "https://web-production-73608.up.railway.app/anexo-legal"

def _bloque_anexo_html() -> str:
    """Recuadro con el hipervínculo al Anexo 1 legal. Lleva el sentinel
    <!--ANEXO1--> para no duplicarlo si el HTML ya lo tenía."""
    return (
        '<!--ANEXO1-->'
        '<div style="margin:22px 0 4px 0;padding:14px 16px;border:1px solid #d9e2e6;'
        'border-left:5px solid #E8601C;border-radius:8px;background:#fff8f3;'
        'font-family:Arial,sans-serif;font-size:13px;color:#1B2A4A;line-height:1.5;">'
        '📎 <strong>Anexo 1 &ndash; Condiciones Legales de la Cotizaci&oacute;n.</strong> '
        'Este documento hace parte integral de la presente oferta y regula seguros, '
        'obligaciones de las partes, exclusiones, m&eacute;rito ejecutivo y dem&aacute;s '
        'condiciones legales.<br>'
        f'<a href="{ANEXO_URL}" target="_blank" '
        'style="display:inline-block;margin-top:8px;color:#0e6b7d;font-weight:bold;'
        'text-decoration:underline;">👉 Ver / descargar el Anexo 1 (PDF)</a>'
        '</div>'
    )

def _inject_anexo(html: str) -> str:
    """Inserta el recuadro del Anexo 1 justo encima de la firma (div.footer).
    Idempotente: si el HTML ya trae el sentinel, no lo vuelve a poner."""
    if not html or "<!--ANEXO1-->" in html:
        return html
    bloque = _bloque_anexo_html()
    if '<div class="footer">' in html:
        return html.replace('<div class="footer">', bloque + '\n<div class="footer">', 1)
    if "</body>" in html:
        return html.replace("</body>", bloque + "\n</body>", 1)
    return html + bloque


def _html_to_pdf_bytes(html_str: str) -> bytes:
    """Convierte HTML → PDF. Prioridad: WeasyPrint > xhtml2pdf."""
    html_str = _inject_anexo(html_str)
    if WEASYPRINT_OK:
        prepared = _inject_pdf_css(html_str)
        return _weasyprint.HTML(string=prepared).write_pdf()

    if XHTML2PDF_OK:
        prepared = _inject_pdf_css(html_str, _XHTML2PDF_CSS_FIXES)
        buf = BytesIO()
        result = _pisa.CreatePDF(prepared, dest=buf, encoding="utf-8")
        if result.err:
            raise RuntimeError(f"xhtml2pdf: {result.err}")
        return buf.getvalue()

    raise RuntimeError("No hay motor PDF disponible. Instala weasyprint o xhtml2pdf.")


@app.post("/api/generar-pdf")
def generar_pdf(data: OfertaHtml):
    try:
        payload = data.dict()
        payload["equipos"] = [e for e in (payload.get("equipos") or [])
                               if e.get("equipo") or e.get("cant")]
        payload["cargo_items"] = [c for c in (payload.get("cargo_items") or [])
                                   if c.get("descripcion") or c.get("dimensiones")]

        ref_fmt = _fmt_ref(data.ref or "260001")
        cliente_slug = re.sub(r"[^a-zA-Z0-9]", "_", (data.cliente or "BOOM"))[:20]
        filename = f"Oferta_{ref_fmt}_{cliente_slug}.pdf"

        if WEASYPRINT_OK or XHTML2PDF_OK:
            # PDF generado desde el mismo HTML del correo — formato idéntico ✓
            html_str  = generar_html_oferta(payload)
            pdf_bytes = _html_to_pdf_bytes(html_str)
        else:
            # Fallback: PDF ReportLab (formato anterior)
            pdf_bytes = generar_pdf_oferta(payload)

        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


@app.post("/api/generar-html")
def generar_html(data: OfertaHtml):
    try:
        payload = data.dict()
        payload["equipos"] = [e for e in (payload.get("equipos") or [])
                               if e.get("equipo") or e.get("cant")]
        payload["cargo_items"] = [c for c in (payload.get("cargo_items") or [])
                                   if c.get("descripcion") or c.get("dimensiones")]
        html_str = generar_html_oferta(payload)
        ref_fmt = _fmt_ref(data.ref or "260001")
        cliente_slug = re.sub(r"[^a-zA-Z0-9]", "_", (data.cliente or "BOOM"))[:20]
        filename = f"Oferta_{ref_fmt}_{cliente_slug}.html"
        return Response(
            content=html_str.encode("utf-8"),
            media_type="text/html",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


class HtmlToPdf(BaseModel):
    html: str
    filename: Optional[str] = "Oferta_BOOM.pdf"


@app.post("/api/oferta-ia")
def oferta_ia_endpoint(body: OfertaIABody):
    """Modo avanzado: la IA arma la oferta HTML completa (compleja, ES/EN, USD, alertas…).
    Devuelve {reply, html, meta, ref}. El número de referencia lo asigna la plataforma."""
    try:
        ref = body.ref
        if not ref:
            with get_conn() as conn:
                cur = conn.cursor()
                cur.execute("SELECT COALESCE(MAX(CAST(num AS INTEGER)), 260000) + 1 AS next FROM ofertas WHERE NOT COALESCE(es_prueba, false)")
                ref = str(fetchone(cur)["next"])
        return _oferta_ia(body.messages, body.fotos or [], ref, body.firmante, body.forma_pago)
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


def _fmt_cop_py(v):
    """Formatea un entero en pesos estilo es-CO ($ 19.300.000)."""
    try:
        return "$ " + "{:,}".format(int(v or 0)).replace(",", ".")
    except Exception:
        return "$ 0"


class OfertaVersionBody(BaseModel):
    num: str
    cliente: Optional[str] = None
    valor: Optional[int] = 0
    moneda: Optional[str] = "COP"
    descripcion: Optional[str] = None
    origen: Optional[str] = None           # ruta para pre-llenar la OSI
    destino: Optional[str] = None
    forma_pago: Optional[str] = None
    tipo: Optional[str] = None
    sector: Optional[str] = None
    unidad: Optional[str] = None
    estado: Optional[str] = "ENVIADO"
    respuesta: Optional[str] = None
    seguimiento: Optional[str] = None
    mes: Optional[str] = None
    fecha: Optional[str] = None
    realizada: Optional[str] = None
    html: Optional[str] = None            # documento de la oferta (ia_html)
    pdf_original: Optional[str] = None    # base64 del PDF original (para v1 externa)
    resumen: Optional[str] = None         # nota manual opcional de qué cambió
    nueva: Optional[bool] = False         # True = oferta NUEVA (no se está versionando una existente)


@app.post("/api/ofertas/guardar-version")
def guardar_version(body: OfertaVersionBody, request: Request):
    """Guarda una oferta desde el asistente COMO NUEVA VERSIÓN, respetando el
    número. Si el número ya existe (interna o importada del Excel) actualiza esa
    misma oferta; si no existe, la crea con ese número. Cada llamada añade una
    versión (v1, v2, …), marca la nueva como vigente y actualiza el valor de la
    oferta al de la versión vigente (cambio de valor automático)."""
    try:
        num = re.sub(r"\D", "", body.num or "")
        if not num:
            raise HTTPException(400, "Falta el número de oferta")
        usuario = ""
        try:
            usuario = request.state.user.get("nombre", "") if hasattr(request.state, "user") else ""
        except Exception:
            pass
        cliente = re.sub(r"\s+", " ", (body.cliente or "").strip()).upper() or "SIN CLIENTE"
        realizada = re.sub(r"\s+", " ", (body.realizada or usuario or "").strip()).upper() or None
        pdf_data = {
            "ref": num, "cliente": cliente, "descripcion": body.descripcion,
            "moneda": body.moneda or "COP", "modo": "ia",
            "forma_pago": body.forma_pago, "ia_html": body.html or "",
            "origen": (body.origen or "").strip(), "destino": (body.destino or "").strip(),
        }
        nuevo_valor = int(body.valor or 0)
        # Candado de consecutivo: si esta es una oferta NUEVA y el número que se
        # mostró ya lo tomó otra oferta mientras se redactaba (o choca al insertar),
        # se reasigna el siguiente número libre y se reintenta. Al VERSIONAR una
        # oferta existente (body.nueva = False) el número manda y se respeta.
        MAX_INTENTOS = 8
        _rechoque = False
        for _intento in range(MAX_INTENTOS):
          try:
            with get_conn() as conn:
                cur = conn.cursor()
                if _rechoque:
                    cur.execute("SELECT COALESCE(MAX(CAST(num AS INTEGER)), 260000) + 1 AS n FROM ofertas WHERE NOT COALESCE(es_prueba, false)")
                    num = str(fetchone(cur)["n"])
                cur.execute("SELECT id, valor FROM ofertas WHERE num = %s", (num,))
                ex = fetchone(cur)
                if ex and body.nueva and not _rechoque:
                    # El consecutivo mostrado ya pertenece a OTRA oferta: no la
                    # sobreescribas; toma el siguiente número libre para esta nueva.
                    cur.execute("SELECT COALESCE(MAX(CAST(num AS INTEGER)), 260000) + 1 AS n FROM ofertas WHERE NOT COALESCE(es_prueba, false)")
                    num = str(fetchone(cur)["n"])
                    cur.execute("SELECT id, valor FROM ofertas WHERE num = %s", (num,))
                    ex = fetchone(cur)
                pdf_data["ref"] = num
                if ex:
                    oferta_id = ex["id"]
                    valor_ant = int(ex.get("valor") or 0)
                    creada = False
                    cur.execute(
                        """UPDATE ofertas SET
                             cliente = %s, valor = %s, moneda = %s,
                             mes = COALESCE(%s, mes), tipo = COALESCE(%s, tipo),
                             sector = COALESCE(%s, sector), unidad = COALESCE(%s, unidad),
                             estado = COALESCE(%s, estado), respuesta = COALESCE(%s, respuesta),
                             seguimiento = COALESCE(%s, seguimiento), general = COALESCE(%s, general),
                             pdf_data = %s
                           WHERE id = %s""",
                        (cliente, nuevo_valor, (body.moneda or "COP"), body.mes, body.tipo, body.sector, body.unidad,
                         body.estado, body.respuesta, body.seguimiento, body.descripcion,
                         json.dumps(pdf_data), oferta_id),
                    )
                else:
                    valor_ant = None
                    creada = True
                    cur.execute(
                        """INSERT INTO ofertas
                             (num, mes, fecha, cliente, realizada, tipo, sector, unidad,
                              valor, moneda, estado, respuesta, seguimiento, general, pdf_data)
                           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                           RETURNING id""",
                        (num, body.mes, body.fecha or None, cliente, realizada, body.tipo,
                         body.sector, body.unidad, nuevo_valor, (body.moneda or "COP"), body.estado or "ENVIADO",
                         body.respuesta, body.seguimiento, body.descripcion, json.dumps(pdf_data)),
                    )
                    oferta_id = fetchone(cur)["id"]
                    cur.execute(
                        """INSERT INTO clientes (nombre_corto)
                           SELECT %s WHERE NOT EXISTS (
                               SELECT 1 FROM clientes WHERE lower(nombre_corto) = lower(%s))""",
                        (cliente, cliente),
                    )

                cur.execute("SELECT COALESCE(MAX(version), 0) + 1 AS n FROM oferta_versiones WHERE oferta_id = %s",
                            (oferta_id,))
                vnum = int(fetchone(cur)["n"])

                resumen = (body.resumen or "").strip()
                cambio_valor = (valor_ant is not None and nuevo_valor != valor_ant)
                valor_txt = ("Valor %s → %s." % (_fmt_cop_py(valor_ant), _fmt_cop_py(nuevo_valor))) if cambio_valor else ""
                if not resumen:
                    if vnum == 1:
                        resumen = "Versión inicial." + (" PDF original adjunto." if body.pdf_original else "")
                    elif cambio_valor:
                        resumen = valor_txt
                    else:
                        resumen = "Ajuste de la cotización."
                elif cambio_valor and _fmt_cop_py(nuevo_valor) not in resumen:
                    # Garantiza la trazabilidad del cambio de valor aunque haya nota manual.
                    resumen = resumen + " — " + valor_txt

                cur.execute("UPDATE oferta_versiones SET vigente = false WHERE oferta_id = %s", (oferta_id,))
                cur.execute(
                    """INSERT INTO oferta_versiones
                         (oferta_id, oferta_num, version, valor, moneda, descripcion,
                          forma_pago, html, pdf_b64, resumen, creado_por, vigente)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s, true)
                       RETURNING id, version, created_at""",
                    (oferta_id, num, vnum, nuevo_valor, body.moneda or "COP", body.descripcion,
                     body.forma_pago, body.html or "", body.pdf_original, resumen, usuario),
                )
                vrow = fetchone(cur)

                if valor_ant is not None and nuevo_valor != valor_ant:
                    cur.execute(
                        """INSERT INTO oferta_historial
                             (oferta_id, oferta_num, campo, valor_ant, valor_nuevo, usuario)
                           VALUES (%s,%s,'valor',%s,%s,%s)""",
                        (oferta_id, num, str(valor_ant), str(nuevo_valor), usuario),
                    )
            num_fmt = num.zfill(6)
            num_fmt = num_fmt[:2] + "-" + num_fmt[2:]
            return {"ok": True, "oferta_id": oferta_id, "num": num, "num_fmt": num_fmt,
                    "version": vrow["version"], "creada": creada, "resumen": resumen}
          except pgdb.DatabaseError as e:
            msg = str(e).lower()
            if "unique" in msg or "duplicate" in msg:
                # Choque de consecutivo (otro usuario tomó ese número casi al mismo
                # tiempo): recalcula el siguiente número libre y reintenta.
                _rechoque = True
                continue
            traceback.print_exc()
            raise HTTPException(500, str(e))
          except HTTPException:
            raise
          except Exception as e:
            traceback.print_exc()
            raise HTTPException(500, str(e))
        raise HTTPException(409, "No se pudo asignar un consecutivo único; intenta guardar de nuevo")
    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


@app.get("/api/ofertas/{oferta_id}/versiones")
def listar_versiones(oferta_id: int):
    """Lista las versiones de una oferta (más reciente primero)."""
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute(
                """SELECT id, version, valor, moneda, resumen, creado_por, vigente, created_at,
                          (html IS NOT NULL AND html <> '')     AS tiene_html,
                          (pdf_b64 IS NOT NULL AND pdf_b64 <> '') AS tiene_pdf
                     FROM oferta_versiones WHERE oferta_id = %s
                     ORDER BY version DESC""",
                (oferta_id,),
            )
            return fetchall(cur)
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


@app.get("/api/ofertas/{oferta_id}/versiones/{ver_id}/documento")
def ver_version_documento(oferta_id: int, ver_id: int):
    """Devuelve el HTML del documento de una versión para previsualizar."""
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT html FROM oferta_versiones WHERE id = %s AND oferta_id = %s",
                        (ver_id, oferta_id))
            r = fetchone(cur)
        if not r:
            raise HTTPException(404, "Versión no encontrada")
        html = r.get("html") or "<p style='font-family:sans-serif;padding:24px'>Esta versión no tiene documento HTML (solo PDF original o datos).</p>"
        return Response(content=html, media_type="text/html; charset=utf-8")
    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


@app.get("/api/ofertas/{oferta_id}/versiones/{ver_id}/pdf")
def ver_version_pdf(oferta_id: int, ver_id: int):
    """Descarga/abre el PDF de una versión: usa el PDF original subido si existe,
    o convierte el HTML de la versión a PDF."""
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT version, pdf_b64, html FROM oferta_versiones WHERE id = %s AND oferta_id = %s",
                        (ver_id, oferta_id))
            r = fetchone(cur)
        if not r:
            raise HTTPException(404, "Versión no encontrada")
        if r.get("pdf_b64"):
            data = base64.b64decode(r["pdf_b64"].split(",", 1)[-1])
        elif r.get("html"):
            data = _html_to_pdf_bytes(r["html"])
        else:
            raise HTTPException(404, "Esta versión no tiene documento")
        fn = "Oferta_v%s.pdf" % r["version"]
        return Response(content=data, media_type="application/pdf",
                        headers={"Content-Disposition": 'inline; filename="%s"' % fn})
    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


@app.post("/api/html-a-pdf")
def html_a_pdf(body: HtmlToPdf):
    """Convierte un HTML de oferta ya generado (modo IA avanzada) en PDF descargable."""
    try:
        # 🛡️ No permitir generar PDF sin nombre de cliente (el filename lo codifica).
        _fn_up = (body.filename or "").upper()
        if "_CLIENTE." in _fn_up or "SIN_CLIENTE" in _fn_up:
            raise HTTPException(status_code=400, detail="La oferta no tiene nombre de cliente. Escríbelo antes de generar el PDF.")
        pdf_bytes = _html_to_pdf_bytes(body.html)
        fn = re.sub(r'[^A-Za-z0-9_.\-]', "_", body.filename or "Oferta_BOOM.pdf")
        if not fn.lower().endswith(".pdf"):
            fn += ".pdf"
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{fn}"'},
        )
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


@app.get("/api/stats")
def get_stats():
    try:
        with get_conn() as conn:
            def q(sql, params=None):
                cur = conn.cursor(); cur.execute(sql, params or ()); return fetchall(cur)
            def q1(sql, params=None):
                cur = conn.cursor(); cur.execute(sql, params or ()); return fetchone(cur)

            # Las ofertas ANULADAS no cuentan en el dashboard (conservan el número
            # para trazabilidad, pero no suman en KPIs ni en valor ofertado).
            total     = q1("SELECT COUNT(*) AS total FROM ofertas WHERE NOT COALESCE(anulada,false)")["total"]
            aceptadas = q1("SELECT COUNT(*) AS total FROM ofertas WHERE UPPER(respuesta)='ACEPTADA' AND NOT COALESCE(anulada,false)")["total"]
            seguim    = q1("SELECT COUNT(*) AS total FROM ofertas WHERE UPPER(respuesta)='EN SEGUIMIENTO' AND NOT COALESCE(anulada,false)")["total"]
            valor     = q1("SELECT COALESCE(SUM(valor),0) AS total FROM ofertas WHERE NOT COALESCE(anulada,false)")["total"]
            meses_es  = {"January":"ENERO","February":"FEBRERO","March":"MARZO",
                          "April":"ABRIL","May":"MAYO","June":"JUNIO",
                          "July":"JULIO","August":"AGOSTO","September":"SEPTIEMBRE",
                          "October":"OCTUBRE","November":"NOVIEMBRE","December":"DICIEMBRE"}
            mes_es    = meses_es.get(datetime.now().strftime("%B"), "")
            este_mes  = q1("SELECT COUNT(*) AS total FROM ofertas WHERE mes=%s AND NOT COALESCE(anulada,false)", (mes_es,))["total"]
            return {
                "total": total,
                "aceptadas": aceptadas,
                "tasa_aceptacion": round(aceptadas*100/total, 1) if total else 0,
                "seguimiento": seguim,
                "valor_total": valor,
                "este_mes": este_mes,
                "aceptadas_por_cliente": q("SELECT cliente, COUNT(*) AS cnt FROM ofertas WHERE UPPER(respuesta)='ACEPTADA' AND NOT COALESCE(anulada,false) AND cliente IS NOT NULL AND cliente<>'' GROUP BY cliente ORDER BY cnt DESC LIMIT 10"),
                "por_tipo":      q("SELECT tipo, COUNT(*) AS cnt FROM ofertas WHERE NOT COALESCE(anulada,false) AND tipo IS NOT NULL AND tipo<>'' GROUP BY tipo ORDER BY cnt DESC"),
                "por_cliente":   q("SELECT cliente, COUNT(*) AS cnt FROM ofertas WHERE NOT COALESCE(anulada,false) AND cliente IS NOT NULL AND cliente<>'' GROUP BY cliente ORDER BY cnt DESC LIMIT 10"),
                "por_unidad":    q("SELECT unidad, COUNT(*) AS cnt FROM ofertas WHERE NOT COALESCE(anulada,false) AND unidad IS NOT NULL AND unidad<>'' GROUP BY unidad ORDER BY cnt DESC"),
                "por_mes":       q("SELECT mes, COUNT(*) AS cnt FROM ofertas WHERE NOT COALESCE(anulada,false) AND mes IS NOT NULL GROUP BY mes"),
                "ultimas":       q("SELECT * FROM ofertas WHERE NOT COALESCE(anulada,false) ORDER BY CAST(num AS INTEGER) DESC LIMIT 8"),
            }
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


# ── TEMP diagnóstico solo lectura: todas las ofertas de TIBA ─────────────────
@app.get("/api/_diagtiba2")
def _diag_tiba2(token: str = ""):
    if token != "k34dZAAnkqpSF_TZwF5e3V3M9l0WTTXe":
        raise HTTPException(403, "no")
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            # Todas las ofertas de TIBA (por nombre)
            cur.execute("""
                SELECT num, fecha, cliente, respuesta, estado, valor, realizada,
                       COALESCE(anulada,false) AS anulada, COALESCE(es_prueba,false) AS es_prueba
                FROM ofertas
                WHERE UPPER(cliente) LIKE '%TIBA%'
                ORDER BY UPPER(respuesta), CAST(num AS INTEGER)
            """)
            tiba = fetchall(cur)
            aceptadas = [r for r in tiba if (r.get("respuesta") or "").upper() == "ACEPTADA"]
            return {
                "ok": True,
                "total_tiba": len(tiba),
                "aceptadas_count": len(aceptadas),
                "aceptadas": aceptadas,
                "todas": tiba,
            }
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


# ── Auth endpoints ────────────────────────────────────────────────────────────
@app.post("/auth/login")
def login(body: LoginBody, response: Response):
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT id, username, nombre, password_hash, rol, activo, modulos FROM usuarios WHERE username = %s",
                (body.username.lower().strip(),)
            )
            row = fetchone(cur)
        if not row or not row.get("activo") or not _verify_pw(body.password, row["password_hash"]):
            raise HTTPException(401, "Usuario o contraseña incorrectos")
        token = secrets.token_hex(32)
        modulos = row.get("modulos") or []
        user_data = {"id": row["id"], "username": row["username"],
                     "nombre": row["nombre"], "rol": row["rol"], "modulos": modulos}
        _session_save(token, user_data, max_age_s=86400 * 7)
        _activity_touch(token)   # arranca el reloj de inactividad
        response.set_cookie("boom_session", token, httponly=True, samesite="lax",
                            max_age=86400 * 7)
        return {"nombre": row["nombre"], "rol": row["rol"], "username": row["username"], "modulos": modulos}
    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


@app.post("/auth/logout")
def logout(request: Request, response: Response):
    token = request.cookies.get("boom_session")
    if token:
        _session_delete(token)
    response.delete_cookie("boom_session")
    return {"ok": True}


@app.get("/auth/me")
def me(request: Request):
    token = request.cookies.get("boom_session")
    user = _session_get(token)
    # Consultar el estado NO cuenta como actividad: solo verifica la inactividad.
    if user and _activity_expired(token):
        _session_delete(token)
        user = None
    if not user:
        raise HTTPException(401, "No autenticado")
    return user


@app.post("/auth/ping")
def auth_ping(request: Request):
    """El navegador llama a esto ante actividad real del usuario (throttled) para
    renovar el reloj de inactividad. Si ya venció, cierra la sesión."""
    token = request.cookies.get("boom_session")
    user = _session_get(token)
    if not user:
        return JSONResponse({"activo": False}, status_code=401)
    if _activity_expired(token):
        _session_delete(token)
        return JSONResponse({"activo": False}, status_code=401)
    _activity_touch(token)
    return {"activo": True}


# ── Gestión de usuarios ───────────────────────────────────────────────────────
@app.get("/api/usuarios")
def list_usuarios(request: Request):
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT id, username, nombre, rol, activo, creado_en, modulos FROM usuarios ORDER BY id")
            return fetchall(cur)
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/api/usuarios", status_code=201)
def create_usuario(body: UsuarioCreate, request: Request):
    if body.rol not in ("admin", "comercial", "operaciones", "viewer"):
        raise HTTPException(400, "Rol inválido")
    if request.state.user["rol"] != "admin":
        raise HTTPException(403, "Solo administradores pueden crear usuarios")
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            modulos = body.modulos or []
            cur.execute(
                "INSERT INTO usuarios (username, nombre, password_hash, rol, modulos) VALUES (%s,%s,%s,%s,%s) RETURNING id",
                (body.username.lower().strip(), body.nombre, _hash_pw(body.password), body.rol, modulos)
            )
            row = fetchone(cur)
            return {"id": row["id"], "username": body.username.lower(), "nombre": body.nombre,
                    "rol": body.rol, "activo": True, "modulos": modulos}
    except Exception as e:
        if "unique" in str(e).lower():
            raise HTTPException(409, f'El usuario "{body.username}" ya existe')
        raise HTTPException(500, str(e))


@app.put("/api/usuarios/{uid}")
def update_usuario(uid: int, body: UsuarioUpdate, request: Request):
    if request.state.user["rol"] != "admin":
        raise HTTPException(403, "Solo administradores")
    sets, params = [], []
    if body.nombre   is not None: sets.append("nombre = %s");        params.append(body.nombre)
    if body.password is not None: sets.append("password_hash = %s"); params.append(_hash_pw(body.password))
    if body.rol      is not None:
        if body.rol not in ("admin", "comercial", "operaciones", "viewer"):
            raise HTTPException(400, "Rol inválido")
        sets.append("rol = %s"); params.append(body.rol)
    if body.activo   is not None: sets.append("activo = %s");        params.append(body.activo)
    if body.modulos  is not None: sets.append("modulos = %s");       params.append(body.modulos)
    if not sets:
        raise HTTPException(400, "Nada que actualizar")
    params.append(uid)
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute(f"UPDATE usuarios SET {', '.join(sets)} WHERE id = %s", params)
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.delete("/api/usuarios/{uid}")
def delete_usuario(uid: int, request: Request):
    if request.state.user["rol"] != "admin":
        raise HTTPException(403, "Solo administradores")
    if uid == request.state.user["id"]:
        raise HTTPException(400, "No puedes desactivar tu propia cuenta")
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("UPDATE usuarios SET activo = FALSE WHERE id = %s", (uid,))
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))


# ══════════════════════════════════════════════════════════════════════════════
# MÓDULO TI
# ══════════════════════════════════════════════════════════════════════════════

class AreaCreate(BaseModel):
    nombre: str
    descripcion: Optional[str] = ""
    icono: Optional[str] = "🏢"

class AreaUpdate(BaseModel):
    nombre: Optional[str] = None
    descripcion: Optional[str] = None
    icono: Optional[str] = None
    activo: Optional[bool] = None

MODULOS_DISPONIBLES = ["dashboard", "generar", "control", "aprobadas", "operaciones", "tarifario"]


def _require_admin(request: Request):
    if request.state.user["rol"] != "admin":
        raise HTTPException(403, "Solo administradores TI")


@app.get("/api/ti/estado")
def ti_estado(request: Request):
    _require_admin(request)
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) AS total FROM ofertas")
            total_ofertas = fetchone(cur)["total"]
            cur.execute("SELECT COUNT(*) AS total FROM ofertas WHERE estado = 'APROBADO'")
            aprobadas = fetchone(cur)["total"]
            cur.execute("SELECT COUNT(*) AS total FROM usuarios WHERE activo = TRUE")
            usuarios_activos = fetchone(cur)["total"]
            cur.execute("SELECT COUNT(*) AS total FROM areas WHERE activo = TRUE")
            areas_activas = fetchone(cur)["total"]
        sesiones_activas = len(_sessions)
        return {
            "db": "ok",
            "sesiones_activas": sesiones_activas,
            "total_ofertas": total_ofertas,
            "ofertas_aprobadas": aprobadas,
            "usuarios_activos": usuarios_activos,
            "areas_activas": areas_activas,
            "sesiones": [
                {"nombre": v["nombre"], "rol": v["rol"]}
                for v in _sessions.values()
            ]
        }
    except Exception as e:
        return {"db": "error", "detalle": str(e)}


@app.get("/api/ti/areas")
def ti_list_areas(request: Request):
    _require_admin(request)
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM areas ORDER BY id")
        areas = fetchall(cur)
        for a in areas:
            cur.execute("SELECT COUNT(*) AS n FROM usuarios WHERE area = %s AND activo = TRUE", (a["nombre"],))
            a["usuarios_count"] = fetchone(cur)["n"]
            cur.execute("SELECT modulo, activo FROM area_permisos WHERE area_id = %s", (a["id"],))
            permisos = {r["modulo"]: r["activo"] for r in fetchall(cur)}
            a["permisos"] = {m: permisos.get(m, True) for m in MODULOS_DISPONIBLES}
        return areas


@app.post("/api/ti/areas", status_code=201)
def ti_create_area(body: AreaCreate, request: Request):
    _require_admin(request)
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute(
                "INSERT INTO areas (nombre, descripcion, icono) VALUES (%s,%s,%s) RETURNING *",
                (body.nombre.strip(), body.descripcion or "", body.icono or "🏢")
            )
            area = fetchone(cur)
            for m in MODULOS_DISPONIBLES:
                cur.execute(
                    "INSERT INTO area_permisos (area_id, modulo, activo) VALUES (%s,%s,%s) ON CONFLICT DO NOTHING",
                    (area["id"], m, True)
                )
            area["permisos"] = {m: True for m in MODULOS_DISPONIBLES}
            area["usuarios_count"] = 0
            return area
    except Exception as e:
        if "unique" in str(e).lower():
            raise HTTPException(409, f'El área "{body.nombre}" ya existe')
        raise HTTPException(500, str(e))


@app.put("/api/ti/areas/{area_id}")
def ti_update_area(area_id: int, body: AreaUpdate, request: Request):
    _require_admin(request)
    sets, params = [], []
    if body.nombre      is not None: sets.append("nombre = %s");      params.append(body.nombre.strip())
    if body.descripcion is not None: sets.append("descripcion = %s"); params.append(body.descripcion)
    if body.icono       is not None: sets.append("icono = %s");       params.append(body.icono)
    if body.activo      is not None: sets.append("activo = %s");      params.append(body.activo)
    if not sets:
        raise HTTPException(400, "Nada que actualizar")
    params.append(area_id)
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute(f"UPDATE areas SET {', '.join(sets)} WHERE id = %s", params)
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.delete("/api/ti/areas/{area_id}")
def ti_delete_area(area_id: int, request: Request):
    _require_admin(request)
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT nombre FROM areas WHERE id = %s", (area_id,))
            row = fetchone(cur)
            if not row:
                raise HTTPException(404, "Área no encontrada")
            if row["nombre"] == "Comercial":
                raise HTTPException(400, "El área Comercial no se puede eliminar")
            cur.execute("UPDATE usuarios SET area = NULL WHERE area = %s", (row["nombre"],))
            cur.execute("DELETE FROM areas WHERE id = %s", (area_id,))
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@app.put("/api/ti/areas/{area_id}/permisos")
def ti_update_permisos(area_id: int, permisos: dict, request: Request):
    _require_admin(request)
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            for modulo, activo in permisos.items():
                if modulo in MODULOS_DISPONIBLES:
                    cur.execute("""
                        INSERT INTO area_permisos (area_id, modulo, activo)
                        VALUES (%s,%s,%s)
                        ON CONFLICT (area_id, modulo) DO UPDATE SET activo = EXCLUDED.activo
                    """, (area_id, modulo, bool(activo)))
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.put("/api/ti/usuarios/{uid}/area")
def ti_asignar_area(uid: int, body: dict, request: Request):
    _require_admin(request)
    area = body.get("area")
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("UPDATE usuarios SET area = %s WHERE id = %s", (area or None, uid))
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))


# ── Notificaciones ────────────────────────────────────────────────────────────
@app.get("/api/notificaciones")
def get_notificaciones(request: Request):
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT n.*, o.pdf_data AS _pdf
                FROM notificaciones n
                LEFT JOIN ofertas o ON o.id = n.oferta_id
                ORDER BY n.created_at DESC LIMIT 100
            """)
            rows = fetchall(cur)
        # La ruta (origen → destino) y la descripción viven en el pdf_data de la oferta IA.
        for r in rows:
            pdf = r.pop("_pdf", None)
            if isinstance(pdf, str):
                try: pdf = json.loads(pdf)
                except Exception: pdf = {}
            pdf = pdf or {}
            if not r.get("origen"):
                r["origen"] = (pdf.get("origen") or "").strip()
            if not r.get("destino"):
                r["destino"] = (pdf.get("destino") or "").strip()
            r["descripcion"] = (pdf.get("descripcion") or "").strip()
        return rows
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/notificaciones/pendientes")
def notificaciones_pendientes(request: Request):
    """FUENTE DE VERDAD del panel de alertas: TODA oferta ACEPTADA/notificada por la
    plataforma que AÚN NO tiene OSI creada. No depende de la tabla notificaciones ni
    de si una tarjeta fue 'leída' — así NUNCA se pierde una OSI pendiente. La tarjeta
    desaparece sola cuando se le crea la OSI. Origen/destino salen del pdf_data de la
    oferta (misma fuente que usaba el trigger de aceptación)."""
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT o.id AS oferta_id,
                       o.num AS oferta_num,
                       o.cliente,
                       COALESCE(o.valor, 0) AS valor,
                       COALESCE(o.pdf_data->>'origen','') AS origen,
                       COALESCE(o.pdf_data->>'destino','') AS destino,
                       COALESCE(o.pdf_data->>'descripcion','') AS descripcion,
                       COALESCE(o.aceptada_fecha, o.created_at) AS created_at
                FROM ofertas o
                WHERE UPPER(COALESCE(o.respuesta,'')) = 'ACEPTADA'
                  AND NOT COALESCE(o.anulada, false)
                  AND NOT COALESCE(o.es_prueba, false)
                  AND NOT EXISTS (SELECT 1 FROM osi s WHERE s.oferta_id = o.id)
                ORDER BY COALESCE(o.aceptada_fecha, o.created_at) DESC NULLS LAST,
                         CAST(o.num AS INTEGER) DESC
            """)
            rows = fetchall(cur)
        for r in rows:
            if r.get("created_at"):
                r["created_at"] = str(r["created_at"])
        return rows
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


@app.get("/api/notificaciones/count")
def count_notificaciones(request: Request):
    """Cuenta las OSI PENDIENTES (ofertas aceptadas sin OSI) para la campana/badge.
    Se alinea con el panel de alertas (mismo criterio: aceptada y sin OSI)."""
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT COUNT(*) AS total
                FROM ofertas o
                WHERE UPPER(COALESCE(o.respuesta,'')) = 'ACEPTADA'
                  AND NOT COALESCE(o.anulada, false)
                  AND NOT COALESCE(o.es_prueba, false)
                  AND NOT EXISTS (SELECT 1 FROM osi s WHERE s.oferta_id = o.id)
            """)
            row = fetchone(cur)
            count = row["total"] if row else 0
            # La oferta aceptada-sin-OSI más reciente (para el sonido/aviso: saber QUÉ llegó)
            cur.execute("""
                SELECT o.id, o.num, o.cliente
                FROM ofertas o
                WHERE UPPER(COALESCE(o.respuesta,'')) = 'ACEPTADA'
                  AND NOT COALESCE(o.anulada, false)
                  AND NOT COALESCE(o.es_prueba, false)
                  AND NOT EXISTS (SELECT 1 FROM osi s WHERE s.oferta_id = o.id)
                ORDER BY o.id DESC LIMIT 1
            """)
            last = fetchone(cur)
            return {
                "count": count,
                "latest_id": last["id"] if last else 0,
                "latest_num": (last["num"] if last else "") or "",
                "latest_cliente": (last["cliente"] if last else "") or "",
            }
    except Exception as e:
        raise HTTPException(500, str(e))


@app.patch("/api/notificaciones/{nid}/leer")
def marcar_leida(nid: int, request: Request):
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("UPDATE notificaciones SET leida = true WHERE id = %s", (nid,))
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.patch("/api/notificaciones/leer-todas")
def marcar_todas_leidas(request: Request):
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("UPDATE notificaciones SET leida = true")
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/api/notificaciones/backfill-aceptadas-sin-osi")
def backfill_notificaciones(request: Request):
    """Crea la TARJETA de alerta (notificación pendiente) para toda oferta que ya
    está ACEPTADA/notificada por la plataforma y AÚN NO tiene OSI creada, pero que
    hoy no tiene tarjeta (p. ej. aceptadas por importación masiva, que no pasan por
    el PATCH que genera la tarjeta). Es IDEMPOTENTE: nunca duplica una tarjeta
    existente, no marca nada como leído ni borra nada. Solo agrega lo que falta.
    Origen/destino se leen del pdf_data de la oferta, igual que el trigger de
    aceptación."""
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("""
                INSERT INTO notificaciones (oferta_id, oferta_num, cliente, origen, destino, valor)
                SELECT o.id, o.num, o.cliente,
                       COALESCE(o.pdf_data->>'origen',''),
                       COALESCE(o.pdf_data->>'destino',''),
                       COALESCE(o.valor, 0)
                FROM ofertas o
                WHERE UPPER(COALESCE(o.respuesta,'')) = 'ACEPTADA'
                  AND NOT COALESCE(o.anulada, false)
                  AND NOT EXISTS (SELECT 1 FROM osi s WHERE s.oferta_id = o.id)
                  AND NOT EXISTS (SELECT 1 FROM notificaciones n WHERE n.oferta_id = o.id)
            """)
            creadas = cur.rowcount or 0
        return {"ok": True, "creadas": creadas}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


# ══════════════════════════════════════════════════════════════════════════
# 🚛 EQUIPOS — catálogo (propios + subcontratos) del módulo Operaciones
# ══════════════════════════════════════════════════════════════════════════
class EquipoBody(BaseModel):
    placa: str = ""
    tipo: str = ""
    propiedad: str = "propio"
    tenedor: str = ""
    conductor: str = ""
    celular: str = ""
    estado: str = "DISPONIBLE"
    # Ficha de flota propia (Fase 1)
    codigo: str = ""
    categoria: str = ""
    marca: str = ""
    clase: str = ""
    config: str = ""
    ejes: str = ""
    ancho: str = ""
    largo: str = ""
    alto: str = ""
    capacidad: str = ""
    peso: str = ""
    descripcion: str = ""
    activo: bool = True


_EQUIPO_COLS = ("placa, tipo, propiedad, tenedor, conductor, celular, estado, "
                "codigo, categoria, marca, clase, config, ejes, ancho, largo, alto, "
                "capacidad, peso, descripcion, activo")


def _equipo_values(body: "EquipoBody"):
    return (body.placa.strip().upper(), body.tipo.strip(), body.propiedad.strip(),
            body.tenedor.strip(), body.conductor.strip(), body.celular.strip(),
            body.estado.strip(), body.codigo.strip().upper(), body.categoria.strip().upper(),
            body.marca.strip().upper(), body.clase.strip(), body.config.strip().upper(),
            body.ejes.strip(), body.ancho.strip(), body.largo.strip(), body.alto.strip(),
            body.capacidad.strip(), body.peso.strip(), body.descripcion.strip(), bool(body.activo))


@app.get("/api/equipos")
def listar_equipos(request: Request):
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT id, %s
                FROM equipos ORDER BY categoria, codigo, placa
            """ % _EQUIPO_COLS)
            cols = [d[0] for d in cur.description]
            return [dict(zip(cols, r)) for r in cur.fetchall()]
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/api/equipos")
def crear_equipo(body: EquipoBody, request: Request):
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            ph = ",".join(["%s"] * 20)
            cur.execute("INSERT INTO equipos (%s) VALUES (%s) RETURNING id" % (_EQUIPO_COLS, ph),
                        _equipo_values(body))
            new_id = fetchone(cur)["id"]
        return {"ok": True, "id": new_id}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.put("/api/equipos/{eid}")
def actualizar_equipo(eid: int, body: EquipoBody, request: Request):
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            sets = ",".join("%s=%%s" % c.strip() for c in _EQUIPO_COLS.split(","))
            cur.execute("UPDATE equipos SET %s WHERE id=%%s" % sets, _equipo_values(body) + (eid,))
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.delete("/api/equipos/{eid}")
def eliminar_equipo(eid: int, request: Request):
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("DELETE FROM equipos WHERE id=%s", (eid,))
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))


# ══════════════════════════════════════════════════════════════════════════
# 📋 OSI — creación manual por el líder (formato real O26XXX)
# ══════════════════════════════════════════════════════════════════════════
class CrearOSIBody(BaseModel):
    oferta_id: int = 0
    oferta_num: str = ""
    cliente: str = ""
    nit: str = ""
    solicitante: str = ""
    contacto_reporte_det: dict = {}
    contacto_cargue_det: dict = {}
    contacto_descargue_det: dict = {}
    lider: str = ""
    tipo_operacion: str = ""
    tipo_carga: str = ""
    tipo_carga_otro: str = ""
    especificacion: str = ""
    origen: str = ""
    destino: str = ""
    lugar_cargue: str = ""
    lugar_descargue: str = ""
    fecha_inicio: str = ""
    hora_servicio: str = ""
    fecha_final: str = ""
    equipo: str = ""
    equipos_asignados: list = []
    operadores: list = []
    conductor: str = ""
    auxiliares: str = ""
    placa: str = ""
    req_sup: bool = False
    req_hse: bool = False
    req_otro: str = ""
    escoltas: str = ""
    dim_ancho: str = ""
    dim_alto: str = ""
    dim_largo: str = ""
    dim_voladizo: str = ""
    dim_peso: str = ""
    via_tipo: str = ""
    via_obstruccion: bool = False
    packing_list: str = ""
    valor: int = 0
    observaciones: str = ""


def _proximo_numero_osi(cur) -> str:
    """Genera el próximo consecutivo con el formato real O26XXX.
    Continúa desde el último real del Excel (O26276)."""
    cur.execute("SELECT numero_osi FROM osi WHERE numero_osi ~ '^O26[0-9]+$'")
    maxn = 276  # piso: último real cargado en el Excel de Proyectos
    for r in cur.fetchall():
        try:
            n = int(r[0][3:])   # después de 'O26'
            if n > maxn:
                maxn = n
        except Exception:
            pass
    return f"O26{str(maxn + 1).zfill(3)}"


@app.post("/api/osi/crear")
def crear_osi(body: CrearOSIBody, request: Request):
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            # Candado de sesión: si dos líderes generan OSI al mismo tiempo, el
            # segundo espera a que el primero termine y toma el siguiente número.
            # Así nunca chocan ni sale error de "número duplicado".
            cur.execute("SELECT pg_advisory_xact_lock(720261)")
            numero_osi = _proximo_numero_osi(cur)
            def _d(v):
                v = (v or "").strip()
                return v or None
            # El VALOR de la OSI se lee directo de la oferta en la base de datos
            # (no del dato que viaja por pantalla), para que sea la base exacta
            # de rentabilidad. Si por algo no se encuentra, usa el valor enviado.
            valor_osi = int(body.valor or 0)
            if body.oferta_id:
                cur.execute("SELECT valor FROM ofertas WHERE id=%s", (body.oferta_id,))
                _vr = fetchone(cur)
                if _vr and _vr.get("valor") is not None:
                    valor_osi = int(_vr["valor"] or 0)
            cur.execute("""
                INSERT INTO osi (numero_osi, oferta_id, oferta_num, responsable, equipo,
                                 cliente, origen, destino, valor, estado,
                                 fecha_despacho, conductor, placa, observaciones)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                RETURNING id
            """, (numero_osi, body.oferta_id or None, body.oferta_num.strip(),
                  body.lider.strip(), body.equipo.strip(), body.cliente.strip(),
                  body.origen.strip(), body.destino.strip(), valor_osi,
                  'PROGRAMADO', _d(body.fecha_inicio), body.conductor.strip(),
                  body.placa.strip().upper(),
                  # Guardamos el detalle operativo en observaciones (mañana se
                  # detallan columnas propias si hace falta).
                  json.dumps({
                      "nit": body.nit, "solicitante": body.solicitante,
                      "contacto_reporte": body.contacto_reporte_det,
                      "contacto_cargue": body.contacto_cargue_det,
                      "contacto_descargue": body.contacto_descargue_det,
                      "tipo_operacion": body.tipo_operacion,
                      "tipo_carga": body.tipo_carga,
                      "tipo_carga_otro": body.tipo_carga_otro,
                      "equipos_asignados": body.equipos_asignados,
                      "operadores": body.operadores,
                      "auxiliares": body.auxiliares,
                      "especificacion": body.especificacion,
                      "lugar_cargue": body.lugar_cargue,
                      "lugar_descargue": body.lugar_descargue,
                      "hora_servicio": body.hora_servicio,
                      "fecha_final": body.fecha_final,
                      "req_sup": body.req_sup,
                      "req_hse": body.req_hse,
                      "req_otro": body.req_otro,
                      "escoltas": body.escoltas,
                      "dimensiones": {
                          "ancho": body.dim_ancho, "alto": body.dim_alto,
                          "largo": body.dim_largo, "voladizo": body.dim_voladizo,
                          "peso": body.dim_peso,
                      },
                      "via_tipo": body.via_tipo,
                      "via_obstruccion": body.via_obstruccion,
                      "packing_list": body.packing_list,
                      "observaciones": body.observaciones,
                  }, ensure_ascii=False)))
            new_id = fetchone(cur)["id"]
            # Al crear la OSI, marca la alerta de esa oferta como leída
            if body.oferta_id:
                cur.execute("UPDATE notificaciones SET leida = true WHERE oferta_id=%s",
                            (body.oferta_id,))
        return {"ok": True, "id": new_id, "numero_osi": numero_osi}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/ofertas/{oferta_id}/osi-prefill")
def osi_prefill(oferta_id: int, request: Request):
    """Datos de la oferta para pre-llenar el formulario de OSI:
    solicitante = comercial que la realizó, tipo, y equipo ofertado."""
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT realizada, tipo, pdf_data FROM ofertas WHERE id=%s", (oferta_id,))
            row = fetchone(cur)
        if not row:
            return {"solicitante": "", "tipo": "", "equipo": "", "origen": "", "destino": "", "descripcion": ""}
        pdf = row.get("pdf_data")
        if isinstance(pdf, str):
            try: pdf = json.loads(pdf)
            except Exception: pdf = {}
        pdf = pdf or {}
        equipo = "; ".join(e.get("equipo", "") for e in (pdf.get("equipos") or []) if e.get("equipo"))
        return {
            "solicitante": row.get("realizada") or "",
            "tipo": row.get("tipo") or "",
            "equipo": equipo,
            "origen": (pdf.get("origen") or "").strip(),
            "destino": (pdf.get("destino") or "").strip(),
            "descripcion": (pdf.get("descripcion") or "").strip(),
        }
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/osi/proximo-numero")
def osi_proximo_numero(request: Request):
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            return {"numero_osi": _proximo_numero_osi(cur)}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/clientes_contrato")
def get_clientes_contrato():
    """Clientes con CONTRATO (facturación recurrente/mensual, no por oferta suelta).
    El panel de Control de Aprobadas los trata aparte: NO se marcan como
    'sin facturar' porque su facturación va por el contrato, no por cada oferta."""
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT nit, razon_social, nombre_corto FROM clientes_contrato "
                "WHERE activo IS NOT FALSE ORDER BY razon_social")
            return fetchall(cur)
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


@app.get("/api/operaciones/aceptadas")
def operaciones_aceptadas(request: Request):
    """Todas las ofertas ACEPTADAS, para que Operaciones sepa lo que se ha
    notificado/ganado a la fecha. Ordenadas de la más reciente a la más antigua."""
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT num, fecha, cliente, tipo, valor, mes_aceptado,
                       seguimiento, no_factura, valor_facturado
                FROM ofertas
                WHERE UPPER(respuesta) = 'ACEPTADA'
                  AND NOT COALESCE(anulada, false)
                ORDER BY fecha DESC NULLS LAST, CAST(num AS INTEGER) DESC
            """)
            filas = fetchall(cur)
        for f in filas:
            if f.get("fecha"):
                f["fecha"] = str(f["fecha"])
        return filas
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


# ── OSI ───────────────────────────────────────────────────────────────────────
class OSIUpdate(BaseModel):
    responsable:    Optional[str] = None
    equipo:         Optional[str] = None
    estado:         Optional[str] = None
    notas:          Optional[str] = None
    fecha:          Optional[str] = None
    fecha_despacho: Optional[str] = None
    conductor:      Optional[str] = None
    placa:          Optional[str] = None
    observaciones:  Optional[str] = None


@app.get("/api/osi")
def get_osi(request: Request):
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT * FROM osi ORDER BY created_at DESC")
            rows = fetchall(cur)
        # Desglosa el detalle operativo guardado como JSON en 'observaciones'
        # para que la tabla pueda mostrar tipo de operación y especificación.
        for r in rows:
            det = r.get("observaciones")
            if isinstance(det, str):
                try: det = json.loads(det)
                except Exception: det = {}
            det = det or {}
            r["tipo_operacion"] = det.get("tipo_operacion") or ""
            r["tipo_carga"] = det.get("tipo_carga") or ""
            r["equipos_asignados"] = det.get("equipos_asignados") or []
            r["operadores"] = det.get("operadores") or []
            r["auxiliares"] = det.get("auxiliares") or ""
            r["escoltas"] = det.get("escoltas") or ""
            r["req_sup"] = bool(det.get("req_sup"))
            r["req_hse"] = bool(det.get("req_hse"))
            r["req_otro"] = det.get("req_otro") or ""
            r["contacto_reporte"] = det.get("contacto_reporte") or {}
            r["contacto_cargue"] = det.get("contacto_cargue") or {}
            r["contacto_descargue"] = det.get("contacto_descargue") or {}
            r["fecha_final"] = det.get("fecha_final") or ""
            r["especificacion"] = det.get("especificacion") or ""
            r["solicitante"] = det.get("solicitante") or ""
        return rows
    except Exception as e:
        raise HTTPException(500, str(e))


@app.patch("/api/osi/{osi_id}")
def update_osi(osi_id: int, body: OSIUpdate, request: Request):
    try:
        fields = {k: v for k, v in body.dict().items() if v is not None}
        if not fields:
            raise HTTPException(400, "Sin campos")
        set_clause = ", ".join(f"{k} = %s" for k in fields)
        values = list(fields.values()) + [osi_id]
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute(f"UPDATE osi SET {set_clause} WHERE id = %s RETURNING *", values)
            row = fetchone(cur)
            if not row:
                raise HTTPException(404, "OSI no encontrada")
            return row
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


# ── Feature 4: Historial de cambios ──────────────────────────────────────────
@app.get("/api/ofertas/{oferta_id}/historial")
def get_oferta_historial(oferta_id: int):
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT * FROM oferta_historial WHERE oferta_id = %s ORDER BY created_at DESC",
                (oferta_id,)
            )
            return fetchall(cur)
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


# ── Catálogo maestro de clientes (fuente única para generar ofertas) ──────────
@app.get("/api/clientes/catalogo")
def get_clientes_catalogo():
    """Lista oficial de clientes para el selector del generador de ofertas."""
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT id, nombre_corto, razon_social, nit
                FROM clientes
                WHERE activo = TRUE
                ORDER BY nombre_corto
            """)
            return fetchall(cur)
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/api/clientes", status_code=201)
def create_cliente(body: ClienteCreate):
    """Registra un cliente nuevo en el catálogo oficial (nombre corto + datos
    oficiales). Bloquea duplicados sin importar mayúsculas/espacios."""
    nombre = re.sub(r"\s+", " ", (body.nombre_corto or "").strip()).upper()
    if not nombre:
        raise HTTPException(400, "El nombre del cliente es obligatorio")
    razon = ((body.razon_social or "").strip() or None)
    if razon:
        razon = razon.upper()
    nit   = (body.nit or "").strip() or None
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            # ¿Ya existe (ignorando mayúsculas/espacios)?
            cur.execute(
                "SELECT nombre_corto FROM clientes WHERE lower(nombre_corto) = lower(%s)",
                (nombre,),
            )
            ya = fetchone(cur)
            if ya:
                raise HTTPException(409, f'El cliente "{ya["nombre_corto"]}" ya existe en el catálogo')
            cur.execute(
                "INSERT INTO clientes (nombre_corto, razon_social, nit) VALUES (%s,%s,%s) RETURNING id",
                (nombre, razon, nit),
            )
            row = fetchone(cur)
            return {"id": row["id"], "nombre_corto": nombre, "razon_social": razon, "nit": nit}
    except HTTPException:
        raise
    except Exception as e:
        if "unique" in str(e).lower() or "duplicate" in str(e).lower():
            raise HTTPException(409, f'El cliente "{nombre}" ya existe')
        raise HTTPException(500, str(e))


@app.put("/api/clientes/{cid}")
def update_cliente(cid: int, body: ClienteUpdate):
    """Edita un cliente del catálogo. Si cambia el NOMBRE CORTO, el cambio se
    propaga a TODAS las ofertas y tablas para que nada quede desligado
    (auditoría 'cero errores'). NIT y razón social se pueden completar aquí."""
    TABLAS = ["ofertas", "contratos", "facturacion_cat", "notificaciones", "ofertas_2025", "osi"]
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT id, nombre_corto FROM clientes WHERE id = %s", (cid,))
            actual = fetchone(cur)
            if not actual:
                raise HTTPException(404, "Cliente no encontrado")
            viejo = actual["nombre_corto"]

            sets, params = [], []
            nuevo_nombre = None
            if body.nombre_corto is not None:
                nuevo_nombre = re.sub(r"\s+", " ", body.nombre_corto.strip()).upper()
                if not nuevo_nombre:
                    raise HTTPException(400, "El nombre del cliente no puede quedar vacío")
                # ¿Choca con OTRO cliente distinto?
                cur.execute(
                    "SELECT id FROM clientes WHERE lower(nombre_corto) = lower(%s) AND id <> %s",
                    (nuevo_nombre, cid),
                )
                if fetchone(cur):
                    raise HTTPException(409, f'Ya existe otro cliente llamado "{nuevo_nombre}"')
                sets.append("nombre_corto = %s"); params.append(nuevo_nombre)
            if body.razon_social is not None:
                sets.append("razon_social = %s"); params.append((body.razon_social.strip() or None) and body.razon_social.strip().upper())
            if body.nit is not None:
                sets.append("nit = %s"); params.append(body.nit.strip() or None)
            if not sets:
                raise HTTPException(400, "Nada que actualizar")

            params.append(cid)
            cur.execute(f"UPDATE clientes SET {', '.join(sets)} WHERE id = %s", params)

            # Propaga el cambio de nombre a todas las tablas.
            filas_afectadas = 0
            if nuevo_nombre and nuevo_nombre != viejo:
                for t in TABLAS:
                    cur.execute(f"UPDATE {t} SET cliente = %s WHERE cliente = %s", (nuevo_nombre, viejo))
                    filas_afectadas += cur.rowcount
            return {"ok": True, "nombre_anterior": viejo, "nombre_nuevo": nuevo_nombre or viejo,
                    "ofertas_actualizadas": filas_afectadas}
    except HTTPException:
        raise
    except Exception as e:
        if "unique" in str(e).lower() or "duplicate" in str(e).lower():
            raise HTTPException(409, "Ese nombre de cliente ya existe")
        raise HTTPException(500, str(e))


# ── Facturación VULCANO: importador y conciliación ────────────────────────────
# Natalia factura en VULCANO (sistema externo) y descarga un Excel. Aquí lo
# importa a la tabla espejo 'vulcano_facturas', marca las facturas que NO cuentan
# ("excluidas") y la app muestra el total real conciliado contra contabilidad.

# Encabezados esperados del Excel de Vulcano (posición 1-based). Se detectan por
# nombre de columna (flexible), con estas posiciones como respaldo.
_VULCANO_COLS = {
    "factura": 1, "fecha": 2, "anio": 3, "mes": 4, "estado": 6,
    "nit": 7, "cliente": 8, "subtotal": 9, "total": 14,
    "valor_pagado": 17, "saldo": 18,
}
# Facturas que contabilidad marcó para NO incluir en el total (las "amarillas").
# Se pre-marcan como excluidas al importar para dejar el total real de una vez.
_VULCANO_EXCLUIDAS_DEFAULT = {
    "BLCE3761", "BLCE3768", "BLCE3773", "BLCE3823", "BLCE3824",
    "BLCE3825", "BLCE3882", "BLCE3883", "BLCE3884",
    "BLCE3578",  # TRANSELCA (enero) — Natalia pidió excluirla siempre
}


def _vul_num(x):
    """Convierte a entero seguro (None/strings/formulas -> 0)."""
    if isinstance(x, bool):
        return 0
    if isinstance(x, (int, float)):
        return int(round(x))
    return 0


def _vul_str(x):
    if x is None:
        return None
    s = str(x).strip()
    return s or None


def _vulcano_fix_cliente(cli):
    """Correcciones fijas de nombre de cliente al importar (typos de la fuente Vulcano).
    Así, aunque el Excel traiga el nombre mal, en la app entra corregido."""
    if cli and "ROTRASHEVI" in cli.upper():
        return "ROTRASVEHI S.A.S."
    return cli


def _app_total_facturado(cur):
    """Total facturado en la app (misma lógica que /api/facturacion/resumen)."""
    cur.execute("SELECT COALESCE(SUM(valor_facturado),0) FROM ofertas "
                "WHERE UPPER(respuesta)='ACEPTADA'")
    fact_2026 = cur.fetchone()[0] or 0
    cur.execute("SELECT COALESCE(SUM(valor_facturado),0) FROM ofertas_2025")
    f2025_num = cur.fetchone()[0] or 0
    cur.execute("SELECT categoria, COALESCE(SUM(valor_facturado),0) "
                "FROM facturacion_cat GROUP BY categoria")
    cat = {r[0]: (r[1] or 0) for r in cur.fetchall()}
    contratos = int(cat.get("CONTRATO", 0))
    ano_pasado = int(cat.get("2025_ANO_PASADO", 0))
    otros = int(cat.get("OTROS", 0))
    return int(fact_2026) + int(f2025_num) + ano_pasado + contratos + otros


def _vulcano_aplicar_a_ofertas(cur):
    """Auto-facturado: cruza vulcano_facturas (NO excluidas) contra las ofertas por su
    número (columna 'Oferta No.') y llena valor_facturado + no_factura + fecha_facturacion.
    Si el facturado cubre el valor de la oferta -> la CIERRA (seguimiento='Facturada').
    Parcial (facturado < valor) -> actualiza el facturado pero NO cierra (queda 🟡 con saldo).
    Objetivo (pedido de Natalia): al cargar el Excel todo se llena y cierra solo, sin trabajo manual.
    Usa MAX(actual, suma Vulcano) para no bajar un facturado ya registrado ni romper la conciliación."""
    cur.execute("""
        SELECT oferta_ref, factura, subtotal, fecha
        FROM vulcano_facturas
        WHERE NOT excluida AND oferta_ref IS NOT NULL AND TRIM(oferta_ref) <> ''
    """)
    por_of = {}   # num26 -> {"suma": int, "facturas": set, "fecha": date|None}
    for oref, factura, subtotal, fecha in cur.fetchall():
        nums = set("26" + m.zfill(4) for m in re.findall(r"26-?(\d{3,4})", str(oref)))
        # Solo auto-atribuir facturas de UNA sola oferta. Las multi-oferta (una factura
        # que cubre varias ofertas) NO se reparten aquí para no inflar el facturado:
        # esas van por el detalle de la hoja "Ofertas Aprobadas" (que ya trae el split).
        if len(nums) != 1:
            continue
        num = next(iter(nums))
        d = por_of.setdefault(num, {"suma": 0, "facturas": set(), "fecha": None})
        d["suma"] += int(subtotal or 0)
        if factura:
            d["facturas"].add(str(factura).strip())
        if fecha and (d["fecha"] is None or fecha > d["fecha"]):
            d["fecha"] = fecha

    n_fact = n_cierre = 0
    for num, d in por_of.items():
        cur.execute(
            "SELECT valor AS valor, COALESCE(valor_facturado,0) AS vf, "
            "UPPER(COALESCE(seguimiento,'')) AS seg "
            "FROM ofertas WHERE num=%s AND UPPER(respuesta)='ACEPTADA'", (num,))
        row = fetchone(cur)
        if not row:
            continue
        valor = int(row["valor"] or 0)
        vf_prev = int(row["vf"] or 0)
        seg = row["seg"] or ""
        facturado = max(vf_prev, int(d["suma"]))
        nofact = ", ".join(sorted(d["facturas"]))[:250] or None
        completa = valor > 0 and facturado >= valor - 1000
        cancelada = "CANCEL" in seg
        if completa and not cancelada:
            cur.execute(
                "UPDATE ofertas SET valor_facturado=%s, no_factura=%s, "
                "fecha_facturacion=%s, seguimiento='Facturada' WHERE num=%s",
                (facturado, nofact, d["fecha"], num))
            n_cierre += 1
        else:
            # Parcial (o cancelada): actualiza el facturado, NO toca el seguimiento.
            cur.execute(
                "UPDATE ofertas SET valor_facturado=%s, no_factura=%s, "
                "fecha_facturacion=%s WHERE num=%s",
                (facturado, nofact, d["fecha"], num))
        n_fact += 1
    return {"ofertas_afectadas": n_fact, "ofertas_cerradas": n_cierre}


@app.post("/api/vulcano/importar")
async def vulcano_importar(archivo: UploadFile = File(...)):
    """Sube el Excel descargado de VULCANO y lo carga a la tabla espejo.
    - Enlaza cada factura por número (BLCE####).
    - Pre-marca como excluidas las facturas que no cuentan al total.
    - Conserva las exclusiones que Natalia haya hecho manualmente (re-importar
      no las pierde)."""
    if not OPENPYXL_OK:
        raise HTTPException(500, "openpyxl no está instalado en el servidor")
    try:
        from openpyxl import load_workbook
    except Exception:
        raise HTTPException(500, "No se pudo cargar openpyxl")

    contenido = await archivo.read()
    try:
        wb = load_workbook(BytesIO(contenido), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"No se pudo leer el Excel: {e}")

    # Buscar la hoja de facturación (por nombre) o usar la primera.
    ws = None
    for nombre in wb.sheetnames:
        if "factur" in nombre.lower():
            ws = wb[nombre]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    # Mapear columnas por encabezado (fila 1); si no, usar posiciones por defecto.
    hdr = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v:
            hdr[str(v).strip().lower()] = c

    def col(key, alts):
        for a in alts:
            if a in hdr:
                return hdr[a]
        return _VULCANO_COLS[key]

    idx = {
        "factura": col("factura", ["factura"]),
        "fecha": col("fecha", ["fecha"]),
        "anio": col("anio", ["año", "ano", "anio"]),
        "mes": col("mes", ["mes_", "mes"]),
        "estado": col("estado", ["estado"]),
        "nit": col("nit", ["id cliente", "nit"]),
        "cliente": col("cliente", ["cliente"]),
        "subtotal": col("subtotal", ["subtotal"]),
        "total": col("total", ["total"]),
        "valor_pagado": col("valor_pagado", ["valor_pagado", "valor pagado"]),
        "saldo": col("saldo", ["saldo"]),
    }

    # "Oferta No." SOLO por encabezado (los dos formatos de Excel tienen columnas en
    # posiciones distintas). Si no está el encabezado, no se captura (evita tomar otra
    # columna por posición). Sirve para el auto-facturado por número de oferta.
    oferta_col = None
    for a in ("oferta no.", "oferta no", "oferta_no", "oferta"):
        if a in hdr:
            oferta_col = hdr[a]
            break

    filas = []
    for r in range(2, ws.max_row + 1):
        factura = _vul_str(ws.cell(r, idx["factura"]).value)
        if not factura:
            continue  # fila vacía
        # Saltar filas de totales/subtotales del Excel (p. ej. "N. registros: 315,0").
        # Los números de factura reales no llevan espacios ni ':'.
        fl = factura.lower()
        if " " in factura or ":" in factura or "registro" in fl or fl.startswith("total"):
            continue
        fecha_val = ws.cell(r, idx["fecha"]).value
        fecha = None
        if isinstance(fecha_val, datetime):
            fecha = fecha_val.date()
        elif isinstance(fecha_val, date):
            fecha = fecha_val
        filas.append({
            "factura": factura,
            "fecha": fecha,
            "mes": _vul_str(ws.cell(r, idx["mes"]).value),
            "anio": _vul_str(ws.cell(r, idx["anio"]).value),
            "estado": _vul_str(ws.cell(r, idx["estado"]).value),
            "nit": _vul_str(ws.cell(r, idx["nit"]).value),
            "cliente": _vulcano_fix_cliente(_vul_str(ws.cell(r, idx["cliente"]).value)),
            "subtotal": _vul_num(ws.cell(r, idx["subtotal"]).value),
            "total": _vul_num(ws.cell(r, idx["total"]).value),
            "valor_pagado": _vul_num(ws.cell(r, idx["valor_pagado"]).value),
            "saldo": _vul_num(ws.cell(r, idx["saldo"]).value),
            "oferta_ref": (_vul_str(ws.cell(r, oferta_col).value) if oferta_col else None),
        })

    if not filas:
        raise HTTPException(400, "El Excel no tiene facturas legibles (revisa la hoja/columnas)")

    nuevas = actualizadas = 0
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            for f in filas:
                # ¿Ya existe? Para conservar la exclusión manual.
                cur.execute("SELECT excluida FROM vulcano_facturas WHERE factura=%s",
                            (f["factura"],))
                prev = fetchone(cur)
                if prev is None:
                    # Nueva: excluir si está en la lista por defecto.
                    excluida = f["factura"] in _VULCANO_EXCLUIDAS_DEFAULT
                    cur.execute("""
                        INSERT INTO vulcano_facturas
                            (factura, fecha, mes, anio, estado, nit, cliente,
                             subtotal, total, valor_pagado, saldo, excluida, oferta_ref)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """, (f["factura"], f["fecha"], f["mes"], f["anio"], f["estado"],
                          f["nit"], f["cliente"], f["subtotal"], f["total"],
                          f["valor_pagado"], f["saldo"], excluida, f.get("oferta_ref")))
                    nuevas += 1
                else:
                    # Existe: actualizar datos, CONSERVAR excluida manual.
                    cur.execute("""
                        UPDATE vulcano_facturas SET
                            fecha=%s, mes=%s, anio=%s, estado=%s, nit=%s, cliente=%s,
                            subtotal=%s, total=%s, valor_pagado=%s, saldo=%s,
                            oferta_ref=COALESCE(%s, oferta_ref),
                            importado_at=now()
                        WHERE factura=%s
                    """, (f["fecha"], f["mes"], f["anio"], f["estado"], f["nit"],
                          f["cliente"], f["subtotal"], f["total"], f["valor_pagado"],
                          f["saldo"], f.get("oferta_ref"), f["factura"]))
                    actualizadas += 1
            # Auto-facturado: cruza las facturas con sus ofertas y cierra las completas.
            aplicado = _vulcano_aplicar_a_ofertas(cur)
            resumen = _vulcano_calc_resumen(cur)
    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))

    return {"ok": True, "leidas": len(filas), "nuevas": nuevas,
            "actualizadas": actualizadas, "resumen": resumen,
            "auto_facturado": aplicado}


def _vulcano_calc_resumen(cur):
    """Calcula totales de Vulcano y la conciliación contra la app."""
    cur.execute("""
        SELECT
            COUNT(*),
            COALESCE(SUM(subtotal),0),
            COUNT(*) FILTER (WHERE excluida),
            COALESCE(SUM(subtotal) FILTER (WHERE excluida),0),
            COALESCE(SUM(subtotal) FILTER (WHERE NOT excluida),0)
        FROM vulcano_facturas
    """)
    n, bruto, n_excl, monto_excl, total_real = cur.fetchone()
    app_total = _app_total_facturado(cur)
    diff = int(app_total) - int(total_real)
    return {
        "n_facturas": int(n or 0),
        "subtotal_bruto": int(bruto or 0),
        "n_excluidas": int(n_excl or 0),
        "monto_excluido": int(monto_excl or 0),
        "total_real_vulcano": int(total_real or 0),
        "total_app": int(app_total),
        "diferencia": diff,
    }


@app.get("/api/vulcano/resumen")
def vulcano_resumen():
    """Totales de Vulcano vs. total de la app (conciliación)."""
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            return _vulcano_calc_resumen(cur)
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


@app.get("/api/vulcano/facturas")
def vulcano_facturas(excluidas: Optional[str] = Query(None),
                     buscar: Optional[str] = Query(None)):
    """Lista las facturas importadas de Vulcano.
    excluidas: 'si' | 'no' (opcional). buscar: por factura o cliente."""
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            where, params = [], []
            if excluidas == "si":
                where.append("excluida = TRUE")
            elif excluidas == "no":
                where.append("excluida = FALSE")
            if buscar:
                where.append("(factura ILIKE %s OR cliente ILIKE %s)")
                params += [f"%{buscar}%", f"%{buscar}%"]
            sql = "SELECT * FROM vulcano_facturas"
            if where:
                sql += " WHERE " + " AND ".join(where)
            sql += " ORDER BY factura"
            cur.execute(sql, params)
            return fetchall(cur)
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


@app.post("/api/vulcano/factura/{factura}/excluir")
def vulcano_excluir(factura: str, body: VulcanoExcluir):
    """Marca/desmarca una factura como excluida del total real."""
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("UPDATE vulcano_facturas SET excluida=%s WHERE factura=%s",
                        (body.excluida, factura))
            if cur.rowcount == 0:
                raise HTTPException(404, "Factura no encontrada")
            resumen = _vulcano_calc_resumen(cur)
            return {"ok": True, "factura": factura, "excluida": body.excluida,
                    "resumen": resumen}
    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


# ── Presupuesto / META de facturación ────────────────────────────────────────
_MES_ORDER = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
              "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]


@app.get("/api/presupuesto")
def get_presupuesto(anio: Optional[int] = Query(None)):
    """Presupuesto (meta) mensual y total del año."""
    anio_val = anio or datetime.now().year
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT mes, monto FROM presupuesto WHERE anio=%s", (anio_val,))
            data = {r["mes"]: int(r["monto"]) for r in fetchall(cur)}
        meses = [{"mes": m, "monto": data.get(m, 0)} for m in _MES_ORDER]
        return {"anio": anio_val, "meses": meses,
                "total": sum(x["monto"] for x in meses)}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


@app.put("/api/presupuesto")
def set_presupuesto(body: PresupuestoItem):
    """Crea/edita la meta de un mes."""
    mes = (body.mes or "").strip().upper()
    if mes not in _MES_ORDER:
        raise HTTPException(400, "Mes inválido")
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("""
                INSERT INTO presupuesto (anio, mes, monto) VALUES (%s,%s,%s)
                ON CONFLICT (anio, mes) DO UPDATE SET monto = EXCLUDED.monto
            """, (body.anio, mes, int(body.monto)))
            return {"ok": True, "anio": body.anio, "mes": mes, "monto": int(body.monto)}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


# ── Feature 2: Mini CRM Clientes ──────────────────────────────────────────────
@app.get("/api/clientes/stats")
def get_clientes_stats():
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT
                    cliente,
                    COUNT(*) AS total_ofertas,
                    COUNT(*) FILTER (WHERE UPPER(respuesta) = 'ACEPTADA') AS aceptadas,
                    COUNT(*) FILTER (WHERE UPPER(respuesta) = 'RECHAZADA') AS rechazadas,
                    COUNT(*) FILTER (WHERE UPPER(respuesta) = 'EN SEGUIMIENTO') AS en_seguimiento,
                    ROUND(
                        COUNT(*) FILTER (WHERE UPPER(respuesta) = 'ACEPTADA') * 100.0
                        / NULLIF(COUNT(*), 0), 1
                    ) AS tasa_cierre,
                    COALESCE(SUM(valor), 0) AS valor_total_ofertado,
                    COALESCE(SUM(valor_facturado), 0) AS valor_total_facturado,
                    MAX(fecha) AS ultima_oferta
                FROM ofertas
                WHERE cliente IS NOT NULL AND cliente != ''
                  AND NOT COALESCE(anulada, false)
                  AND NOT COALESCE(es_prueba, false)
                GROUP BY cliente
                ORDER BY total_ofertas DESC
            """)
            return fetchall(cur)
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


# ── Feature 1: Exportar a Excel ───────────────────────────────────────────────
def _make_excel_ofertas(rows: list) -> bytes:
    if not OPENPYXL_OK:
        raise RuntimeError("openpyxl no está instalado. Ejecuta: pip install openpyxl")

    NAVY  = "1B2A4A"
    ALT   = "EEF2FF"

    def _v(val):
        """Convierte cualquier valor a algo seguro para openpyxl."""
        if val is None:
            return ""
        if isinstance(val, (dict, list)):
            return ""
        return val

    wb = Workbook()

    # ── Hoja 1: Ofertas ───────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Ofertas"
    ws.freeze_panes = "A2"

    hdr_fill = PatternFill("solid", fgColor=NAVY)
    alt_fill = PatternFill("solid", fgColor=ALT)
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    center   = Alignment(horizontal="center", vertical="center", wrap_text=False)

    headers    = ["No. Oferta","Mes","Fecha","Cliente","Realizada por",
                  "Unidad","Tipo","Valor COP","Respuesta","Estado",
                  "Seguimiento","Mes Aceptado","Facturación","Valor Facturado","No. Factura"]
    col_widths = [12, 12, 13, 30, 22, 14, 22, 18, 16, 14, 16, 14, 14, 18, 16]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = center
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[1].height = 22

    for ri, o in enumerate(rows, 2):
        vals = [
            _v(o.get("num")),      _v(o.get("mes")),     _v(o.get("fecha")),
            _v(o.get("cliente")),  _v(o.get("realizada")),
            _v(o.get("unidad")),   _v(o.get("tipo")),
            o.get("valor") or 0,
            _v(o.get("respuesta")), _v(o.get("estado")),
            _v(o.get("seguimiento")), _v(o.get("mes_aceptado")),
            _v(o.get("facturacion")),
            o.get("valor_facturado") or 0,
            _v(o.get("no_factura")),
        ]
        fill = alt_fill if ri % 2 == 0 else None
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.alignment = Alignment(vertical="center")
            if fill:
                c.fill = fill

    # ── Hoja 2: Resumen ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumen")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 20

    def _hdr2(row, col, text):
        c = ws2.cell(row=row, column=col, value=text)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = center

    def _title2(row, text):
        c = ws2.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, size=11, color=NAVY)

    # Bloque 1 — por respuesta
    _title2(1, "RESUMEN POR RESPUESTA")
    _hdr2(2, 1, "Respuesta")
    _hdr2(2, 2, "Cantidad")

    resp_map: dict = {}
    for o in rows:
        k = (o.get("respuesta") or "Sin respuesta").upper()
        resp_map[k] = resp_map.get(k, 0) + 1
    for ri2, (k, v) in enumerate(sorted(resp_map.items()), 3):
        ws2.cell(row=ri2, column=1, value=k)
        ws2.cell(row=ri2, column=2, value=v)

    # Bloque 2 — por mes
    off = 3 + len(resp_map) + 2
    _title2(off, "RESUMEN POR MES")
    _hdr2(off + 1, 1, "Mes")
    _hdr2(off + 1, 2, "Cantidad")
    _hdr2(off + 1, 3, "Valor Total COP")

    mes_order = ["ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO",
                 "JULIO","AGOSTO","SEPTIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"]
    mes_cnt: dict = {}
    mes_val: dict = {}
    for o in rows:
        m = (o.get("mes") or "").upper()
        if m:
            mes_cnt[m] = mes_cnt.get(m, 0) + 1
            mes_val[m] = mes_val.get(m, 0) + (o.get("valor") or 0)
    for ri2, m in enumerate([m for m in mes_order if m in mes_cnt], off + 2):
        ws2.cell(row=ri2, column=1, value=m)
        ws2.cell(row=ri2, column=2, value=mes_cnt[m])
        ws2.cell(row=ri2, column=3, value=mes_val[m])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@app.get("/api/exportar/ofertas")
def exportar_ofertas_excel(
    mes:    Optional[str] = Query(None),
    anio:   Optional[int] = Query(None),
    estado: Optional[str] = Query(None),
):
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT * FROM ofertas ORDER BY CAST(num AS INTEGER) DESC")
            rows = fetchall(cur)

        # Filtrar
        if mes:
            rows = [r for r in rows if (r.get("mes") or "").upper() == mes.upper()]
        if anio:
            rows = [r for r in rows if r.get("fecha") and str(r["fecha"])[:4] == str(anio)]
        if estado:
            rows = [r for r in rows if (r.get("respuesta") or "").upper() == estado.upper()
                    or (r.get("estado") or "").upper() == estado.upper()]

        excel_bytes = _make_excel_ofertas(rows)
        fname = "Ofertas_BOOM"
        if mes:
            fname += f"_{mes}"
        if anio:
            fname += f"_{anio}"
        fname += ".xlsx"
        return StreamingResponse(
            BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


# ── Feature 5: Informe Mensual de Facturación ─────────────────────────────────
@app.get("/api/reportes/facturacion")
def reporte_facturacion(anio: Optional[int] = Query(None)):
    try:
        anio_val = anio or datetime.now().year
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT
                    mes_aceptado AS mes,
                    COUNT(*) AS total_aceptadas,
                    COUNT(*) FILTER (WHERE no_factura IS NOT NULL AND no_factura != '') AS facturadas,
                    COUNT(*) FILTER (WHERE no_factura IS NULL OR no_factura = '') AS sin_facturar,
                    COALESCE(SUM(valor), 0) AS valor_ofertado,
                    COALESCE(SUM(valor_facturado), 0) AS valor_facturado,
                    ROUND(
                        COALESCE(SUM(valor_facturado), 0) * 100.0
                        / NULLIF(COALESCE(SUM(valor), 0), 0), 1
                    ) AS tasa_facturacion
                FROM ofertas
                WHERE UPPER(respuesta) = 'ACEPTADA'
                  AND mes_aceptado IS NOT NULL AND mes_aceptado != ''
                  AND (
                      fecha_facturacion IS NULL
                      OR EXTRACT(YEAR FROM fecha_facturacion) = %s
                      OR fecha IS NULL
                      OR EXTRACT(YEAR FROM fecha) = %s
                  )
                GROUP BY mes_aceptado
                ORDER BY
                    CASE mes_aceptado
                        WHEN 'ENERO'      THEN 1  WHEN 'FEBRERO'    THEN 2
                        WHEN 'MARZO'      THEN 3  WHEN 'ABRIL'      THEN 4
                        WHEN 'MAYO'       THEN 5  WHEN 'JUNIO'      THEN 6
                        WHEN 'JULIO'      THEN 7  WHEN 'AGOSTO'     THEN 8
                        WHEN 'SEPTIEMBRE' THEN 9  WHEN 'OCTUBRE'    THEN 10
                        WHEN 'NOVIEMBRE'  THEN 11 WHEN 'DICIEMBRE'   THEN 12
                        ELSE 99
                    END
            """, (anio_val, anio_val))
            filas = fetchall(cur)
            # Presupuesto (meta) del año para calcular cumplimiento por mes.
            cur.execute("SELECT mes, monto FROM presupuesto WHERE anio=%s", (anio_val,))
            presu = {r["mes"]: int(r["monto"]) for r in fetchall(cur)}
        for f in filas:
            m = (f.get("mes") or "").upper()
            meta = presu.get(m, 0)
            fact = float(f.get("valor_facturado") or 0)
            f["presupuesto"] = meta
            f["cumplimiento"] = round(fact * 100.0 / meta, 1) if meta else None
        return filas
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))



@app.get("/api/reportes/top-clientes")
def reporte_top_clientes(limit: int = Query(10)):
    """Top de clientes por facturación REAL (fuente: VULCANO conciliado).
    Usa el subtotal de las facturas NO excluidas, agrupado por cliente.
    Devuelve el total general para poder calcular el % de cada cliente."""
    try:
        limit = max(1, min(limit, 50))
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT cliente,
                       COUNT(*) AS n_facturas,
                       COALESCE(SUM(subtotal), 0) AS valor
                FROM vulcano_facturas
                WHERE excluida = false
                  AND cliente IS NOT NULL AND cliente <> ''
                GROUP BY cliente
                HAVING COALESCE(SUM(subtotal), 0) > 0
                ORDER BY valor DESC
                LIMIT %s
            """, (limit,))
            filas = fetchall(cur)
            cur.execute("""
                SELECT COALESCE(SUM(subtotal), 0) AS total
                FROM vulcano_facturas
                WHERE excluida = false
            """)
            total = int(fetchone(cur)["total"] or 0)
        clientes = []
        for f in filas:
            valor = int(f["valor"] or 0)
            clientes.append({
                "cliente": f["cliente"],
                "n_facturas": int(f["n_facturas"] or 0),
                "valor": valor,
                "porcentaje": round(valor * 100.0 / total, 1) if total else 0.0,
            })
        return {"total": total, "clientes": clientes}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


@app.get("/api/reportes/tablero")
def reporte_tablero():
    """Resumen ejecutivo para la pestaña Tablero:
    - Efectividad (ofertas aceptadas / total)
    - Valor ganado (facturado real conciliado con Vulcano)
    - Seguimiento pendiente (ofertas RECIBIDO con +15 días sin respuesta)
    - Ranking por ejecutivo (quién generó cada oferta)."""
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT
                    COUNT(*) AS total,
                    COUNT(*) FILTER (WHERE UPPER(respuesta) = 'ACEPTADA') AS aceptadas,
                    COUNT(*) FILTER (
                        WHERE UPPER(respuesta) = 'RECIBIDO'
                          AND fecha IS NOT NULL
                          AND fecha <= CURRENT_DATE - INTERVAL '15 days'
                    ) AS seguimiento
                FROM ofertas
            """)
            r = fetchone(cur)
            total = int(r["total"] or 0)
            aceptadas = int(r["aceptadas"] or 0)
            seguimiento = int(r["seguimiento"] or 0)

            # Valor ganado = facturación real conciliada (Vulcano). Si aún no se
            # ha importado Vulcano, cae al total facturado interno de la app.
            cur.execute("""
                SELECT COALESCE(SUM(subtotal), 0) AS t
                FROM vulcano_facturas WHERE excluida = false
            """)
            valor_ganado = int(fetchone(cur)["t"] or 0)
            if not valor_ganado:
                valor_ganado = int(_app_total_facturado(cur) or 0)

            cur.execute("""
                SELECT COALESCE(NULLIF(TRIM(realizada), ''), '(sin asignar)') AS ejecutivo,
                       COUNT(*) AS n_ofertas,
                       COUNT(*) FILTER (WHERE UPPER(respuesta) = 'ACEPTADA') AS n_aceptadas,
                       COALESCE(SUM(valor_facturado), 0) AS valor_facturado
                FROM ofertas
                GROUP BY 1
                ORDER BY n_ofertas DESC
            """)
            ejec = fetchall(cur)

        por_ejecutivo = [{
            "ejecutivo": e["ejecutivo"],
            "n_ofertas": int(e["n_ofertas"] or 0),
            "n_aceptadas": int(e["n_aceptadas"] or 0),
            "valor_facturado": int(e["valor_facturado"] or 0),
        } for e in ejec]

        return {
            "total_ofertas": total,
            "aceptadas": aceptadas,
            "efectividad": round(aceptadas * 100.0 / total, 1) if total else 0.0,
            "valor_ganado": valor_ganado,
            "seguimiento_pendiente": seguimiento,
            "por_ejecutivo": por_ejecutivo,
        }
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


if __name__ == '__main__':
    import uvicorn
    port = int(os.environ.get('PORT', 8000))
    uvicorn.run(app, host='0.0.0.0', port=port)
